"""
Test suite for export features with Corrected Light-Induced HRV (Detrended) data
Features tested:
- PDF export: Light-Induced HRV Analysis page with original + detrended tables
- PDF export: Detrending Visualization page with 5 stim panels (A, B, C)
- XLSX export: Light Stimulus HRV sheet with both original and detrended sections
"""

import pytest
import requests
import os
import io
import json
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestExportWithDetrendedData:
    """Tests for export endpoints with detrended light HRV data"""
    
    @pytest.fixture(scope="class")
    def session_data(self):
        """Upload ABF file and get session data"""
        with open('/app/4D030006.abf', 'rb') as f:
            files = {'files': ('4D030006.abf', f, 'application/octet-stream')}
            response = requests.post(f"{BASE_URL}/api/upload", files=files, timeout=60)
        
        assert response.status_code == 200, f"Upload failed: {response.text}"
        data = response.json()
        assert 'session_id' in data
        return data
    
    @pytest.fixture(scope="class")
    def metrics_data(self, session_data):
        """Get beat metrics from uploaded file"""
        file_data = session_data['files'][0]
        beat_times_sec = [b['time_sec'] for b in file_data['beats']]
        
        response = requests.post(f"{BASE_URL}/api/compute-metrics", json={
            "beat_times_sec": beat_times_sec,
            "filter_lower_pct": 50,
            "filter_upper_pct": 200,
        }, timeout=60)
        
        assert response.status_code == 200
        return response.json()
    
    @pytest.fixture(scope="class")
    def pulses_data(self, metrics_data):
        """Detect light pulses"""
        response = requests.post(f"{BASE_URL}/api/light-detect", json={
            "start_time_sec": 180,
            "pulse_duration_sec": 20,
            "interval_sec": "decreasing",
            "n_pulses": 5,
            "auto_detect": True,
            "beat_times_min": metrics_data['filtered_beat_times_min'],
            "bf_filtered": metrics_data['filtered_bf_bpm'],
            "search_range_sec": 20,
        }, timeout=60)
        
        assert response.status_code == 200
        return response.json()
    
    @pytest.fixture(scope="class")
    def light_hrv_data(self, metrics_data, pulses_data):
        """Get original Light HRV data"""
        response = requests.post(f"{BASE_URL}/api/light-hrv", json={
            "beat_times_min": metrics_data['filtered_beat_times_min'],
            "bf_filtered": metrics_data['filtered_bf_bpm'],
            "pulses": pulses_data['pulses'],
        }, timeout=60)
        
        assert response.status_code == 200
        return response.json()
    
    @pytest.fixture(scope="class")
    def light_hrv_detrended_data(self, metrics_data, pulses_data):
        """Get Corrected (Detrended) Light HRV data"""
        response = requests.post(f"{BASE_URL}/api/light-hrv-detrended", json={
            "beat_times_min": metrics_data['filtered_beat_times_min'],
            "bf_filtered": metrics_data['filtered_bf_bpm'],
            "pulses": pulses_data['pulses'],
            "loess_frac": 0.25,
        }, timeout=60)
        
        assert response.status_code == 200
        return response.json()
    
    @pytest.fixture(scope="class")
    def full_export_payload(self, metrics_data, pulses_data, light_hrv_data, light_hrv_detrended_data):
        """Build complete export payload with all data"""
        per_beat = metrics_data['beat_times_min'][:-1]
        per_beat_data = [{
            'time_min': t,
            'bf_bpm': metrics_data['beat_freq_bpm'][i],
            'nn_ms': metrics_data['nn_intervals_ms'][i],
            'status': 'kept' if metrics_data['artifact_mask'][i] else 'filtered',
        } for i, t in enumerate(per_beat)]
        
        return {
            "per_beat_data": per_beat_data,
            "light_metrics": light_hrv_data.get('per_pulse', []),
            "light_metrics_detrended": light_hrv_detrended_data,
            "light_pulses": pulses_data['pulses'],
            "filename": "test_export_detrended",
            "recording_name": "Test Recording",
        }

    # ==== PDF Export Tests ====
    
    def test_pdf_export_with_detrended_returns_200(self, full_export_payload):
        """Test PDF export with detrended data returns 200"""
        response = requests.post(
            f"{BASE_URL}/api/export/pdf", 
            json=full_export_payload,
            timeout=120
        )
        
        assert response.status_code == 200, f"PDF export failed: {response.text}"
        assert response.headers.get('content-type') == 'application/pdf' or 'pdf' in response.headers.get('content-type', '')
    
    def test_pdf_export_creates_valid_pdf(self, full_export_payload):
        """Test PDF export creates a valid PDF file"""
        response = requests.post(
            f"{BASE_URL}/api/export/pdf", 
            json=full_export_payload,
            timeout=120
        )
        
        assert response.status_code == 200
        
        # Check PDF magic bytes (PDF files start with %PDF)
        content = response.content
        assert content[:4] == b'%PDF', "Response is not a valid PDF file"
        assert len(content) > 1000, "PDF file seems too small"
    
    def test_pdf_export_without_detrended_works(self, light_hrv_data, pulses_data):
        """Test PDF export works without detrended data (backward compatibility)"""
        payload = {
            "light_metrics": light_hrv_data.get('per_pulse', []),
            "light_pulses": pulses_data['pulses'],
            "filename": "test_no_detrend",
        }
        
        response = requests.post(
            f"{BASE_URL}/api/export/pdf", 
            json=payload,
            timeout=120
        )
        
        assert response.status_code == 200

    # ==== XLSX Export Tests ====
    
    def test_xlsx_export_with_detrended_returns_200(self, full_export_payload):
        """Test XLSX export with detrended data returns 200"""
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=full_export_payload,
            timeout=120
        )
        
        assert response.status_code == 200, f"XLSX export failed: {response.text}"
    
    def test_xlsx_export_creates_valid_xlsx(self, full_export_payload):
        """Test XLSX export creates a valid Excel file"""
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=full_export_payload,
            timeout=120
        )
        
        assert response.status_code == 200
        
        # Check Excel magic bytes (xlsx files start with PK)
        content = response.content
        assert content[:2] == b'PK', "Response is not a valid XLSX file"
    
    def test_xlsx_has_light_stim_hrv_sheet(self, full_export_payload):
        """Test XLSX contains Light Stimulus HRV sheet"""
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=full_export_payload,
            timeout=120
        )
        
        assert response.status_code == 200
        
        # Load workbook and check sheets
        wb = load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert 'Light Stimulus HRV' in sheet_names, f"Missing 'Light Stimulus HRV' sheet. Found: {sheet_names}"
    
    def test_xlsx_light_stim_hrv_has_original_data(self, full_export_payload):
        """Test Light Stimulus HRV sheet has original HRV data"""
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=full_export_payload,
            timeout=120
        )
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb['Light Stimulus HRV']
        
        # Check header row has original HRV columns
        header_values = [cell.value for cell in ws[1] if cell.value]
        
        # Should have Stim # and ln(RMSSD) columns
        assert any('Stim' in str(v) for v in header_values), "Missing Stim # header"
        assert any('RMSSD' in str(v) for v in header_values), "Missing RMSSD header"
    
    def test_xlsx_light_stim_hrv_has_detrended_section(self, full_export_payload):
        """Test Light Stimulus HRV sheet has Corrected HRV (Detrended) section"""
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=full_export_payload,
            timeout=120
        )
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb['Light Stimulus HRV']
        
        # Search for "Corrected" or "Detrended" text in the sheet
        found_detrended = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and ('Detrended' in str(cell.value) or 'Corrected' in str(cell.value)):
                    found_detrended = True
                    break
            if found_detrended:
                break
        
        assert found_detrended, "Missing 'Corrected Light-Induced HRV (Detrended)' section in Light Stimulus HRV sheet"
    
    def test_xlsx_detrended_section_has_correct_columns(self, full_export_payload):
        """Test detrended section has correct column headers"""
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=full_export_payload,
            timeout=120
        )
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb['Light Stimulus HRV']
        
        # Find the detrended header row
        detrended_header_row = None
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                if cell.value and 'ln(RMSSD' in str(cell.value) and '_det' in str(cell.value):
                    detrended_header_row = row_idx
                    break
            if detrended_header_row:
                break
        
        assert detrended_header_row is not None, "Could not find detrended header row"
        
        # Check header values
        header_values = [cell.value for cell in ws[detrended_header_row] if cell.value]
        header_str = ' '.join(str(v) for v in header_values)
        
        assert 'RMSSD' in header_str, "Detrended section missing RMSSD column"
        assert 'SDNN' in header_str, "Detrended section missing SDNN column"
        assert 'pNN50' in header_str, "Detrended section missing pNN50 column"
    
    def test_xlsx_detrended_has_5_stim_rows_plus_median(self, full_export_payload):
        """Test detrended section has data rows for 5 stims plus median"""
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=full_export_payload,
            timeout=120
        )
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb['Light Stimulus HRV']
        
        # Count total rows in sheet
        max_row = ws.max_row
        
        # Should have at least:
        # - Original header + 5 stims + median (7 rows)
        # - Detrended header + 5 stims + median (7 rows)
        # - Title rows and spacing
        assert max_row >= 14, f"Light Stimulus HRV sheet has only {max_row} rows, expected at least 14 (original + detrended)"
    
    def test_xlsx_without_detrended_works(self, light_hrv_data, pulses_data):
        """Test XLSX export works without detrended data (backward compatibility)"""
        payload = {
            "light_metrics": light_hrv_data.get('per_pulse', []),
            "light_pulses": pulses_data['pulses'],
            "filename": "test_no_detrend",
        }
        
        response = requests.post(
            f"{BASE_URL}/api/export/xlsx", 
            json=payload,
            timeout=120
        )
        
        assert response.status_code == 200
        
        # Should still create valid workbook
        wb = load_workbook(io.BytesIO(response.content))
        assert 'Light Stimulus HRV' in wb.sheetnames


class TestExportDetrendedVisualization:
    """Tests for detrending visualization data in exports"""
    
    @pytest.fixture(scope="class")
    def detrended_data_with_viz(self):
        """Get detrended data with visualization arrays"""
        # Upload and process file
        with open('/app/4D030006.abf', 'rb') as f:
            files = {'files': ('4D030006.abf', f, 'application/octet-stream')}
            upload_resp = requests.post(f"{BASE_URL}/api/upload", files=files, timeout=60)
        
        file_data = upload_resp.json()['files'][0]
        beat_times_sec = [b['time_sec'] for b in file_data['beats']]
        
        metrics_resp = requests.post(f"{BASE_URL}/api/compute-metrics", json={
            "beat_times_sec": beat_times_sec,
        }, timeout=60)
        metrics = metrics_resp.json()
        
        pulses_resp = requests.post(f"{BASE_URL}/api/light-detect", json={
            "start_time_sec": 180,
            "pulse_duration_sec": 20,
            "n_pulses": 5,
            "auto_detect": True,
            "beat_times_min": metrics['filtered_beat_times_min'],
            "bf_filtered": metrics['filtered_bf_bpm'],
        }, timeout=60)
        pulses = pulses_resp.json()
        
        detrended_resp = requests.post(f"{BASE_URL}/api/light-hrv-detrended", json={
            "beat_times_min": metrics['filtered_beat_times_min'],
            "bf_filtered": metrics['filtered_bf_bpm'],
            "pulses": pulses['pulses'],
            "loess_frac": 0.25,
        }, timeout=60)
        
        return detrended_resp.json()
    
    def test_detrended_has_viz_arrays(self, detrended_data_with_viz):
        """Test detrended data includes visualization arrays"""
        per_pulse = detrended_data_with_viz['per_pulse']
        
        valid_pulses = [p for p in per_pulse if p is not None]
        assert len(valid_pulses) >= 1, "Need at least one valid pulse"
        
        pulse = valid_pulses[0]
        assert 'viz' in pulse, "Pulse missing 'viz' object"
        
        viz = pulse['viz']
        assert 'time_rel' in viz, "viz missing time_rel"
        assert 'nn_70' in viz, "viz missing nn_70"
        assert 'trend' in viz, "viz missing trend"
        assert 'residual' in viz, "viz missing residual"
    
    def test_viz_arrays_have_data(self, detrended_data_with_viz):
        """Test visualization arrays contain actual data"""
        per_pulse = detrended_data_with_viz['per_pulse']
        valid_pulses = [p for p in per_pulse if p is not None]
        
        pulse = valid_pulses[0]
        viz = pulse['viz']
        
        assert len(viz['time_rel']) > 0, "time_rel is empty"
        assert len(viz['nn_70']) > 0, "nn_70 is empty"
        assert len(viz['trend']) > 0, "trend is empty"
        assert len(viz['residual']) > 0, "residual is empty"
        
        # All arrays should have same length
        assert len(viz['time_rel']) == len(viz['nn_70']) == len(viz['trend']) == len(viz['residual'])


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
