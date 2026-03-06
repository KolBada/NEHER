"""
Test Excel Spontaneous BF and HRV sheets - per-drug row highlighting
Tests:
1. Spontaneous BF sheet - drug readout rows highlighted with per-drug fill colors and bold font
2. Spontaneous HRV sheet - drug readout rows highlighted with per-drug fill colors and bold font  
3. Multiple drug rows (2 drugs) should have different purple shades (F3E8FF for drug 1, EDE9FE for drug 2)
4. Baseline row should still be highlighted with baseline_fill color (E0F2FE)
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Per-drug colors as defined in export_utils.py
DRUG_FILL_COLORS = [
    'F3E8FF',  # Drug 1 - Light purple
    'EDE9FE',  # Drug 2 - Light violet
    'E9D5FF',  # Drug 3 - Light fuchsia
    'DDD6FE',  # Drug 4 - Light indigo
]

BASELINE_FILL_COLOR = 'E0F2FE'  # Light blue


def create_test_payload_two_drugs():
    """Create test payload with 2 enabled drugs as per test requirements
    
    Drugs:
    - Tetrodotoxin: start=3, delay=0, bfReadoutMinute=2, hrvReadoutMinute=1
    - Acetylcholine: start=6, delay=0, bfReadoutMinute=4, hrvReadoutMinute=3, enabled=true
    
    Expected BF readout windows:
    - Drug 1 (Tetrodotoxin): minute = 0 + 3 + 2 = 5 -> window "5-6"
    - Drug 2 (Acetylcholine): minute = 0 + 6 + 4 = 10 -> window "10-11"
    
    Wait, let me recalculate per the test requirements:
    - Drug 1: bfReadoutMinute=2 + start=3 + delay=0 = 5 -> window "5-6"
    - Drug 2: bfReadoutMinute=4 + start=6 + delay=0 = 10 -> window "10-11"
    
    Expected HRV readout minutes:
    - Drug 1: hrvReadoutMinute=1 + start=3 + delay=0 = 4
    - Drug 2: hrvReadoutMinute=3 + start=6 + delay=0 = 9
    """
    # Generate per_minute_data covering windows 0-15
    per_minute_data = []
    for i in range(16):
        per_minute_data.append({
            "minute": i,
            "mean_bf": 60 + i,
            "avg_bf": 60 + i,
            "mean_nn": 1000 - i * 10,
            "avg_nn": 1000 - i * 10
        })
    
    # Generate hrv_windows covering minutes 0-15
    hrv_windows = []
    for i in range(16):
        hrv_windows.append({
            "window": f"{i}-{i+3}",
            "minute": i,
            "ln_rmssd70": 3.5 + i * 0.05,
            "rmssd70": 30 + i,
            "sdnn": 50 + i,
            "pnn50": 10 + i * 0.5,
            "mean_bf": 60 + i
        })
    
    return {
        "recording_name": "Test Recording - Two Drugs",
        "filename": "test_two_drugs.txt",
        "baseline_enabled": True,
        "baseline": {
            "baseline_bf_minute": 1,  # baseline at minute 1 -> window "1-2"
            "baseline_bf_range": "1-2",
            "baseline_hrv_minute": 0,  # baseline HRV at minute 0
            "baseline_bf": 61,
            "baseline_ln_rmssd70": 3.55,
            "baseline_sdnn": 51,
            "baseline_pnn50": 10.5
        },
        "drug_readout_enabled": True,
        "all_drugs": [
            {"name": "Tetrodotoxin", "start": 3, "delay": 0},
            {"name": "Acetylcholine", "start": 6, "delay": 0}
        ],
        "drug_readout_settings": {
            "enableHrvReadout": True,
            "enableBfReadout": True,
            "perDrug": {
                "tetrodotoxin": {
                    "hrvReadoutMinute": "1",
                    "bfReadoutMinute": "2"
                },
                "acetylcholine": {
                    "hrvReadoutMinute": "3",
                    "bfReadoutMinute": "4",
                    "enabled": True
                }
            }
        },
        "per_minute_data": per_minute_data,
        "hrv_windows": hrv_windows
    }


class TestExcelSpontaneousBFSheetHighlighting:
    """Tests for Spontaneous BF sheet per-drug row highlighting"""
    
    def test_bf_sheet_exists(self):
        """Test that Spontaneous BF sheet is created"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        wb = load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert 'Spontaneous BF' in sheet_names, f"Spontaneous BF sheet not found. Sheets: {sheet_names}"
        print("PASS: Spontaneous BF sheet exists")
    
    def test_bf_sheet_baseline_row_highlighted(self):
        """Test baseline row in BF sheet has baseline fill color (E0F2FE)"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_bf = wb['Spontaneous BF']
        
        # Baseline is at minute 1 -> window "1-2"
        # Find the row with window "1-2"
        baseline_row = None
        for row in range(5, ws_bf.max_row + 1):
            cell_value = ws_bf.cell(row=row, column=1).value
            if cell_value == "1-2":
                baseline_row = row
                break
        
        assert baseline_row is not None, "Could not find baseline row '1-2' in BF sheet"
        
        # Check fill color of first cell in baseline row
        cell = ws_bf.cell(row=baseline_row, column=1)
        fill_color = cell.fill.start_color.rgb
        
        # RGB color includes alpha channel (ARGB), so remove alpha if present
        if len(fill_color) == 8:
            fill_color = fill_color[2:]  # Remove 'FF' alpha prefix
        
        assert fill_color.upper() == BASELINE_FILL_COLOR.upper(), \
            f"Baseline row fill color mismatch. Expected {BASELINE_FILL_COLOR}, got {fill_color}"
        
        # Check if font is bold
        assert cell.font.bold == True, "Baseline row font should be bold"
        
        print(f"PASS: Baseline row '1-2' has correct fill color {fill_color} and bold font")
    
    def test_bf_sheet_drug1_row_highlighted(self):
        """Test Drug 1 (Tetrodotoxin) row has per-drug fill color F3E8FF"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_bf = wb['Spontaneous BF']
        
        # Drug 1 BF readout: bfReadoutMinute=2 + start=3 + delay=0 = 5 -> window "5-6"
        drug1_row = None
        for row in range(5, ws_bf.max_row + 1):
            cell_value = ws_bf.cell(row=row, column=1).value
            if cell_value == "5-6":
                drug1_row = row
                break
        
        assert drug1_row is not None, "Could not find Drug 1 row '5-6' in BF sheet"
        
        # Check fill color
        cell = ws_bf.cell(row=drug1_row, column=1)
        fill_color = cell.fill.start_color.rgb
        
        if len(fill_color) == 8:
            fill_color = fill_color[2:]
        
        assert fill_color.upper() == DRUG_FILL_COLORS[0].upper(), \
            f"Drug 1 row fill color mismatch. Expected {DRUG_FILL_COLORS[0]}, got {fill_color}"
        
        # Check if font is bold
        assert cell.font.bold == True, "Drug 1 row font should be bold"
        
        print(f"PASS: Drug 1 row '5-6' has correct fill color {fill_color} and bold font")
    
    def test_bf_sheet_drug2_row_highlighted(self):
        """Test Drug 2 (Acetylcholine) row has different per-drug fill color EDE9FE"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_bf = wb['Spontaneous BF']
        
        # Drug 2 BF readout: bfReadoutMinute=4 + start=6 + delay=0 = 10 -> window "10-11"
        drug2_row = None
        for row in range(5, ws_bf.max_row + 1):
            cell_value = ws_bf.cell(row=row, column=1).value
            if cell_value == "10-11":
                drug2_row = row
                break
        
        assert drug2_row is not None, "Could not find Drug 2 row '10-11' in BF sheet"
        
        # Check fill color
        cell = ws_bf.cell(row=drug2_row, column=1)
        fill_color = cell.fill.start_color.rgb
        
        if len(fill_color) == 8:
            fill_color = fill_color[2:]
        
        assert fill_color.upper() == DRUG_FILL_COLORS[1].upper(), \
            f"Drug 2 row fill color mismatch. Expected {DRUG_FILL_COLORS[1]}, got {fill_color}"
        
        # Check if font is bold
        assert cell.font.bold == True, "Drug 2 row font should be bold"
        
        print(f"PASS: Drug 2 row '10-11' has correct fill color {fill_color} and bold font")
    
    def test_bf_sheet_drugs_have_different_colors(self):
        """Test that Drug 1 and Drug 2 rows have different fill colors"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_bf = wb['Spontaneous BF']
        
        # Find both drug rows
        drug1_color = None
        drug2_color = None
        
        for row in range(5, ws_bf.max_row + 1):
            cell_value = ws_bf.cell(row=row, column=1).value
            if cell_value == "5-6":  # Drug 1
                color = ws_bf.cell(row=row, column=1).fill.start_color.rgb
                drug1_color = color[2:] if len(color) == 8 else color
            elif cell_value == "10-11":  # Drug 2
                color = ws_bf.cell(row=row, column=1).fill.start_color.rgb
                drug2_color = color[2:] if len(color) == 8 else color
        
        assert drug1_color is not None, "Drug 1 row not found"
        assert drug2_color is not None, "Drug 2 row not found"
        assert drug1_color.upper() != drug2_color.upper(), \
            f"Drug 1 and Drug 2 should have different colors, both have {drug1_color}"
        
        print(f"PASS: Drug rows have different colors - Drug 1: {drug1_color}, Drug 2: {drug2_color}")


class TestExcelSpontaneousHRVSheetHighlighting:
    """Tests for Spontaneous HRV sheet per-drug row highlighting"""
    
    def test_hrv_sheet_exists(self):
        """Test that Spontaneous HRV sheet is created"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert 'Spontaneous HRV' in sheet_names, f"Spontaneous HRV sheet not found. Sheets: {sheet_names}"
        print("PASS: Spontaneous HRV sheet exists")
    
    def test_hrv_sheet_baseline_row_highlighted(self):
        """Test baseline row in HRV sheet has baseline fill color (E0F2FE)"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_hrv = wb['Spontaneous HRV']
        
        # Baseline HRV is at minute 0 -> window "0-3"
        # The window column format may vary, let's check the minute in the window
        baseline_row = None
        for row in range(5, ws_hrv.max_row + 1):
            cell_value = str(ws_hrv.cell(row=row, column=1).value or '')
            # Window format is "X-X+3", so minute 0 would be "0-3"
            if cell_value.startswith('0-') or cell_value == '0':
                baseline_row = row
                break
        
        assert baseline_row is not None, "Could not find baseline row at minute 0 in HRV sheet"
        
        # Check fill color
        cell = ws_hrv.cell(row=baseline_row, column=1)
        fill_color = cell.fill.start_color.rgb
        
        if len(fill_color) == 8:
            fill_color = fill_color[2:]
        
        assert fill_color.upper() == BASELINE_FILL_COLOR.upper(), \
            f"Baseline row fill color mismatch. Expected {BASELINE_FILL_COLOR}, got {fill_color}"
        
        # Check if font is bold
        assert cell.font.bold == True, "Baseline row font should be bold"
        
        print(f"PASS: Baseline HRV row has correct fill color {fill_color} and bold font")
    
    def test_hrv_sheet_drug1_row_highlighted(self):
        """Test Drug 1 (Tetrodotoxin) HRV row has per-drug fill color F3E8FF"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_hrv = wb['Spontaneous HRV']
        
        # Drug 1 HRV readout: hrvReadoutMinute=1 + start=3 + delay=0 = 4 -> window "4-7"
        drug1_row = None
        for row in range(5, ws_hrv.max_row + 1):
            cell_value = str(ws_hrv.cell(row=row, column=1).value or '')
            if cell_value.startswith('4-'):
                drug1_row = row
                break
        
        assert drug1_row is not None, "Could not find Drug 1 HRV row at minute 4"
        
        # Check fill color
        cell = ws_hrv.cell(row=drug1_row, column=1)
        fill_color = cell.fill.start_color.rgb
        
        if len(fill_color) == 8:
            fill_color = fill_color[2:]
        
        assert fill_color.upper() == DRUG_FILL_COLORS[0].upper(), \
            f"Drug 1 HRV row fill color mismatch. Expected {DRUG_FILL_COLORS[0]}, got {fill_color}"
        
        # Check if font is bold
        assert cell.font.bold == True, "Drug 1 HRV row font should be bold"
        
        print(f"PASS: Drug 1 HRV row (minute 4) has correct fill color {fill_color} and bold font")
    
    def test_hrv_sheet_drug2_row_highlighted(self):
        """Test Drug 2 (Acetylcholine) HRV row has different per-drug fill color EDE9FE"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_hrv = wb['Spontaneous HRV']
        
        # Drug 2 HRV readout: hrvReadoutMinute=3 + start=6 + delay=0 = 9 -> window "9-12"
        drug2_row = None
        for row in range(5, ws_hrv.max_row + 1):
            cell_value = str(ws_hrv.cell(row=row, column=1).value or '')
            if cell_value.startswith('9-'):
                drug2_row = row
                break
        
        assert drug2_row is not None, "Could not find Drug 2 HRV row at minute 9"
        
        # Check fill color
        cell = ws_hrv.cell(row=drug2_row, column=1)
        fill_color = cell.fill.start_color.rgb
        
        if len(fill_color) == 8:
            fill_color = fill_color[2:]
        
        assert fill_color.upper() == DRUG_FILL_COLORS[1].upper(), \
            f"Drug 2 HRV row fill color mismatch. Expected {DRUG_FILL_COLORS[1]}, got {fill_color}"
        
        # Check if font is bold
        assert cell.font.bold == True, "Drug 2 HRV row font should be bold"
        
        print(f"PASS: Drug 2 HRV row (minute 9) has correct fill color {fill_color} and bold font")
    
    def test_hrv_sheet_drugs_have_different_colors(self):
        """Test that Drug 1 and Drug 2 HRV rows have different fill colors"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_hrv = wb['Spontaneous HRV']
        
        # Find both drug rows by minute
        drug1_color = None
        drug2_color = None
        
        for row in range(5, ws_hrv.max_row + 1):
            cell_value = str(ws_hrv.cell(row=row, column=1).value or '')
            if cell_value.startswith('4-'):  # Drug 1 at minute 4
                color = ws_hrv.cell(row=row, column=1).fill.start_color.rgb
                drug1_color = color[2:] if len(color) == 8 else color
            elif cell_value.startswith('9-'):  # Drug 2 at minute 9
                color = ws_hrv.cell(row=row, column=1).fill.start_color.rgb
                drug2_color = color[2:] if len(color) == 8 else color
        
        assert drug1_color is not None, "Drug 1 HRV row not found"
        assert drug2_color is not None, "Drug 2 HRV row not found"
        assert drug1_color.upper() != drug2_color.upper(), \
            f"Drug 1 and Drug 2 HRV rows should have different colors, both have {drug1_color}"
        
        print(f"PASS: Drug HRV rows have different colors - Drug 1: {drug1_color}, Drug 2: {drug2_color}")


class TestExcelAllColumnsHighlighted:
    """Test that all columns in a highlighted row have the same fill color"""
    
    def test_bf_sheet_all_columns_highlighted(self):
        """Test all columns in BF drug row have same fill color"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_bf = wb['Spontaneous BF']
        
        # Find Drug 1 row
        drug1_row = None
        for row in range(5, ws_bf.max_row + 1):
            if ws_bf.cell(row=row, column=1).value == "5-6":
                drug1_row = row
                break
        
        assert drug1_row is not None
        
        # Check all 3 columns have same fill
        colors = []
        for col in range(1, 4):
            color = ws_bf.cell(row=drug1_row, column=col).fill.start_color.rgb
            colors.append(color[2:] if len(color) == 8 else color)
        
        assert all(c.upper() == colors[0].upper() for c in colors), \
            f"Not all columns have same fill color: {colors}"
        
        print(f"PASS: All 3 columns in BF drug row have same fill color {colors[0]}")
    
    def test_hrv_sheet_all_columns_highlighted(self):
        """Test all columns in HRV drug row have same fill color"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_hrv = wb['Spontaneous HRV']
        
        # Find Drug 1 row at minute 4
        drug1_row = None
        for row in range(5, ws_hrv.max_row + 1):
            cell_value = str(ws_hrv.cell(row=row, column=1).value or '')
            if cell_value.startswith('4-'):
                drug1_row = row
                break
        
        assert drug1_row is not None
        
        # Check all 7 columns have same fill (HRV sheet has 7 columns)
        colors = []
        for col in range(1, 8):
            color = ws_hrv.cell(row=drug1_row, column=col).fill.start_color.rgb
            colors.append(color[2:] if len(color) == 8 else color)
        
        assert all(c.upper() == colors[0].upper() for c in colors), \
            f"Not all columns have same fill color: {colors}"
        
        print(f"PASS: All 7 columns in HRV drug row have same fill color {colors[0]}")


class TestNonHighlightedRowsNoFill:
    """Test that non-readout rows don't have drug/baseline highlighting"""
    
    def test_bf_sheet_normal_row_no_highlight(self):
        """Test non-readout BF rows have no special fill"""
        payload = create_test_payload_two_drugs()
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_bf = wb['Spontaneous BF']
        
        # Find a row that is NOT baseline (1-2), Drug 1 (5-6), or Drug 2 (10-11)
        # Let's check row "3-4"
        normal_row = None
        for row in range(5, ws_bf.max_row + 1):
            if ws_bf.cell(row=row, column=1).value == "3-4":
                normal_row = row
                break
        
        if normal_row is not None:
            cell = ws_bf.cell(row=normal_row, column=1)
            fill_color = cell.fill.start_color.rgb if cell.fill.start_color.rgb else None
            
            # Normal rows should not have drug or baseline colors
            if fill_color and len(fill_color) == 8:
                fill_color = fill_color[2:]
            
            if fill_color:
                assert fill_color.upper() not in [BASELINE_FILL_COLOR.upper(), DRUG_FILL_COLORS[0].upper(), DRUG_FILL_COLORS[1].upper()], \
                    f"Normal row should not have drug/baseline highlight, got {fill_color}"
            
            # Check font is not bold
            assert cell.font.bold != True, "Normal row font should not be bold"
            
            print(f"PASS: Normal row '3-4' has no special highlighting")
        else:
            print("SKIP: Normal row '3-4' not found")


class TestDisabledDrugNotHighlighted:
    """Test that disabled drug readouts don't get highlighting"""
    
    def test_disabled_drug_not_highlighted(self):
        """Test that a drug with enabled=false doesn't get highlighted"""
        # Create payload with Drug 2 disabled
        payload = create_test_payload_two_drugs()
        payload['drug_readout_settings']['perDrug']['acetylcholine']['enabled'] = False
        
        response = requests.post(f"{BASE_URL}/api/export/xlsx", json=payload)
        
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws_bf = wb['Spontaneous BF']
        
        # Find Drug 2 row (10-11) - should NOT be highlighted now
        drug2_row = None
        for row in range(5, ws_bf.max_row + 1):
            if ws_bf.cell(row=row, column=1).value == "10-11":
                drug2_row = row
                break
        
        if drug2_row is not None:
            cell = ws_bf.cell(row=drug2_row, column=1)
            fill_color = cell.fill.start_color.rgb if cell.fill.start_color.rgb else None
            
            if fill_color and len(fill_color) == 8:
                fill_color = fill_color[2:]
            
            # Drug 2 should NOT have the drug highlight color
            if fill_color:
                assert fill_color.upper() != DRUG_FILL_COLORS[1].upper(), \
                    f"Disabled drug should not have highlight, got {fill_color}"
            
            print(f"PASS: Disabled drug row '10-11' is not highlighted")


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
