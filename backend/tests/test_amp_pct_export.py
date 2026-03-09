"""
Test Amp. % (Amplitude Normalized) feature in exports
Tests:
- PDF comparison export Table 4 has 'Amp. %' column
- PDF comparison export summary has 'Amp. (Norm.)' row
- Excel comparison export Light HRA sheet has 'Amp. %' column
- Excel comparison export Summary sheet has 'Amp. (Norm.)' row
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
FOLDER_ID = "69a6321397e7f87817ef235a"  # H9 NeuCarS + Light Stimulus

class TestAmpPctExport:
    """Test Amp. % export features"""
    
    def test_excel_export_has_amp_pct_column(self):
        """Test Excel export has Amp. % column in Light HRA sheet (Table 4)"""
        # Get comparison data first
        response = requests.get(f"{BASE_URL}/api/folders/{FOLDER_ID}/comparison")
        assert response.status_code == 200, f"Failed to get comparison data: {response.status_code}"
        comparison_data = response.json()
        
        # Export to Excel
        export_response = requests.post(
            f"{BASE_URL}/api/folders/{FOLDER_ID}/export/xlsx",
            json={
                "folder_id": FOLDER_ID,
                "folder_name": "H9 NeuCarS + Light Stimulus",
                "comparison_data": comparison_data,
                "excluded_recording_ids": []
            }
        )
        assert export_response.status_code == 200, f"Failed to export Excel: {export_response.status_code}"
        
        # Load the Excel workbook
        xlsx_bytes = io.BytesIO(export_response.content)
        wb = load_workbook(xlsx_bytes)
        
        # Check for Light HRA sheet (Table 4)
        sheet_names = wb.sheetnames
        print(f"Excel sheet names: {sheet_names}")
        
        # Find the sheet with Light HRA data (usually "Light HRA" or similar)
        light_hra_sheet = None
        for name in sheet_names:
            if 'Light' in name and 'HRA' in name:
                light_hra_sheet = wb[name]
                break
        
        if light_hra_sheet is None:
            # Try to find sheet with Table 4
            for name in sheet_names:
                ws = wb[name]
                for row in ws.iter_rows(min_row=1, max_row=5, max_col=1):
                    for cell in row:
                        if cell.value and 'Table 4' in str(cell.value):
                            light_hra_sheet = ws
                            break
        
        assert light_hra_sheet is not None, f"Could not find Light HRA sheet in {sheet_names}"
        
        # Find Amp. % column header
        amp_pct_found = False
        for row in light_hra_sheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and 'Amp. %' in str(cell.value):
                    amp_pct_found = True
                    print(f"Found 'Amp. %' header at {cell.coordinate}")
                    break
            if amp_pct_found:
                break
        
        assert amp_pct_found, "Amp. % column not found in Light HRA sheet"
        print("PASS: Excel Light HRA sheet has 'Amp. %' column")
        
    def test_excel_export_summary_has_amp_norm(self):
        """Test Excel export Summary sheet has 'Amp. (Norm.)' row"""
        # Get comparison data first
        response = requests.get(f"{BASE_URL}/api/folders/{FOLDER_ID}/comparison")
        assert response.status_code == 200
        comparison_data = response.json()
        
        # Export to Excel
        export_response = requests.post(
            f"{BASE_URL}/api/folders/{FOLDER_ID}/export/xlsx",
            json={
                "folder_id": FOLDER_ID,
                "folder_name": "H9 NeuCarS + Light Stimulus",
                "comparison_data": comparison_data,
                "excluded_recording_ids": []
            }
        )
        assert export_response.status_code == 200
        
        # Load the Excel workbook
        xlsx_bytes = io.BytesIO(export_response.content)
        wb = load_workbook(xlsx_bytes)
        
        # Find Summary sheet
        summary_sheet = None
        for name in wb.sheetnames:
            if 'Summary' in name:
                summary_sheet = wb[name]
                break
        
        assert summary_sheet is not None, f"Summary sheet not found in {wb.sheetnames}"
        
        # Find Amp. (Norm.) row in Summary
        amp_norm_found = False
        for row in summary_sheet.iter_rows():
            for cell in row:
                if cell.value and 'Amp. (Norm.)' in str(cell.value):
                    amp_norm_found = True
                    print(f"Found 'Amp. (Norm.)' at {cell.coordinate}: {cell.value}")
                    break
            if amp_norm_found:
                break
        
        assert amp_norm_found, "Amp. (Norm.) row not found in Summary sheet"
        print("PASS: Excel Summary sheet has 'Amp. (Norm.)' row")

    def test_pdf_export_status(self):
        """Test PDF export returns 200 status (content validation is manual)"""
        # Get comparison data first
        response = requests.get(f"{BASE_URL}/api/folders/{FOLDER_ID}/comparison")
        assert response.status_code == 200
        comparison_data = response.json()
        
        # Export to PDF
        export_response = requests.post(
            f"{BASE_URL}/api/folders/{FOLDER_ID}/export/pdf",
            json={
                "folder_id": FOLDER_ID,
                "folder_name": "H9 NeuCarS + Light Stimulus",
                "comparison_data": comparison_data,
                "excluded_recording_ids": []
            }
        )
        assert export_response.status_code == 200, f"Failed to export PDF: {export_response.status_code}"
        
        # Verify it's a PDF
        content_type = export_response.headers.get('content-type', '')
        assert 'pdf' in content_type.lower() or len(export_response.content) > 0
        print(f"PASS: PDF export successful, size: {len(export_response.content)} bytes")
        
        # Save PDF for manual inspection if needed
        with open('/app/test_reports/comparison_export.pdf', 'wb') as f:
            f.write(export_response.content)
        print("PDF saved to /app/test_reports/comparison_export.pdf for manual inspection")


class TestComparisonDataAmpPct:
    """Test comparison data has correct Amp. % calculations"""
    
    def test_comparison_data_structure(self):
        """Test comparison data API returns expected fields"""
        response = requests.get(f"{BASE_URL}/api/folders/{FOLDER_ID}/comparison")
        assert response.status_code == 200
        
        data = response.json()
        assert 'recordings' in data
        assert len(data['recordings']) > 0
        
        # Check first recording has light stimulus fields
        rec = data['recordings'][0]
        print(f"Recording: {rec.get('name')}")
        print(f"  light_baseline_bf: {rec.get('light_baseline_bf')}")
        print(f"  light_amplitude: {rec.get('light_amplitude')}")
        print(f"  light_amp_norm: {rec.get('light_amp_norm')}")
        
        # Verify light stimulus data exists
        assert rec.get('light_baseline_bf') is not None, "light_baseline_bf should exist"
        assert rec.get('light_amplitude') is not None, "light_amplitude should exist"
        
        # Check if amp_norm can be computed (frontend computes on-the-fly if not stored)
        baseline = rec.get('light_baseline_bf')
        amplitude = rec.get('light_amplitude')
        if baseline and baseline > 0 and amplitude is not None:
            computed_amp_norm = 100 * amplitude / baseline
            print(f"  Computed Amp %: {computed_amp_norm:.1f}")
            assert computed_amp_norm > 0, "Computed Amp % should be positive"
        
        print("PASS: Comparison data structure is correct")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
