from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import tempfile
import io
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import numpy as np

import analysis
import storage
import export_utils

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Dependency to get database
async def get_db():
    return db

app = FastAPI()
api_router = APIRouter(prefix="/api")

# In-memory session store (for active analysis sessions only)
sessions = {}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# --- Chunked Upload Models ---
class ChunkInitRequest(BaseModel):
    filename: str
    total_size: int
    total_chunks: int

class ChunkCompleteRequest(BaseModel):
    upload_id: str


# --- Pydantic Models ---
class DetectBeatsRequest(BaseModel):
    session_id: str
    file_id: str
    threshold: Optional[float] = None
    min_distance: Optional[float] = None
    prominence: Optional[float] = None
    invert: bool = False
    bidirectional: bool = False  # Detect peaks both above AND below threshold


class ComputeMetricsRequest(BaseModel):
    beat_times_sec: List[float]
    filter_lower_pct: float = 50.0
    filter_upper_pct: float = 200.0


class HRVAnalysisRequest(BaseModel):
    beat_times_min: List[float]
    bf_filtered: List[float]
    readout_minute: Optional[int] = None
    baseline_hrv_minute: int = 0  # HRV readout at this minute (uses 3-min window)
    baseline_bf_minute: int = 1   # BF readout at this minute


class LightDetectRequest(BaseModel):
    start_time_sec: float = 180.0
    pulse_duration_sec: float = 20.0
    interval_sec: Optional[str] = 'decreasing'
    n_pulses: int = 5
    auto_detect: bool = False
    beat_times_min: Optional[List[float]] = None
    bf_filtered: Optional[List[float]] = None
    search_range_sec: float = 3.0  # Default search window ±3 seconds


class LightHRVRequest(BaseModel):
    beat_times_min: List[float]
    bf_filtered: List[float]
    pulses: List[dict]


class LightResponseRequest(BaseModel):
    beat_times_min: List[float]
    bf_filtered: List[float]
    pulses: List[dict]


class PerMinuteRequest(BaseModel):
    beat_times_min: List[float]
    bf_filtered: List[float]


class ExportRequest(BaseModel):
    per_beat_data: Optional[List[dict]] = None
    hrv_windows: Optional[List[dict]] = None
    light_metrics: Optional[List[dict]] = None
    light_metrics_detrended: Optional[dict] = None  # Corrected HRV (Detrended) data
    light_response: Optional[List[dict]] = None
    light_pulses: Optional[List[dict]] = None  # For showing light stim zones on charts
    light_enabled: Optional[bool] = True  # Whether light stim is enabled
    light_stim_count: Optional[int] = 0  # Number of stims detected
    light_params: Optional[dict] = None  # Light stimulation parameters (duration, ISI, etc.)
    baseline_enabled: Optional[bool] = True  # Whether baseline is enabled
    drug_readout_enabled: Optional[bool] = False  # Whether drug readout is enabled
    drug_readout_settings: Optional[dict] = None  # Drug readout settings
    summary: Optional[dict] = None
    filename: str = "analysis"
    recording_name: Optional[str] = None
    drug_used: Optional[str] = None
    all_drugs: Optional[List[dict]] = None  # Full drug details with start/delay/end
    per_minute_data: Optional[List[dict]] = None
    baseline: Optional[dict] = None
    drug_readout: Optional[dict] = None  # Drug readout timing info for highlighting
    perfusion_params: Optional[dict] = None  # Perfusion parameters for export
    # Full recording beat-by-beat data
    full_recording_data: Optional[List[dict]] = None
    # Light stim isolated beat-by-beat data
    light_stim_data: Optional[List[dict]] = None
    # Original ABF filename
    original_filename: Optional[str] = None
    # Recording metadata
    recording_date: Optional[str] = None  # Date of recording
    fusion_date: Optional[str] = None  # Fusion date (shared for all samples)
    days_since_fusion: Optional[int] = None  # Calculated days since fusion
    organoid_info: Optional[List[dict]] = None  # List of sample info with transfection
    recording_description: Optional[str] = None  # Description/notes


# --- Endpoints ---
@api_router.get("/")
async def root():
    return {"message": "NEHER API"}


# --- Chunked Upload Endpoints (for large files) ---
# Store chunks directly in MongoDB GridFS-style for persistence across restarts

@api_router.post("/upload/init")
async def init_chunked_upload(request: ChunkInitRequest, db=Depends(get_db)):
    """Initialize a chunked upload session - stored in MongoDB for persistence"""
    upload_id = str(uuid.uuid4())
    
    # Store upload session in MongoDB
    await db.upload_sessions.insert_one({
        '_id': upload_id,
        'filename': request.filename,
        'total_size': request.total_size,
        'total_chunks': request.total_chunks,
        'received_chunks': [],
        'created_at': datetime.now(timezone.utc)
    })
    
    logging.info(f"Initialized chunked upload {upload_id} for {request.filename} ({request.total_size} bytes, {request.total_chunks} chunks)")
    
    return {'upload_id': upload_id}


@api_router.post("/upload/chunk/{upload_id}/{chunk_index}")
async def upload_chunk(upload_id: str, chunk_index: int, file: UploadFile = File(...), db=Depends(get_db)):
    """Upload a single chunk - stored in MongoDB for persistence"""
    # Check if session exists
    session = await db.upload_sessions.find_one({'_id': upload_id})
    if not session:
        raise HTTPException(404, "Upload session not found or expired. Please retry the upload.")
    
    if chunk_index < 0 or chunk_index >= session['total_chunks']:
        raise HTTPException(400, f"Invalid chunk index: {chunk_index}")
    
    # Read chunk data
    chunk_data = await file.read()
    
    # Store chunk in MongoDB
    await db.upload_chunks.update_one(
        {'upload_id': upload_id, 'chunk_index': chunk_index},
        {'$set': {
            'upload_id': upload_id,
            'chunk_index': chunk_index,
            'data': chunk_data,
            'size': len(chunk_data)
        }},
        upsert=True
    )
    
    # Update received chunks list
    await db.upload_sessions.update_one(
        {'_id': upload_id},
        {'$addToSet': {'received_chunks': chunk_index}}
    )
    
    # Get updated session
    session = await db.upload_sessions.find_one({'_id': upload_id})
    received_count = len(session.get('received_chunks', []))
    
    logging.info(f"Received chunk {chunk_index + 1}/{session['total_chunks']} for upload {upload_id}")
    
    return {
        'chunk_index': chunk_index,
        'received': received_count,
        'total': session['total_chunks']
    }


@api_router.post("/upload/complete")
async def complete_chunked_upload(request: ChunkCompleteRequest, db=Depends(get_db)):
    """Complete the chunked upload and process the file - reads chunks from MongoDB"""
    import pyabf
    import gc
    
    upload_id = request.upload_id
    
    # Get session from MongoDB
    session = await db.upload_sessions.find_one({'_id': upload_id})
    if not session:
        raise HTTPException(404, "Upload session not found or expired. Please retry the upload.")
    
    # Verify all chunks received
    received_chunks = set(session.get('received_chunks', []))
    total_chunks = session['total_chunks']
    
    if len(received_chunks) != total_chunks:
        missing = set(range(total_chunks)) - received_chunks
        raise HTTPException(400, f"Missing chunks: {sorted(list(missing)[:10])}...")
    
    filename = session['filename']
    temp_path = None
    
    try:
        # Reassemble file from MongoDB chunks
        temp_path = os.path.join(tempfile.gettempdir(), f"assembled_{upload_id}.abf")
        
        with open(temp_path, 'wb') as f:
            for chunk_index in range(total_chunks):
                chunk_doc = await db.upload_chunks.find_one({
                    'upload_id': upload_id,
                    'chunk_index': chunk_index
                })
                if chunk_doc and 'data' in chunk_doc:
                    f.write(chunk_doc['data'])
                else:
                    raise HTTPException(500, f"Chunk {chunk_index} data not found")
        
        logging.info(f"Reassembled file {filename} from {total_chunks} chunks")
        
        # Process the assembled file
        session_id = str(uuid.uuid4())
        sessions[session_id] = {}
        file_id = str(uuid.uuid4())
        
        try:
            abf = pyabf.ABF(temp_path)
        except Exception as parse_err:
            logging.error(f"Failed to parse ABF file '{filename}': {parse_err}")
            raise HTTPException(400, f"Failed to parse ABF file '{filename}': {str(parse_err)}")
        
        abf.setSweep(0, channel=0)
        trace = abf.sweepY.copy().astype(np.float64)
        times = abf.sweepX.copy().astype(np.float64)
        sample_rate = abf.dataRate
        
        # Handle multi-sweep
        if abf.sweepCount > 1:
            all_traces = [trace]
            all_times = [times]
            offset = times[-1] + 1.0 / sample_rate
            for sw in range(1, abf.sweepCount):
                abf.setSweep(sw, channel=0)
                all_traces.append(abf.sweepY.copy().astype(np.float64))
                sw_times = abf.sweepX.copy().astype(np.float64) + offset
                all_times.append(sw_times)
                offset = sw_times[-1] + 1.0 / sample_rate
            trace = np.concatenate(all_traces)
            times = np.concatenate(all_times)
            del all_traces, all_times
            gc.collect()
        
        sessions[session_id][file_id] = {
            'filename': filename,
            'trace': trace,
            'times': times,
            'sample_rate': sample_rate,
        }
        
        # Decimate for response
        dec_times, dec_voltages = analysis.decimate_trace(times, trace)
        beat_indices = analysis.detect_beats(trace, sample_rate)
        beat_times_sec = [float(times[i]) for i in beat_indices if i < len(times)]
        beat_voltages = [float(trace[i]) for i in beat_indices if i < len(trace)]
        
        # Use RAW signal stats so threshold slider matches what user sees in the trace
        signal_stats = {
            'min': float(np.min(trace)),
            'max': float(np.max(trace)),
            'mean': float(np.mean(trace)),
            'std': float(np.std(trace))
        }
        
        result = {
            'session_id': session_id,
            'files': [{
                'file_id': file_id,
                'filename': filename,
                'sample_rate': sample_rate,
                'duration_sec': float(len(trace) / sample_rate),
                'n_samples': len(trace),
                'n_channels': abf.channelCount,
                'n_sweeps': abf.sweepCount,
                'trace_times': dec_times,
                'trace_voltages': dec_voltages,
                'beats': [{'time_sec': t, 'voltage': v} for t, v in zip(beat_times_sec, beat_voltages)],
                'signal_stats': signal_stats,
                'n_beats_detected': len(beat_indices)
            }]
        }
        
        logging.info(f"Successfully processed chunked upload '{filename}': {len(trace)} samples, {len(beat_indices)} beats")
        
        return result
        
    finally:
        # Cleanup: delete temp file
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except:
                pass
        
        # Cleanup: delete chunks and session from MongoDB
        try:
            await db.upload_chunks.delete_many({'upload_id': upload_id})
            await db.upload_sessions.delete_one({'_id': upload_id})
        except:
            pass
        
        gc.collect()


@api_router.post("/upload")
async def upload_files(files: List[UploadFile] = File(...)):
    import pyabf
    import gc

    session_id = str(uuid.uuid4())
    sessions[session_id] = {}
    result_files = []

    for uploaded in files:
        fname = uploaded.filename or ''
        if not fname.lower().endswith('.abf'):
            raise HTTPException(400, f"Only .abf files are supported. Got: '{fname}'. Please rename your file with .abf extension if needed.")

        file_id = str(uuid.uuid4())
        
        # Read file in chunks to handle large files better
        try:
            content = await uploaded.read()
        except Exception as read_err:
            logging.error(f"Failed to read uploaded file '{fname}': {read_err}")
            raise HTTPException(400, f"Failed to read file '{fname}': {str(read_err)}")

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.abf', delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            
            # Free memory from content after writing to temp file
            del content
            gc.collect()

            try:
                abf = pyabf.ABF(tmp_path)
            except Exception as parse_err:
                logging.error(f"Failed to parse ABF file '{fname}': {parse_err}")
                raise HTTPException(400, f"Failed to parse ABF file '{fname}': {str(parse_err)}")
            
            abf.setSweep(0, channel=0)
            trace = abf.sweepY.copy().astype(np.float64)
            times = abf.sweepX.copy().astype(np.float64)
            sample_rate = abf.dataRate

            # Handle multi-sweep by concatenating
            if abf.sweepCount > 1:
                all_traces = [trace]
                all_times = [times]
                offset = times[-1] + 1.0 / sample_rate
                for sw in range(1, abf.sweepCount):
                    abf.setSweep(sw, channel=0)
                    all_traces.append(abf.sweepY.copy().astype(np.float64))
                    sw_times = abf.sweepX.copy().astype(np.float64) + offset
                    all_times.append(sw_times)
                    offset = sw_times[-1] + 1.0 / sample_rate
                trace = np.concatenate(all_traces)
                times = np.concatenate(all_times)
                # Free memory
                del all_traces, all_times
                gc.collect()

            sessions[session_id][file_id] = {
                'filename': uploaded.filename,
                'trace': trace,
                'times': times,
                'sample_rate': sample_rate,
            }

            # Decimate trace for response (reduce data size significantly)
            dec_times, dec_voltages = analysis.decimate_trace(times, trace)
            beat_indices = analysis.detect_beats(trace, sample_rate)
            beat_times_sec = [float(times[i]) for i in beat_indices if i < len(times)]
            beat_voltages = [float(trace[i]) for i in beat_indices if i < len(trace)]

            # Use RAW signal stats so threshold slider matches what user sees in the trace
            signal_stats = {
                'min': float(np.min(trace)),
                'max': float(np.max(trace)),
                'mean': float(np.mean(trace)),
                'std': float(np.std(trace))
            }

            result_files.append({
                'file_id': file_id,
                'filename': uploaded.filename,
                'sample_rate': sample_rate,
                'duration_sec': float(len(trace) / sample_rate),
                'n_samples': len(trace),
                'n_channels': abf.channelCount,
                'n_sweeps': abf.sweepCount,
                'trace_times': dec_times,
                'trace_voltages': dec_voltages,
                'beats': [{'time_sec': t, 'voltage': v} for t, v in zip(beat_times_sec, beat_voltages)],
                'signal_stats': signal_stats,
                'n_beats_detected': len(beat_indices)
            })
            
            logging.info(f"Successfully processed file '{fname}': {len(trace)} samples, {len(beat_indices)} beats")
            
        except HTTPException:
            raise
        except Exception as e:
            logging.error(f"Unexpected error processing '{fname}': {e}")
            raise HTTPException(500, f"Error processing file '{fname}': {str(e)}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except:
                    pass

    return {'session_id': session_id, 'files': result_files}


@api_router.post("/detect-beats")
async def detect_beats_endpoint(request: DetectBeatsRequest):
    session = sessions.get(request.session_id)
    if not session:
        raise HTTPException(404, "Session not found")
    file_data = session.get(request.file_id)
    if not file_data:
        raise HTTPException(404, "File not found")

    trace = file_data['trace']
    times = file_data['times']
    sr = file_data['sample_rate']

    min_dist = int(request.min_distance * sr) if request.min_distance else None
    beat_indices = analysis.detect_beats(
        trace, sr,
        threshold=request.threshold,
        min_distance=min_dist,
        prominence=request.prominence,
        invert=request.invert,
        bidirectional=request.bidirectional
    )

    beat_times_sec = [float(times[i]) for i in beat_indices if i < len(times)]
    beat_voltages = [float(trace[i]) for i in beat_indices if i < len(trace)]

    return {
        'beats': [{'time_sec': t, 'voltage': v} for t, v in zip(beat_times_sec, beat_voltages)],
        'n_beats': len(beat_indices)
    }


@api_router.post("/compute-metrics")
async def compute_metrics_endpoint(request: ComputeMetricsRequest):
    if len(request.beat_times_sec) < 2:
        raise HTTPException(400, "Need at least 2 beats")

    beat_times_sec = sorted(request.beat_times_sec)
    beat_times_min, nn_ms, bf_bpm = analysis.compute_beat_metrics(beat_times_sec)
    # artifact_filter returns (mask, filtered_bf)
    filter_mask, _ = analysis.artifact_filter(
        bf_bpm, 
        lower_pct=request.filter_lower_pct, 
        upper_pct=request.filter_upper_pct
    )

    bt_arr = np.array(beat_times_min)
    nn_arr = np.array(nn_ms)
    bf_arr = np.array(bf_bpm)
    mask_arr = np.array(filter_mask)

    # Filtered values (NN/BF arrays are N-1, correspond to intervals starting at beat k)
    filtered_bt = bt_arr[:-1][mask_arr].tolist()
    filtered_nn = nn_arr[mask_arr].tolist()
    filtered_bf = bf_arr[mask_arr].tolist()

    return {
        'beat_times_min': [float(x) for x in beat_times_min],
        'nn_intervals_ms': [float(x) for x in nn_ms],
        'beat_freq_bpm': [float(x) for x in bf_bpm],
        'artifact_mask': filter_mask,
        'filtered_beat_times_min': [float(x) for x in filtered_bt],
        'filtered_nn_ms': [float(x) for x in filtered_nn],
        'filtered_bf_bpm': [float(x) for x in filtered_bf],
        'n_total': len(beat_times_sec),  # Total beats detected
        'n_removed': int(np.sum(~mask_arr)),  # Beats removed by filter
        'n_kept': len(beat_times_sec) - int(np.sum(~mask_arr)),  # Kept = Detected - Removed
        'filter_settings': {
            'lower_pct': request.filter_lower_pct,
            'upper_pct': request.filter_upper_pct
        }
    }


@api_router.post("/hrv-analysis")
async def hrv_analysis_endpoint(request: HRVAnalysisRequest):
    if len(request.beat_times_min) < 6:
        raise HTTPException(400, "Need at least 6 beats for HRV")
    results, readout = analysis.spontaneous_hrv_analysis(
        request.beat_times_min, request.bf_filtered, request.readout_minute
    )
    
    # Compute baseline metrics - pass hrv_windows to ensure numerical consistency
    baseline = analysis.compute_baseline_metrics(
        request.beat_times_min, request.bf_filtered,
        hrv_windows=results,  # Pass pre-computed windows for HRV baseline lookup
        hrv_minute=request.baseline_hrv_minute,
        bf_minute=request.baseline_bf_minute
    )
    
    return {'windows': results, 'readout': readout, 'baseline': baseline}


@api_router.post("/light-detect")
async def light_detect_endpoint(request: LightDetectRequest):
    start_sec = request.start_time_sec

    # Auto-detect light start from BF increase
    if request.auto_detect and request.beat_times_min and request.bf_filtered:
        start_sec = analysis.auto_detect_light_start(
            request.beat_times_min, request.bf_filtered,
            request.start_time_sec, request.search_range_sec
        )

    # Parse interval
    interval_arg = request.interval_sec
    if interval_arg == 'decreasing' or interval_arg is None:
        interval_arg = None  # Will use default decreasing
    else:
        try:
            interval_arg = float(interval_arg)
        except (ValueError, TypeError):
            interval_arg = None

    # Use guided detection if auto-detect is on and we have BF data
    if request.auto_detect and request.beat_times_min and request.bf_filtered:
        pulses = analysis.generate_pulses_guided(
            start_sec, request.pulse_duration_sec,
            interval_pattern=interval_arg if interval_arg else 'decreasing',
            n_pulses=request.n_pulses,
            beat_times_min_list=request.beat_times_min,
            bf_filtered_list=request.bf_filtered,
            search_window_sec=request.search_range_sec  # Use the request's search range
        )
    else:
        pulses = analysis.generate_pulses(
            start_sec, request.pulse_duration_sec,
            interval_pattern=interval_arg if interval_arg else 'decreasing', 
            n_pulses=request.n_pulses
        )
    return {'pulses': pulses, 'detected_start_sec': start_sec}


class LightAutoDetectAllRequest(BaseModel):
    beat_times_min: List[float]
    bf_filtered: List[float]
    expected_n_pulses: int = 5
    pulse_duration_sec: float = 20.0
    first_pulse_start_sec: Optional[float] = None  # Start time of first pulse (for guided detection)
    pulse_interval_sec: Optional[float] = None      # Interval between pulses (for guided detection)
    search_window_sec: float = 3.0                  # Search window around expected times


@api_router.post("/light-detect-all")
async def light_detect_all_endpoint(request: LightAutoDetectAllRequest):
    """
    Fully automatic detection of all light stimulation pulses.
    Uses BF pattern analysis to find characteristic stim responses.
    
    If first_pulse_start_sec and pulse_interval_sec are provided, uses guided
    detection: searches within ±search_window_sec around each expected pulse time.
    """
    pulses = analysis.auto_detect_all_pulses(
        request.beat_times_min, 
        request.bf_filtered,
        expected_n_pulses=request.expected_n_pulses,
        pulse_duration_sec=request.pulse_duration_sec,
        first_pulse_start_sec=request.first_pulse_start_sec,
        pulse_interval_sec=request.pulse_interval_sec,
        search_window_sec=request.search_window_sec
    )
    
    if pulses:
        return {'success': True, 'pulses': pulses, 'n_detected': len(pulses)}
    else:
        return {'success': False, 'pulses': None, 'message': 'Could not auto-detect pulses. Please set manually.'}


@api_router.post("/light-hrv")
async def light_hrv_endpoint(request: LightHRVRequest):
    hrv_result = analysis.compute_light_hrv(
        request.beat_times_min, request.bf_filtered, request.pulses
    )
    return {'per_pulse': hrv_result['per_pulse'], 'final': hrv_result['final']}


@api_router.post("/light-response")
async def light_response_endpoint(request: LightResponseRequest):
    per_stim, mean_metrics, baseline_bf = analysis.compute_light_response_v2(
        request.beat_times_min, request.bf_filtered, request.pulses
    )
    return {
        'per_stim': per_stim,
        'mean_metrics': mean_metrics,
        'baseline_bf': baseline_bf
    }


class LightHRVDetrendedRequest(BaseModel):
    beat_times_min: List[float]
    bf_filtered: List[float]
    pulses: List[dict]
    loess_frac: float = 0.25  # LOESS span (20-30% of stim duration)


@api_router.post("/light-hrv-detrended")
async def light_hrv_detrended_endpoint(request: LightHRVDetrendedRequest):
    """
    Compute Corrected Light-Induced HRV using LOESS detrending.
    Removes slow deterministic adaptation curves to isolate true beat-to-beat variability.
    """
    result = analysis.compute_light_hrv_detrended(
        request.beat_times_min, request.bf_filtered, request.pulses, request.loess_frac
    )
    return {'per_pulse': result['per_pulse'], 'final': result['final']}


@api_router.post("/per-minute-metrics")
async def per_minute_metrics_endpoint(request: PerMinuteRequest):
    if len(request.beat_times_min) < 2:
        raise HTTPException(400, "Need at least 2 beats")
    # per_minute_aggregation requires nn_70, compute it from bf_filtered
    nn_values = analysis.bf_to_nn(request.bf_filtered)
    nn_70 = analysis.normalize_nn_70_windowing(request.beat_times_min, nn_values)
    rows = analysis.per_minute_aggregation(request.beat_times_min, request.bf_filtered, nn_70)
    return {'rows': rows}


# ==============================================================================
# STORAGE API - Folders and Recordings
# ==============================================================================

@api_router.get("/folders")
async def get_folders_endpoint():
    """Get all folders with recording counts."""
    folders = await storage.get_folders(db)
    return {"folders": folders}


@api_router.post("/folders")
async def create_folder_endpoint(request: storage.FolderCreate):
    """Create a new folder."""
    folder = await storage.create_folder(db, request.name)
    return folder


@api_router.get("/folders/{folder_id}")
async def get_folder_endpoint(folder_id: str):
    """Get a single folder."""
    folder = await storage.get_folder(db, folder_id)
    if not folder:
        raise HTTPException(404, "Folder not found")
    return folder


@api_router.put("/folders/{folder_id}")
async def update_folder_endpoint(folder_id: str, request: storage.FolderUpdate):
    """Update a folder's name, color, or section."""
    folder = await storage.update_folder(db, folder_id, request.name, request.color, request.section_id)
    if not folder:
        raise HTTPException(404, "Folder not found")
    return folder


@api_router.delete("/folders/{folder_id}")
async def delete_folder_endpoint(folder_id: str):
    """Delete a folder and all its recordings."""
    success = await storage.delete_folder(db, folder_id)
    if not success:
        raise HTTPException(404, "Folder not found")
    return {"success": True}


# ==============================================================================
# Section Endpoints
# ==============================================================================

@api_router.get("/sections")
async def get_sections_endpoint():
    """Get all sections."""
    sections = await storage.get_sections(db)
    return {"sections": sections}


@api_router.post("/sections")
async def create_section_endpoint(request: storage.SectionCreate):
    """Create a new section."""
    section = await storage.create_section(db, request.name)
    return section


@api_router.put("/sections/{section_id}")
async def update_section_endpoint(section_id: str, request: storage.SectionUpdate):
    """Update a section."""
    section = await storage.update_section(db, section_id, request.name, request.order, request.expanded)
    if not section:
        raise HTTPException(404, "Section not found")
    return section


@api_router.delete("/sections/{section_id}")
async def delete_section_endpoint(section_id: str):
    """Delete a section."""
    success = await storage.delete_section(db, section_id)
    if not success:
        raise HTTPException(404, "Section not found")
    return {"success": True}


@api_router.post("/sections/reorder")
async def reorder_sections_endpoint(section_ids: List[str]):
    """Reorder sections."""
    success = await storage.reorder_sections(db, section_ids)
    if not success:
        raise HTTPException(500, "Failed to reorder sections")
    return {"success": True}


@api_router.get("/folders/{folder_id}/recordings")
async def get_recordings_endpoint(folder_id: str):
    """Get all recordings in a folder."""
    # Verify folder exists
    folder = await storage.get_folder(db, folder_id)
    if not folder:
        raise HTTPException(404, "Folder not found")
    recordings = await storage.get_recordings_in_folder(db, folder_id)
    return {"folder": folder, "recordings": recordings}


@api_router.post("/recordings")
async def create_recording_endpoint(request: storage.RecordingCreate):
    """Create a new recording in a folder."""
    # Verify folder exists
    folder = await storage.get_folder(db, request.folder_id)
    if not folder:
        raise HTTPException(404, "Folder not found")
    
    # Check for duplicates
    is_duplicate = await storage.check_duplicate_recording(db, request.folder_id, request.filename)
    if is_duplicate:
        raise HTTPException(400, f"A recording with filename '{request.filename}' already exists in this folder")
    
    recording = await storage.create_recording(db, request.folder_id, request.name, request.filename, request.analysis_state)
    return recording


@api_router.get("/recordings/{recording_id}")
async def get_recording_endpoint(recording_id: str):
    """Get a single recording with full analysis state."""
    recording = await storage.get_recording(db, recording_id)
    if not recording:
        raise HTTPException(404, "Recording not found")
    
    # Fix n_kept to ensure Detected = Kept + Removed for saved recordings
    if recording.get('analysis_state') and recording['analysis_state'].get('metrics'):
        metrics = recording['analysis_state']['metrics']
        if 'n_total' in metrics and 'n_removed' in metrics:
            # Recalculate n_kept to ensure the equation holds
            metrics['n_kept'] = metrics['n_total'] - metrics['n_removed']
    
    return recording


@api_router.put("/recordings/{recording_id}")
async def update_recording_endpoint(recording_id: str, request: storage.RecordingUpdate):
    """Update a recording's name and/or analysis state."""
    print(f"[DEBUG] Updating recording with ID: {recording_id}")
    print(f"[DEBUG] Request name: {request.name}")
    print(f"[DEBUG] Has analysis_state: {request.analysis_state is not None}")
    recording = await storage.update_recording(db, recording_id, request.name, request.analysis_state)
    if not recording:
        print(f"[DEBUG] Recording not found or update failed for ID: {recording_id}")
        raise HTTPException(404, "Recording not found")
    print(f"[DEBUG] Recording updated successfully: {recording.get('name')}")
    return recording


@api_router.delete("/recordings/{recording_id}")
async def delete_recording_endpoint(recording_id: str):
    """Delete a recording."""
    success = await storage.delete_recording(db, recording_id)
    if not success:
        raise HTTPException(404, "Recording not found")
    return {"success": True}


@api_router.post("/recordings/{recording_id}/move")
async def move_recording_endpoint(recording_id: str, request: storage.RecordingMove):
    """Move a recording to a different folder."""
    recording = await storage.move_recording(db, recording_id, request.target_folder_id)
    if not recording:
        raise HTTPException(400, "Could not move recording. Target folder may not exist or a duplicate filename exists.")
    return recording


@api_router.post("/recordings/batch-update")
async def batch_update_recordings():
    """
    Check all recordings for outdated metrics and recompute them.
    Only recomputes sections that were already computed in each recording.
    Returns list of updated recording names.
    """
    import logging
    
    outdated = await storage.get_outdated_recordings(db)
    updated_recordings = []
    
    for rec in outdated:
        try:
            state = rec["analysis_state"]
            recording_name = rec["name"]
            updated_sections = []
            
            # Get the required data for recomputation
            metrics = state.get("metrics")
            if not metrics:
                continue  # Can't recompute without base metrics
            
            beat_times_min = metrics.get("filtered_beat_times_min", [])
            bf_filtered = metrics.get("filtered_bf_bpm", [])
            
            if not beat_times_min or not bf_filtered:
                continue  # Can't recompute without beat data
            
            # Check and recompute HRV results (Spontaneous Activity)
            if state.get("hrvResults"):
                try:
                    # Compute nn_70 from beat frequency
                    nn_values = analysis.bf_to_nn(bf_filtered)
                    nn_70 = analysis.normalize_nn_70_windowing(beat_times_min, nn_values)
                    hrv_windows = analysis.rolling_3min_hrv(beat_times_min, nn_70, bf_filtered)
                    baseline = None
                    if hrv_windows:
                        baseline = analysis.compute_baseline_metrics(
                            beat_times_min, bf_filtered, 
                            hrv_windows=hrv_windows, hrv_minute=0, bf_minute=1
                        )
                    state["hrvResults"] = {
                        "windows": hrv_windows,
                        "baseline": baseline
                    }
                    updated_sections.append("HRV")
                except Exception as e:
                    logging.warning(f"HRV recomputation failed for {recording_name}: {e}")
            
            # Check and recompute Light Response (HRA)
            if state.get("lightResponse") and state.get("lightPulses"):
                try:
                    pulses = state["lightPulses"]
                    per_stim, mean_metrics, baseline_bf = analysis.compute_light_response_v2(
                        beat_times_min, bf_filtered, pulses
                    )
                    state["lightResponse"] = {
                        "per_stim": per_stim,
                        "mean_metrics": mean_metrics,
                        "baseline_bf": baseline_bf
                    }
                    updated_sections.append("Light HRA")
                except Exception as e:
                    logging.warning(f"Light HRA recomputation failed for {recording_name}: {e}")
            
            # Check and recompute Light HRV
            if state.get("lightHrv") and state.get("lightPulses"):
                try:
                    pulses = state["lightPulses"]
                    per_pulse, final_metrics = analysis.compute_light_hrv(
                        beat_times_min, bf_filtered, pulses
                    )
                    state["lightHrv"] = {
                        "per_pulse": per_pulse,
                        "final": final_metrics
                    }
                    updated_sections.append("Light HRV")
                except Exception as e:
                    logging.warning(f"Light HRV recomputation failed for {recording_name}: {e}")
            
            # Check and recompute Detrended HRV
            if state.get("lightHrvDetrended") and state.get("lightPulses"):
                try:
                    pulses = state["lightPulses"]
                    loess_frac = state.get("lightParams", {}).get("loessFrac", 0.25)
                    detrended_results = analysis.compute_light_hrv_detrended(
                        beat_times_min, bf_filtered, pulses, loess_frac
                    )
                    state["lightHrvDetrended"] = detrended_results
                    updated_sections.append("Detrended HRV")
                except Exception as e:
                    logging.warning(f"Detrended HRV recomputation failed for {recording_name}: {e}")
            
            # Save updated state if any sections were updated
            if updated_sections:
                success = await storage.update_recording_metrics_version(db, rec["id"], state)
                if success:
                    updated_recordings.append({
                        "name": recording_name,
                        "sections": updated_sections
                    })
                    logging.info(f"Updated recording '{recording_name}': {', '.join(updated_sections)}")
        
        except Exception as e:
            logging.error(f"Error updating recording {rec.get('name', 'unknown')}: {e}")
            continue
    
    return {
        "updated_count": len(updated_recordings),
        "current_version": storage.METRICS_VERSION,
        "recordings": updated_recordings
    }


# ==============================================================================
# FOLDER COMPARISON API
# ==============================================================================

class FolderComparisonExportRequest(BaseModel):
    folder_id: str
    folder_name: str
    comparison_data: dict


def extract_comparison_metrics(recording: dict) -> dict:
    """Extract comparison metrics from a recording's analysis_state.
    
    Note: Frontend saves with camelCase keys, so we need to check both cases.
    """
    state = recording.get('analysis_state', {})
    
    # Basic info
    result = {
        'id': recording.get('id', ''),
        'name': recording.get('name', ''),
        'filename': recording.get('filename', ''),
    }
    
    # Recording metadata - check both camelCase and snake_case
    result['recording_date'] = state.get('recordingDate') or state.get('recording_date', '')
    result['recording_description'] = state.get('recordingDescription') or state.get('recording_description', '')
    result['fusion_date'] = state.get('fusionDate') or state.get('fusion_date', '')
    
    # Calculate fusion age (days from fusion to recording)
    result['fusion_age'] = None
    fusion_date = state.get('fusionDate') or state.get('fusion_date', '')
    rec_date = state.get('recordingDate') or state.get('recording_date', '')
    if fusion_date and rec_date:
        try:
            from datetime import datetime
            fd = datetime.strptime(fusion_date, '%Y-%m-%d')
            rd = datetime.strptime(rec_date, '%Y-%m-%d')
            result['fusion_age'] = (rd - fd).days
        except (ValueError, TypeError):
            pass
    
    # Organoid/Cell info - frontend uses camelCase 'organoidInfo'
    organoid_info = state.get('organoidInfo') or state.get('organoid_info', [])
    
    # Initialize organoid-related fields
    result['hspo_info'] = None
    result['hco_info'] = None
    result['other_info'] = None
    result['hspo_age'] = None
    result['hco_age'] = None
    
    if organoid_info:
        for sample in organoid_info:
            cell_type = sample.get('cell_type', '')
            line_name = sample.get('line_name', '')
            passage = sample.get('passage_number', '')
            
            # Calculate age from birth_date if available
            age = sample.get('age_at_recording')
            if age is None:
                birth_date = sample.get('birth_date') or sample.get('diff_date', '')
                rec_date = state.get('recordingDate') or state.get('recording_date', '')
                if birth_date and rec_date:
                    try:
                        from datetime import datetime
                        bd = datetime.strptime(birth_date, '%Y-%m-%d')
                        rd = datetime.strptime(rec_date, '%Y-%m-%d')
                        age = (rd - bd).days
                    except (ValueError, TypeError):
                        pass
            
            # Check for transduction/transfection
            transfection = sample.get('transfection', {})
            has_transfection = bool(transfection and transfection.get('technique'))
            
            sample_info = {
                'line_name': line_name,
                'passage': passage,
                'age': age,
                'has_transduction': has_transfection,
                'transfection_details': transfection,
            }
            
            if cell_type == 'hSpO':
                result['hspo_info'] = sample_info
                result['hspo_age'] = age
            elif cell_type == 'hCO':
                result['hco_info'] = sample_info
                result['hco_age'] = age
            elif cell_type:
                other_type = sample.get('other_cell_type', cell_type)
                sample_info['cell_type'] = other_type
                result['other_info'] = sample_info
    
    # Drug info - frontend uses camelCase
    selected_drugs = state.get('selectedDrugs') or state.get('selected_drugs', [])
    drug_settings = state.get('drugSettings') or state.get('drug_settings', {})
    other_drugs_list = state.get('otherDrugs') or state.get('other_drugs', [])
    drug_readout_settings = state.get('drugReadoutSettings') or state.get('drug_readout_settings', {})
    
    # DRUG_CONFIG equivalent for BF readout times
    DRUG_BF_READOUTS = {
        'tetrodotoxin': 12,
        'isoproterenol': None,  # manual peak
        'acetylcholine': 3,
        'propranolol': 12,
        'nepicastat': 42,
        'ruxolitinib': 15,
    }
    
    # Build drug info list
    result['drug_info'] = []
    result['has_drug'] = bool(selected_drugs) or bool(other_drugs_list)
    
    if selected_drugs:
        for drug in selected_drugs:
            drug_name = drug
            settings = drug_settings.get(drug, {})
            concentration = settings.get('concentration', '')
            
            # Get BF readout time (perfusion time for BF)
            bf_readout = DRUG_BF_READOUTS.get(drug.lower(), None)
            # Check if there's a custom override in drugReadoutSettings
            custom_bf = drug_readout_settings.get('bfReadoutMinute')
            if custom_bf:
                bf_readout = custom_bf
            
            result['drug_info'].append({
                'name': drug_name,
                'concentration': concentration,
                'bf_readout_time': bf_readout,
            })
    
    # Add other drugs (custom drugs added by user)
    if other_drugs_list and isinstance(other_drugs_list, list):
        for drug in other_drugs_list:
            if isinstance(drug, dict) and drug.get('name'):
                # For custom drugs, use their perfusionTime as bf_readout_time
                perf_time = drug.get('perfusionTime', drug.get('perfusion_time', ''))
                result['drug_info'].append({
                    'name': drug.get('name', ''),
                    'concentration': drug.get('concentration', ''),
                    'bf_readout_time': perf_time if perf_time else None,
                })
    
    if result['drug_info']:
        # For backwards compat
        result['drug_names'] = ', '.join([d['name'] for d in result['drug_info']])
        result['drug_concentrations'] = ', '.join([str(d['concentration']) for d in result['drug_info'] if d['concentration']])
        result['condition'] = f"Drug: {result['drug_names']}"
    else:
        result['drug_names'] = ''
        result['drug_concentrations'] = ''
        result['condition'] = 'Control'
    
    # Light stim info - use ORIGINAL search parameters (not adjusted pulses)
    light_enabled = state.get('lightEnabled', False)
    light_params = state.get('lightParams') or state.get('light_params', {})
    light_pulses = state.get('lightPulses') or state.get('light_pulses', [])
    
    result['has_light_stim'] = light_enabled and bool(light_pulses)
    
    # Use original search parameters from lightParams
    pulse_duration = light_params.get('pulseDuration', light_params.get('pulse_duration_sec', 20)) if light_params else 20
    result['stim_duration'] = pulse_duration
    
    # Get ISI structure from original interval setting, not from calculated pulse timings
    interval_setting = light_params.get('interval', 'decreasing') if light_params else 'decreasing'
    if interval_setting == 'decreasing':
        result['isi_structure'] = '60s-30s-20s-10s'
    elif interval_setting == 'constant':
        result['isi_structure'] = 'constant'
    else:
        # Custom or unknown - fall back to calculated ISI from pulses
        if light_pulses and len(light_pulses) > 1:
            intervals = []
            for i in range(1, min(len(light_pulses), 5)):
                start_curr = light_pulses[i].get('start_min', 0)
                start_prev = light_pulses[i-1].get('start_min', 0)
                interval_sec = (start_curr - start_prev) * 60
                intervals.append(int(round(interval_sec)))
            result['isi_structure'] = '-'.join([f"{i}s" for i in intervals])
        else:
            result['isi_structure'] = ''
    
    # Spontaneous Activity - Baseline metrics from hrvResults
    hrv_results = state.get('hrvResults') or state.get('hrv_results', {})
    baseline = hrv_results.get('baseline', {})
    
    # Check if baseline is enabled (default True for backward compatibility)
    baseline_enabled = state.get('baselineEnabled', True)
    
    if baseline_enabled:
        result['baseline_bf'] = baseline.get('baseline_bf')
        result['baseline_ln_rmssd70'] = baseline.get('baseline_ln_rmssd70')
        baseline_sdnn = baseline.get('baseline_sdnn')
        result['baseline_ln_sdnn70'] = np.log(baseline_sdnn) if baseline_sdnn and baseline_sdnn > 0 else None
        result['baseline_pnn50'] = baseline.get('baseline_pnn50')
    else:
        result['baseline_bf'] = None
        result['baseline_ln_rmssd70'] = None
        result['baseline_ln_sdnn70'] = None
        result['baseline_pnn50'] = None
    
    result['baseline_enabled'] = baseline_enabled
    
    # Spontaneous Activity - Drug metrics
    drug_readout = state.get('drugReadoutSettings') or state.get('drug_readout', {})
    hrv_windows = hrv_results.get('windows', [])
    per_minute_data = state.get('perMinuteData') or state.get('per_minute_data', [])
    
    result['drug_bf'] = None
    result['drug_ln_rmssd70'] = None
    result['drug_ln_sdnn70'] = None
    result['drug_pnn50'] = None
    
    if drug_readout and (drug_readout.get('enableHrvReadout') or drug_readout.get('enableBfReadout')):
        # Frontend uses 'bfReadoutMinute' and 'hrvReadoutMinute' (as strings)
        drug_bf_minute_str = drug_readout.get('bfReadoutMinute') or drug_readout.get('bfMinute') or drug_readout.get('bf_minute')
        drug_hrv_minute_str = drug_readout.get('hrvReadoutMinute') or drug_readout.get('hrvMinute') or drug_readout.get('hrv_minute')
        
        # Convert to int
        drug_bf_minute = None
        drug_hrv_minute = None
        try:
            if drug_bf_minute_str is not None:
                drug_bf_minute = int(drug_bf_minute_str)
        except (ValueError, TypeError):
            pass
        try:
            if drug_hrv_minute_str is not None:
                drug_hrv_minute = int(drug_hrv_minute_str)
        except (ValueError, TypeError):
            pass
        
        # Get drug BF from per_minute_data
        if drug_bf_minute is not None and per_minute_data:
            for pm in per_minute_data:
                minute_val = pm.get('minute', '0')
                try:
                    minute_num = int(str(minute_val).split('-')[0])
                    if minute_num == drug_bf_minute:
                        result['drug_bf'] = pm.get('avg_bf') or pm.get('mean_bf')
                        break
                except (ValueError, TypeError, AttributeError):
                    pass
        
        # Get drug HRV from hrv_windows
        if drug_hrv_minute is not None and hrv_windows:
            for w in hrv_windows:
                if w.get('minute') == drug_hrv_minute:
                    result['drug_ln_rmssd70'] = w.get('ln_rmssd70')
                    sdnn = w.get('sdnn')
                    result['drug_ln_sdnn70'] = np.log(sdnn) if sdnn and sdnn > 0 else None
                    result['drug_pnn50'] = w.get('pnn50')
                    break
    
    # Light HRA metrics - frontend uses camelCase
    # lightResponse is a dict: {per_stim: [...], mean_metrics: {...}, baseline_bf: number}
    light_response = state.get('lightResponse') or state.get('light_response')
    
    # Initialize light metrics
    for key in ['light_baseline_bf', 'light_avg_bf', 'light_peak_bf', 'light_peak_norm', 
               'light_ttp_first', 'light_ttp_avg', 'light_recovery_bf', 'light_recovery_pct',
               'light_amplitude', 'light_roc']:
        result[key] = None
    
    if light_response and isinstance(light_response, dict):
        # Use mean_metrics if available (pre-computed averages)
        mean_metrics = light_response.get('mean_metrics', {})
        per_stim = light_response.get('per_stim', [])
        
        if mean_metrics:
            # Use baseline from lightResponse or fallback to hrvResults baseline
            result['light_baseline_bf'] = light_response.get('baseline_bf') or baseline.get('baseline_bf')
            
            # Direct extraction from mean_metrics
            result['light_avg_bf'] = mean_metrics.get('avg_bf')
            result['light_peak_bf'] = mean_metrics.get('peak_bf')
            result['light_peak_norm'] = mean_metrics.get('peak_norm_pct')
            result['light_ttp_avg'] = mean_metrics.get('time_to_peak_sec')
            result['light_recovery_bf'] = mean_metrics.get('bf_end')
            result['light_recovery_pct'] = mean_metrics.get('bf_end_pct')
            result['light_amplitude'] = mean_metrics.get('amplitude')
            result['light_roc'] = mean_metrics.get('rate_of_change')
            
            # Time to Peak for first stim (from per_stim if available)
            if per_stim and len(per_stim) > 0:
                result['light_ttp_first'] = per_stim[0].get('time_to_peak_sec')
        elif per_stim:
            # Fallback: compute from per_stim if mean_metrics not present
            valid_resp = [r for r in per_stim if r is not None and isinstance(r, dict)]
            if valid_resp:
                result['light_baseline_bf'] = light_response.get('baseline_bf') or baseline.get('baseline_bf')
                
                avg_bf_vals = [r.get('avg_bf') for r in valid_resp if r.get('avg_bf') is not None]
                result['light_avg_bf'] = float(np.mean(avg_bf_vals)) if avg_bf_vals else None
                
                peak_bf_vals = [r.get('peak_bf') for r in valid_resp if r.get('peak_bf') is not None]
                result['light_peak_bf'] = float(np.mean(peak_bf_vals)) if peak_bf_vals else None
                
                peak_norm_vals = [r.get('peak_norm_pct') for r in valid_resp if r.get('peak_norm_pct') is not None]
                result['light_peak_norm'] = float(np.mean(peak_norm_vals)) if peak_norm_vals else None
                
                result['light_ttp_first'] = valid_resp[0].get('time_to_peak_sec') if valid_resp else None
                ttp_vals = [r.get('time_to_peak_sec') for r in valid_resp if r.get('time_to_peak_sec') is not None]
                result['light_ttp_avg'] = float(np.mean(ttp_vals)) if ttp_vals else None
                
                recovery_bf_vals = [r.get('bf_end') for r in valid_resp if r.get('bf_end') is not None]
                result['light_recovery_bf'] = float(np.mean(recovery_bf_vals)) if recovery_bf_vals else None
                
                recovery_pct_vals = [r.get('bf_end_pct') for r in valid_resp if r.get('bf_end_pct') is not None]
                result['light_recovery_pct'] = float(np.mean(recovery_pct_vals)) if recovery_pct_vals else None
                
                amplitude_vals = [r.get('amplitude') for r in valid_resp if r.get('amplitude') is not None]
                result['light_amplitude'] = float(np.mean(amplitude_vals)) if amplitude_vals else None
                
                roc_vals = [r.get('rate_of_change') for r in valid_resp if r.get('rate_of_change') is not None]
                result['light_roc'] = float(np.mean(roc_vals)) if roc_vals else None
    
    # Corrected Light HRV (detrended) - frontend uses camelCase
    light_hrv_detrended = state.get('lightHrvDetrended') or state.get('light_metrics_detrended', {})
    
    result['light_hrv_ln_rmssd70'] = None
    result['light_hrv_ln_sdnn70'] = None
    result['light_hrv_pnn50'] = None
    
    if light_hrv_detrended and light_hrv_detrended.get('final'):
        final = light_hrv_detrended['final']
        result['light_hrv_ln_rmssd70'] = final.get('ln_rmssd70_detrended')
        result['light_hrv_ln_sdnn70'] = final.get('ln_sdnn70_detrended')
        result['light_hrv_pnn50'] = final.get('pnn50_detrended')
    
    return result


def compute_folder_averages(recordings_data: list, metrics: list) -> dict:
    """Compute folder averages for specified metrics, ignoring None values."""
    averages = {}
    counts = {}
    
    for metric in metrics:
        values = [r.get(metric) for r in recordings_data if r.get(metric) is not None]
        if values:
            averages[metric] = float(np.mean(values))
            counts[metric] = len(values)
        else:
            averages[metric] = None
            counts[metric] = 0
    
    return {'averages': averages, 'counts': counts}


@api_router.get("/folders/{folder_id}/comparison")
async def get_folder_comparison(folder_id: str):
    """Get comparison data for all recordings in a folder."""
    # Get folder info
    folder = await storage.get_folder(db, folder_id)
    if not folder:
        raise HTTPException(status_code=404, detail="Folder not found")
    
    # Get all recordings with full analysis_state (limit to 500 for performance)
    recordings_data = []
    async for rec in db.recordings.find({"folder_id": folder_id}).limit(500):
        recording = {
            "id": str(rec["_id"]),
            "name": rec["name"],
            "filename": rec["filename"],
            "analysis_state": rec.get("analysis_state", {}),
        }
        metrics = extract_comparison_metrics(recording)
        recordings_data.append(metrics)
    
    # Compute age ranges
    hspo_ages = [r['hspo_age'] for r in recordings_data if r.get('hspo_age') is not None]
    hco_ages = [r['hco_age'] for r in recordings_data if r.get('hco_age') is not None]
    fusion_ages = [r['fusion_age'] for r in recordings_data if r.get('fusion_age') is not None]
    
    # Compute averages for spontaneous activity
    spontaneous_metrics = [
        'baseline_bf', 'baseline_ln_rmssd70', 'baseline_ln_sdnn70', 'baseline_pnn50',
        'drug_bf', 'drug_ln_rmssd70', 'drug_ln_sdnn70', 'drug_pnn50'
    ]
    spontaneous_averages = compute_folder_averages(recordings_data, spontaneous_metrics)
    
    # Compute averages for Light HRA
    light_hra_metrics = [
        'light_baseline_bf', 'light_avg_bf', 'light_peak_bf', 'light_peak_norm',
        'light_ttp_first', 'light_ttp_avg', 'light_recovery_bf', 'light_recovery_pct',
        'light_amplitude', 'light_roc'
    ]
    light_hra_averages = compute_folder_averages(recordings_data, light_hra_metrics)
    
    # Compute averages for Corrected Light HRV
    light_hrv_metrics = ['light_hrv_ln_rmssd70', 'light_hrv_ln_sdnn70', 'light_hrv_pnn50']
    light_hrv_averages = compute_folder_averages(recordings_data, light_hrv_metrics)
    
    return {
        "folder": folder,
        "summary": {
            "recording_count": len(recordings_data),
            "hspo_age_range": {"min": min(hspo_ages) if hspo_ages else None, "max": max(hspo_ages) if hspo_ages else None, "n": len(hspo_ages)},
            "hco_age_range": {"min": min(hco_ages) if hco_ages else None, "max": max(hco_ages) if hco_ages else None, "n": len(hco_ages)},
            "fusion_age_range": {"min": min(fusion_ages) if fusion_ages else None, "max": max(fusion_ages) if fusion_ages else None, "n": len(fusion_ages)},
        },
        "recordings": recordings_data,
        "spontaneous_averages": spontaneous_averages,
        "light_hra_averages": light_hra_averages,
        "light_hrv_averages": light_hrv_averages,
    }


@api_router.post("/folders/{folder_id}/export/xlsx")
async def export_folder_comparison_xlsx(folder_id: str, request: FolderComparisonExportRequest):
    """Export folder comparison data to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    data = request.comparison_data
    recordings = data.get('recordings', [])
    summary = data.get('summary', {})
    
    wb = openpyxl.Workbook()
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill_baseline = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")  # Amber for baseline
    header_fill_drug = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")  # Purple for drug
    header_fill_light = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")  # Cyan for light
    avg_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    data_font = Font(size=9)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    
    def format_value(val, decimals=2):
        if val is None:
            return '—'
        if isinstance(val, (int, float)):
            if decimals == 0:
                return f"{val:.0f}"
            elif decimals == 1:
                return f"{val:.1f}"
            elif decimals == 3:
                return f"{val:.3f}"
            elif decimals == 4:
                return f"{val:.4f}"
            return f"{val:.{decimals}f}"
        return str(val) if val else '—'
    
    # Sheet 1: Folder Summary
    ws_summary = wb.active
    ws_summary.title = "Folder Summary"
    
    # Title styling
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    title_font = Font(bold=True, size=16, color="FFFFFF")
    section_font = Font(bold=True, size=11, color="374151")
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    
    # Title row
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = f'{request.folder_name} - Folder Comparison'
    ws_summary['A1'].font = title_font
    ws_summary['A1'].fill = title_fill
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 35
    
    # Section: Overview
    ws_summary['A3'] = 'OVERVIEW'
    ws_summary['A3'].font = section_font
    ws_summary.merge_cells('A3:D3')
    
    overview_data = [
        ('Recordings', summary.get('recording_count', 0), '', ''),
        ('Generated', datetime.now().strftime('%B %d, %Y at %H:%M'), '', ''),
    ]
    
    for row_idx, (label, value, _, _) in enumerate(overview_data, start=4):
        ws_summary[f'A{row_idx}'] = label
        ws_summary[f'B{row_idx}'] = value
        ws_summary[f'A{row_idx}'].font = label_font
        ws_summary[f'B{row_idx}'].font = value_font
        ws_summary[f'A{row_idx}'].border = thin_border
        ws_summary[f'B{row_idx}'].border = thin_border
    
    # Section: Age Ranges
    ws_summary['A7'] = 'AGE RANGES'
    ws_summary['A7'].font = section_font
    ws_summary.merge_cells('A7:D7')
    
    # Helper for age range formatting
    def format_age_range(age_dict):
        if not age_dict or age_dict.get('min') is None:
            return '—'
        return f"{age_dict.get('min')} - {age_dict.get('max')} days"
    
    age_headers = ['Type', 'Range', 'n']
    for col_idx, header in enumerate(age_headers, start=1):
        cell = ws_summary.cell(row=8, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    age_data = [
        ('hSpOs', format_age_range(summary.get('hspo_age_range')), summary.get('hspo_age_range', {}).get('n', 0)),
        ('hCOs', format_age_range(summary.get('hco_age_range')), summary.get('hco_age_range', {}).get('n', 0)),
        ('Fusion', format_age_range(summary.get('fusion_age_range')), summary.get('fusion_age_range', {}).get('n', 0)),
    ]
    
    for row_idx, (type_name, range_val, n_val) in enumerate(age_data, start=9):
        ws_summary.cell(row=row_idx, column=1, value=type_name).font = label_font
        ws_summary.cell(row=row_idx, column=2, value=range_val).font = value_font
        ws_summary.cell(row=row_idx, column=3, value=n_val).font = value_font
        for col in range(1, 4):
            ws_summary.cell(row=row_idx, column=col).border = thin_border
            ws_summary.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center' if col > 1 else 'left')
    
    # Section: Folder Averages (Spontaneous)
    spont_averages = data.get('spontaneous_averages', {}).get('averages', {})
    hra_averages = data.get('light_hra_averages', {}).get('averages', {})
    hrv_averages = data.get('light_hrv_averages', {}).get('averages', {})
    
    ws_summary['A13'] = 'FOLDER AVERAGES - SPONTANEOUS ACTIVITY'
    ws_summary['A13'].font = section_font
    ws_summary.merge_cells('A13:D13')
    
    avg_headers = ['Metric', 'Baseline', 'Drug']
    for col_idx, header in enumerate(avg_headers, start=1):
        cell = ws_summary.cell(row=14, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        if col_idx == 2:
            cell.fill = header_fill_baseline
        elif col_idx == 3:
            cell.fill = header_fill_drug
        else:
            cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    
    spont_avg_data = [
        ('Mean BF (bpm)', format_value(spont_averages.get('baseline_bf'), 1), format_value(spont_averages.get('drug_bf'), 1)),
        ('ln(RMSSD70)', format_value(spont_averages.get('baseline_ln_rmssd70'), 3), format_value(spont_averages.get('drug_ln_rmssd70'), 3)),
        ('ln(SDNN70)', format_value(spont_averages.get('baseline_ln_sdnn70'), 3), format_value(spont_averages.get('drug_ln_sdnn70'), 3)),
        ('pNN50 (%)', format_value(spont_averages.get('baseline_pnn50'), 1), format_value(spont_averages.get('drug_pnn50'), 1)),
    ]
    
    for row_idx, (metric, baseline, drug) in enumerate(spont_avg_data, start=15):
        ws_summary.cell(row=row_idx, column=1, value=metric).font = label_font
        ws_summary.cell(row=row_idx, column=2, value=baseline).font = value_font
        ws_summary.cell(row=row_idx, column=3, value=drug).font = value_font
        for col in range(1, 4):
            ws_summary.cell(row=row_idx, column=col).border = thin_border
            ws_summary.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center' if col > 1 else 'left')
    
    # Section: Light Metrics Averages (ALL metrics)
    ws_summary['A20'] = 'FOLDER AVERAGES - LIGHT STIMULUS'
    ws_summary['A20'].font = section_font
    ws_summary.merge_cells('A20:D20')
    
    # HRA Metrics subsection
    ws_summary['A21'] = 'Heart Rate Adaptation (HRA)'
    ws_summary['A21'].font = Font(bold=True, size=9, italic=True)
    ws_summary.merge_cells('A21:D21')
    
    hra_headers_sum = ['Metric', 'Value']
    for col_idx, header in enumerate(hra_headers_sum, start=1):
        cell = ws_summary.cell(row=22, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_light
        cell.border = thin_border
    
    hra_avg_data = [
        ('Baseline BF (bpm)', format_value(hra_averages.get('light_baseline_bf'), 1)),
        ('Avg BF (bpm)', format_value(hra_averages.get('light_avg_bf'), 1)),
        ('Peak BF (bpm)', format_value(hra_averages.get('light_peak_bf'), 1)),
        ('Normalized Peak (%)', format_value(hra_averages.get('light_peak_norm'), 1)),
        ('Time to Peak 1st (s)', format_value(hra_averages.get('light_ttp_first'), 1)),
        ('Time to Peak Avg (s)', format_value(hra_averages.get('light_ttp_avg'), 1)),
        ('Recovery BF (bpm)', format_value(hra_averages.get('light_recovery_bf'), 1)),
        ('Recovery (%)', format_value(hra_averages.get('light_recovery_pct'), 1)),
        ('Amplitude (bpm)', format_value(hra_averages.get('light_amplitude'), 1)),
        ('Rate of Change', format_value(hra_averages.get('light_roc'), 4)),
    ]
    
    for row_idx, (metric, val) in enumerate(hra_avg_data, start=23):
        ws_summary.cell(row=row_idx, column=1, value=metric).font = label_font
        ws_summary.cell(row=row_idx, column=2, value=val).font = value_font
        for col in range(1, 3):
            ws_summary.cell(row=row_idx, column=col).border = thin_border
    
    # Corrected HRV Metrics subsection
    hrv_start = 23 + len(hra_avg_data) + 1
    ws_summary.cell(row=hrv_start, column=1, value='Corrected HRV').font = Font(bold=True, size=9, italic=True)
    ws_summary.merge_cells(f'A{hrv_start}:D{hrv_start}')
    
    for col_idx, header in enumerate(hra_headers_sum, start=1):
        cell = ws_summary.cell(row=hrv_start + 1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_light
        cell.border = thin_border
    
    hrv_avg_data = [
        ('ln(RMSSD70) corr.', format_value(hrv_averages.get('light_hrv_ln_rmssd70'), 3)),
        ('ln(SDNN70) corr.', format_value(hrv_averages.get('light_hrv_ln_sdnn70'), 3)),
        ('pNN50 corr. (%)', format_value(hrv_averages.get('light_hrv_pnn50'), 1)),
    ]
    
    for row_idx, (metric, val) in enumerate(hrv_avg_data, start=hrv_start + 2):
        ws_summary.cell(row=row_idx, column=1, value=metric).font = label_font
        ws_summary.cell(row=row_idx, column=2, value=val).font = value_font
        for col in range(1, 3):
            ws_summary.cell(row=row_idx, column=col).border = thin_border
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    
    # Sheet 2: Spontaneous Activity Comparison
    ws_spont = wb.create_sheet("Spontaneous Activity")
    
    # Headers
    spont_headers = [
        ('Recording', None),
        ('Baseline BF (bpm)', 'baseline'),
        ('Baseline ln(RMSSD70)', 'baseline'),
        ('Baseline ln(SDNN70)', 'baseline'),
        ('Baseline pNN50 (%)', 'baseline'),
        ('Drug BF (bpm)', 'drug'),
        ('Drug ln(RMSSD70)', 'drug'),
        ('Drug ln(SDNN70)', 'drug'),
        ('Drug pNN50 (%)', 'drug'),
    ]
    
    for col_idx, (header, header_type) in enumerate(spont_headers, start=1):
        cell = ws_spont.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if header_type == 'baseline':
            cell.fill = header_fill_baseline
        elif header_type == 'drug':
            cell.fill = header_fill_drug
        else:
            cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    
    # Data rows
    spont_averages = data.get('spontaneous_averages', {}).get('averages', {})
    for row_idx, rec in enumerate(recordings, start=2):
        ws_spont.cell(row=row_idx, column=1, value=rec.get('name', '')).font = data_font
        ws_spont.cell(row=row_idx, column=2, value=format_value(rec.get('baseline_bf'), 1)).font = data_font
        ws_spont.cell(row=row_idx, column=3, value=format_value(rec.get('baseline_ln_rmssd70'), 3)).font = data_font
        ws_spont.cell(row=row_idx, column=4, value=format_value(rec.get('baseline_ln_sdnn70'), 3)).font = data_font
        ws_spont.cell(row=row_idx, column=5, value=format_value(rec.get('baseline_pnn50'), 1)).font = data_font
        ws_spont.cell(row=row_idx, column=6, value=format_value(rec.get('drug_bf'), 1)).font = data_font
        ws_spont.cell(row=row_idx, column=7, value=format_value(rec.get('drug_ln_rmssd70'), 3)).font = data_font
        ws_spont.cell(row=row_idx, column=8, value=format_value(rec.get('drug_ln_sdnn70'), 3)).font = data_font
        ws_spont.cell(row=row_idx, column=9, value=format_value(rec.get('drug_pnn50'), 1)).font = data_font
        
        for col in range(1, 10):
            ws_spont.cell(row=row_idx, column=col).border = thin_border
            # Color code baseline vs drug columns
            if 2 <= col <= 5:
                ws_spont.cell(row=row_idx, column=col).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif 6 <= col <= 9:
                ws_spont.cell(row=row_idx, column=col).fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    
    # Average row
    avg_row = len(recordings) + 2
    ws_spont.cell(row=avg_row, column=1, value=f"Folder Average (n={len(recordings)})").font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=2, value=format_value(spont_averages.get('baseline_bf'), 1)).font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=3, value=format_value(spont_averages.get('baseline_ln_rmssd70'), 3)).font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=4, value=format_value(spont_averages.get('baseline_ln_sdnn70'), 3)).font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=5, value=format_value(spont_averages.get('baseline_pnn50'), 1)).font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=6, value=format_value(spont_averages.get('drug_bf'), 1)).font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=7, value=format_value(spont_averages.get('drug_ln_rmssd70'), 3)).font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=8, value=format_value(spont_averages.get('drug_ln_sdnn70'), 3)).font = Font(bold=True, size=9)
    ws_spont.cell(row=avg_row, column=9, value=format_value(spont_averages.get('drug_pnn50'), 1)).font = Font(bold=True, size=9)
    
    for col in range(1, 10):
        ws_spont.cell(row=avg_row, column=col).fill = avg_fill
        ws_spont.cell(row=avg_row, column=col).border = thin_border
    
    # Add Normalized to Baseline section below the main table
    norm_start_row = avg_row + 3
    
    # Calculate cohort baseline averages for normalization
    baseline_bfs = [r.get('baseline_bf') for r in recordings if r.get('baseline_bf') is not None]
    baseline_ln_rmssds = [r.get('baseline_ln_rmssd70') for r in recordings if r.get('baseline_ln_rmssd70') is not None]
    baseline_ln_sdnns = [r.get('baseline_ln_sdnn70') for r in recordings if r.get('baseline_ln_sdnn70') is not None]
    baseline_pnn50s = [r.get('baseline_pnn50') for r in recordings if r.get('baseline_pnn50') is not None]
    
    avg_baseline_bf = sum(baseline_bfs) / len(baseline_bfs) if baseline_bfs else 1
    avg_baseline_ln_rmssd = sum(baseline_ln_rmssds) / len(baseline_ln_rmssds) if baseline_ln_rmssds else 1
    avg_baseline_ln_sdnn = sum(baseline_ln_sdnns) / len(baseline_ln_sdnns) if baseline_ln_sdnns else 1
    avg_baseline_pnn50 = sum(baseline_pnn50s) / len(baseline_pnn50s) if baseline_pnn50s else 1
    
    def norm_val(val, avg):
        if val is None or avg == 0:
            return None
        return 100 * val / avg
    
    # Section title
    ws_spont.merge_cells(f'A{norm_start_row}:I{norm_start_row}')
    ws_spont.cell(row=norm_start_row, column=1, value='Normalized to Baseline — Spontaneous Activity').font = Font(bold=True, size=11, color="FFFFFF")
    ws_spont.cell(row=norm_start_row, column=1).fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    ws_spont.cell(row=norm_start_row, column=1).alignment = Alignment(horizontal='center')
    
    norm_spont_headers = [
        ('Recording', None),
        ('Baseline BF (%)', 'baseline'),
        ('Baseline ln(RMSSD) (%)', 'baseline'),
        ('Baseline ln(SDNN) (%)', 'baseline'),
        ('Baseline pNN50 (%)', 'baseline'),
        ('Drug BF (%)', 'drug'),
        ('Drug ln(RMSSD) (%)', 'drug'),
        ('Drug ln(SDNN) (%)', 'drug'),
        ('Drug pNN50 (%)', 'drug'),
    ]
    
    for col_idx, (header, header_type) in enumerate(norm_spont_headers, start=1):
        cell = ws_spont.cell(row=norm_start_row + 1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if header_type == 'baseline':
            cell.fill = header_fill_baseline
        elif header_type == 'drug':
            cell.fill = header_fill_drug
        else:
            cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    
    # Data rows with normalized values
    norm_sums = {'baseline_bf': [], 'baseline_ln_rmssd': [], 'baseline_ln_sdnn': [], 'baseline_pnn50': [],
                 'drug_bf': [], 'drug_ln_rmssd': [], 'drug_ln_sdnn': [], 'drug_pnn50': []}
    
    for row_idx, rec in enumerate(recordings, start=norm_start_row + 2):
        ws_spont.cell(row=row_idx, column=1, value=rec.get('name', '')).font = data_font
        
        n_baseline_bf = norm_val(rec.get('baseline_bf'), avg_baseline_bf)
        n_baseline_ln_rmssd = norm_val(rec.get('baseline_ln_rmssd70'), avg_baseline_ln_rmssd)
        n_baseline_ln_sdnn = norm_val(rec.get('baseline_ln_sdnn70'), avg_baseline_ln_sdnn)
        n_baseline_pnn50 = norm_val(rec.get('baseline_pnn50'), avg_baseline_pnn50)
        n_drug_bf = norm_val(rec.get('drug_bf'), avg_baseline_bf)
        n_drug_ln_rmssd = norm_val(rec.get('drug_ln_rmssd70'), avg_baseline_ln_rmssd)
        n_drug_ln_sdnn = norm_val(rec.get('drug_ln_sdnn70'), avg_baseline_ln_sdnn)
        n_drug_pnn50 = norm_val(rec.get('drug_pnn50'), avg_baseline_pnn50)
        
        ws_spont.cell(row=row_idx, column=2, value=format_value(n_baseline_bf, 1)).font = data_font
        ws_spont.cell(row=row_idx, column=3, value=format_value(n_baseline_ln_rmssd, 1)).font = data_font
        ws_spont.cell(row=row_idx, column=4, value=format_value(n_baseline_ln_sdnn, 1)).font = data_font
        ws_spont.cell(row=row_idx, column=5, value=format_value(n_baseline_pnn50, 1)).font = data_font
        ws_spont.cell(row=row_idx, column=6, value=format_value(n_drug_bf, 1)).font = data_font
        ws_spont.cell(row=row_idx, column=7, value=format_value(n_drug_ln_rmssd, 1)).font = data_font
        ws_spont.cell(row=row_idx, column=8, value=format_value(n_drug_ln_sdnn, 1)).font = data_font
        ws_spont.cell(row=row_idx, column=9, value=format_value(n_drug_pnn50, 1)).font = data_font
        
        # Collect for averages
        for key, val in [('baseline_bf', n_baseline_bf), ('baseline_ln_rmssd', n_baseline_ln_rmssd),
                         ('baseline_ln_sdnn', n_baseline_ln_sdnn), ('baseline_pnn50', n_baseline_pnn50),
                         ('drug_bf', n_drug_bf), ('drug_ln_rmssd', n_drug_ln_rmssd),
                         ('drug_ln_sdnn', n_drug_ln_sdnn), ('drug_pnn50', n_drug_pnn50)]:
            if val is not None:
                norm_sums[key].append(val)
        
        for col in range(1, 10):
            ws_spont.cell(row=row_idx, column=col).border = thin_border
            if 2 <= col <= 5:
                ws_spont.cell(row=row_idx, column=col).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif 6 <= col <= 9:
                ws_spont.cell(row=row_idx, column=col).fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    
    # Folder Average row for normalized section
    norm_avg_row = norm_start_row + 2 + len(recordings)
    ws_spont.cell(row=norm_avg_row, column=1, value=f"Folder Average (n={len(recordings)})").font = Font(bold=True, size=9)
    
    for col_idx, key in enumerate(['baseline_bf', 'baseline_ln_rmssd', 'baseline_ln_sdnn', 'baseline_pnn50',
                                   'drug_bf', 'drug_ln_rmssd', 'drug_ln_sdnn', 'drug_pnn50'], start=2):
        avg_val_norm = sum(norm_sums[key]) / len(norm_sums[key]) if norm_sums[key] else None
        ws_spont.cell(row=norm_avg_row, column=col_idx, value=format_value(avg_val_norm, 1)).font = Font(bold=True, size=9)
    
    for col in range(1, 10):
        ws_spont.cell(row=norm_avg_row, column=col).fill = avg_fill
        ws_spont.cell(row=norm_avg_row, column=col).border = thin_border
    
    auto_width(ws_spont)
    
    # Sheet 3: Light Stimulus (Combined HRA and Corrected HRV)
    ws_light = wb.create_sheet("Light Stimulus")
    
    # HRA Section Title
    ws_light.merge_cells('A1:K1')
    ws_light['A1'] = 'Light-Induced Heart Rate Adaptation (HRA)'
    ws_light['A1'].font = Font(bold=True, size=11)
    ws_light['A1'].fill = header_fill_light
    ws_light['A1'].alignment = Alignment(horizontal='center')
    
    hra_headers = [
        'Recording', 'Baseline BF (bpm)', 'Avg BF (bpm)', 'Peak BF (bpm)', 
        'Norm. Peak (%)', 'TTP 1st (s)', 'TTP Avg (s)', 
        'Recovery BF (bpm)', 'Recovery (%)', 'Amplitude (bpm)', 'Rate of Change'
    ]
    
    for col_idx, header in enumerate(hra_headers, start=1):
        cell = ws_light.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_light
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    hra_averages = data.get('light_hra_averages', {}).get('averages', {})
    for row_idx, rec in enumerate(recordings, start=3):
        ws_light.cell(row=row_idx, column=1, value=rec.get('name', '')).font = data_font
        ws_light.cell(row=row_idx, column=2, value=format_value(rec.get('light_baseline_bf'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=3, value=format_value(rec.get('light_avg_bf'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=4, value=format_value(rec.get('light_peak_bf'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=5, value=format_value(rec.get('light_peak_norm'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=6, value=format_value(rec.get('light_ttp_first'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=7, value=format_value(rec.get('light_ttp_avg'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=8, value=format_value(rec.get('light_recovery_bf'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=9, value=format_value(rec.get('light_recovery_pct'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=10, value=format_value(rec.get('light_amplitude'), 1)).font = data_font
        ws_light.cell(row=row_idx, column=11, value=format_value(rec.get('light_roc'), 4)).font = data_font
        
        for col in range(1, 12):
            ws_light.cell(row=row_idx, column=col).border = thin_border
    
    # HRA Average row
    hra_avg_row = len(recordings) + 3
    ws_light.cell(row=hra_avg_row, column=1, value=f"Folder Average (n={len(recordings)})").font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=2, value=format_value(hra_averages.get('light_baseline_bf'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=3, value=format_value(hra_averages.get('light_avg_bf'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=4, value=format_value(hra_averages.get('light_peak_bf'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=5, value=format_value(hra_averages.get('light_peak_norm'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=6, value=format_value(hra_averages.get('light_ttp_first'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=7, value=format_value(hra_averages.get('light_ttp_avg'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=8, value=format_value(hra_averages.get('light_recovery_bf'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=9, value=format_value(hra_averages.get('light_recovery_pct'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=10, value=format_value(hra_averages.get('light_amplitude'), 1)).font = Font(bold=True, size=9)
    ws_light.cell(row=hra_avg_row, column=11, value=format_value(hra_averages.get('light_roc'), 4)).font = Font(bold=True, size=9)
    
    for col in range(1, 12):
        ws_light.cell(row=hra_avg_row, column=col).fill = avg_fill
        ws_light.cell(row=hra_avg_row, column=col).border = thin_border
    
    # Add Normalized to Baseline - Light HRA section below the main HRA table
    norm_hra_start_row = hra_avg_row + 3
    
    ws_light.merge_cells(f'A{norm_hra_start_row}:E{norm_hra_start_row}')
    ws_light.cell(row=norm_hra_start_row, column=1, value='Normalized to Baseline — Light-Induced Heart Rate Adaptation (HRA)').font = Font(bold=True, size=11, color="FFFFFF")
    ws_light.cell(row=norm_hra_start_row, column=1).fill = header_fill_light
    ws_light.cell(row=norm_hra_start_row, column=1).alignment = Alignment(horizontal='center')
    
    norm_light_headers = ['Recording', 'Baseline BF (%)', 'Avg BF (%)', 'Peak BF (%)', 'Recovery BF (%)']
    
    for col_idx, header in enumerate(norm_light_headers, start=1):
        cell = ws_light.cell(row=norm_hra_start_row + 1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_light
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Calculate the average of light_baseline_bf for Light HRA normalization
    light_baseline_bfs = [r.get('light_baseline_bf') for r in recordings if r.get('light_baseline_bf') is not None]
    avg_light_baseline_bf = sum(light_baseline_bfs) / len(light_baseline_bfs) if light_baseline_bfs else 1
    
    # Data rows with normalized values (using avg_light_baseline_bf - NOT spontaneous baseline)
    light_norm_sums = {'baseline_bf': [], 'avg_bf': [], 'peak_bf': [], 'recovery_bf': []}
    
    for row_idx, rec in enumerate(recordings, start=norm_hra_start_row + 2):
        ws_light.cell(row=row_idx, column=1, value=rec.get('name', '')).font = data_font
        
        n_light_baseline_bf = norm_val(rec.get('light_baseline_bf'), avg_light_baseline_bf)
        n_light_avg_bf = norm_val(rec.get('light_avg_bf'), avg_light_baseline_bf)
        n_light_peak_bf = norm_val(rec.get('light_peak_bf'), avg_light_baseline_bf)
        n_light_recovery_bf = norm_val(rec.get('light_recovery_bf'), avg_light_baseline_bf)
        
        ws_light.cell(row=row_idx, column=2, value=format_value(n_light_baseline_bf, 1)).font = data_font
        ws_light.cell(row=row_idx, column=3, value=format_value(n_light_avg_bf, 1)).font = data_font
        ws_light.cell(row=row_idx, column=4, value=format_value(n_light_peak_bf, 1)).font = data_font
        ws_light.cell(row=row_idx, column=5, value=format_value(n_light_recovery_bf, 1)).font = data_font
        
        # Collect for averages
        for key, val in [('baseline_bf', n_light_baseline_bf), ('avg_bf', n_light_avg_bf),
                         ('peak_bf', n_light_peak_bf), ('recovery_bf', n_light_recovery_bf)]:
            if val is not None:
                light_norm_sums[key].append(val)
        
        for col in range(1, 6):
            ws_light.cell(row=row_idx, column=col).border = thin_border
    
    # Folder Average row for normalized Light HRA
    light_norm_avg_row = norm_hra_start_row + 2 + len(recordings)
    ws_light.cell(row=light_norm_avg_row, column=1, value=f"Folder Average (n={len(recordings)})").font = Font(bold=True, size=9)
    
    for col_idx, key in enumerate(['baseline_bf', 'avg_bf', 'peak_bf', 'recovery_bf'], start=2):
        avg_val_light = sum(light_norm_sums[key]) / len(light_norm_sums[key]) if light_norm_sums[key] else None
        ws_light.cell(row=light_norm_avg_row, column=col_idx, value=format_value(avg_val_light, 1)).font = Font(bold=True, size=9)
    
    for col in range(1, 6):
        ws_light.cell(row=light_norm_avg_row, column=col).fill = avg_fill
        ws_light.cell(row=light_norm_avg_row, column=col).border = thin_border
    
    # Spacer row - adjust for Corrected HRV section
    hrv_start_row = light_norm_avg_row + 3
    
    # Corrected HRV Section Title
    ws_light.merge_cells(f'A{hrv_start_row}:D{hrv_start_row}')
    ws_light.cell(row=hrv_start_row, column=1, value='Corrected Light-Induced Heart Rate Variability (HRV)').font = Font(bold=True, size=11)
    ws_light.cell(row=hrv_start_row, column=1).fill = header_fill_light
    ws_light.cell(row=hrv_start_row, column=1).alignment = Alignment(horizontal='center')
    
    hrv_headers = ['Recording', 'ln(RMSSD70) corr.', 'ln(SDNN70) corr.', 'pNN50 corr. (%)']
    
    for col_idx, header in enumerate(hrv_headers, start=1):
        cell = ws_light.cell(row=hrv_start_row + 1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_light
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    hrv_averages = data.get('light_hrv_averages', {}).get('averages', {})
    for row_idx, rec in enumerate(recordings, start=hrv_start_row + 2):
        actual_rec_idx = row_idx - (hrv_start_row + 2)
        if actual_rec_idx < len(recordings):
            rec = recordings[actual_rec_idx]
            ws_light.cell(row=row_idx, column=1, value=rec.get('name', '')).font = data_font
            ws_light.cell(row=row_idx, column=2, value=format_value(rec.get('light_hrv_ln_rmssd70'), 3)).font = data_font
            ws_light.cell(row=row_idx, column=3, value=format_value(rec.get('light_hrv_ln_sdnn70'), 3)).font = data_font
            ws_light.cell(row=row_idx, column=4, value=format_value(rec.get('light_hrv_pnn50'), 1)).font = data_font
            
            for col in range(1, 5):
                ws_light.cell(row=row_idx, column=col).border = thin_border
    
    # HRV Average row
    hrv_avg_row = hrv_start_row + 2 + len(recordings)
    ws_light.cell(row=hrv_avg_row, column=1, value=f"Folder Average (n={len(recordings)})").font = Font(bold=True, size=9)
    ws_light.cell(row=hrv_avg_row, column=2, value=format_value(hrv_averages.get('light_hrv_ln_rmssd70'), 3)).font = Font(bold=True, size=9)
    ws_light.cell(row=hrv_avg_row, column=3, value=format_value(hrv_averages.get('light_hrv_ln_sdnn70'), 3)).font = Font(bold=True, size=9)
    ws_light.cell(row=hrv_avg_row, column=4, value=format_value(hrv_averages.get('light_hrv_pnn50'), 1)).font = Font(bold=True, size=9)
    
    for col in range(1, 5):
        ws_light.cell(row=hrv_avg_row, column=col).fill = avg_fill
        ws_light.cell(row=hrv_avg_row, column=col).border = thin_border
    
    auto_width(ws_light)
    
    # Sheet 4: Recording Metadata
    ws_meta = wb.create_sheet("Recording Metadata")
    
    meta_headers = [
        'Recording', 'Filename', 'Date', 
        'hSpO Line', 'hSpO P#', 'hSpO Age', 'hSpO Transd.',
        'hCO Line', 'hCO P#', 'hCO Age',
        'Fusion Date', 'Condition',
        'Drug(s)', 'Concentration(s)', 'Perfusion Time(s)',
        'Stim Duration', 'ISI Structure', 'Notes'
    ]
    
    for col_idx, header in enumerate(meta_headers, start=1):
        cell = ws_meta.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        # Color code headers
        if 4 <= col_idx <= 7:  # hSpO columns
            cell.fill = header_fill_baseline  # Amber
        elif 8 <= col_idx <= 10:  # hCO columns
            cell.fill = header_fill_drug  # Purple
        elif 13 <= col_idx <= 15:  # Drug columns
            cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")  # Green
        elif 16 <= col_idx <= 17:  # Light columns
            cell.fill = header_fill_light  # Cyan
        else:
            cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    
    for row_idx, rec in enumerate(recordings, start=2):
        ws_meta.cell(row=row_idx, column=1, value=rec.get('name', '')).font = data_font
        ws_meta.cell(row=row_idx, column=2, value=rec.get('filename', '')).font = data_font
        ws_meta.cell(row=row_idx, column=3, value=rec.get('recording_date', '')).font = data_font
        
        # hSpO info
        hspo = rec.get('hspo_info') or {}
        ws_meta.cell(row=row_idx, column=4, value=hspo.get('line_name', '') if hspo else '').font = data_font
        ws_meta.cell(row=row_idx, column=5, value=hspo.get('passage', '') if hspo else '').font = data_font
        ws_meta.cell(row=row_idx, column=6, value=format_value(hspo.get('age'), 0) if hspo and hspo.get('age') else '').font = data_font
        ws_meta.cell(row=row_idx, column=7, value='Yes' if hspo and hspo.get('has_transduction') else '').font = data_font
        
        # hCO info
        hco = rec.get('hco_info') or {}
        ws_meta.cell(row=row_idx, column=8, value=hco.get('line_name', '') if hco else '').font = data_font
        ws_meta.cell(row=row_idx, column=9, value=hco.get('passage', '') if hco else '').font = data_font
        ws_meta.cell(row=row_idx, column=10, value=format_value(hco.get('age'), 0) if hco and hco.get('age') else '').font = data_font
        
        ws_meta.cell(row=row_idx, column=11, value=rec.get('fusion_date', '')).font = data_font
        ws_meta.cell(row=row_idx, column=12, value=rec.get('condition', '')).font = data_font
        
        # Drug info - use dashes if no drug
        drug_info = rec.get('drug_info', [])
        if rec.get('has_drug') and drug_info:
            drug_names = ', '.join([d.get('name', '') for d in drug_info]) if drug_info else ''
            drug_concs = ', '.join([str(d.get('concentration', '')) for d in drug_info if d.get('concentration')]) if drug_info else ''
            drug_times = ', '.join([str(d.get('perfusion_time', '')) for d in drug_info if d.get('perfusion_time')]) if drug_info else ''
        else:
            drug_names = '—'
            drug_concs = '—'
            drug_times = '—'
        ws_meta.cell(row=row_idx, column=13, value=drug_names).font = data_font
        ws_meta.cell(row=row_idx, column=14, value=drug_concs).font = data_font
        ws_meta.cell(row=row_idx, column=15, value=drug_times).font = data_font
        
        # Light stim info - use dashes if no light stim
        if rec.get('has_light_stim'):
            ws_meta.cell(row=row_idx, column=16, value=f"{rec.get('stim_duration', '')}s").font = data_font
            ws_meta.cell(row=row_idx, column=17, value=rec.get('isi_structure', '') or '—').font = data_font
        else:
            ws_meta.cell(row=row_idx, column=16, value='—').font = data_font
            ws_meta.cell(row=row_idx, column=17, value='—').font = data_font
        ws_meta.cell(row=row_idx, column=18, value=rec.get('recording_description', '') or '—').font = data_font
        
        for col in range(1, 19):
            ws_meta.cell(row=row_idx, column=col).border = thin_border
    
    auto_width(ws_meta)
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{request.folder_name}_comparison.xlsx".replace(' ', '_')
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.post("/folders/{folder_id}/export/pdf")
async def export_folder_comparison_pdf(folder_id: str, request: FolderComparisonExportRequest):
    """Export folder comparison data to PDF using the new Nature-style format."""
    output = export_utils.create_comparison_pdf(request.folder_name, request.comparison_data)
    filename = f"{request.folder_name}_comparison.pdf".replace(' ', '_')
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/export/csv")
async def export_csv(request: ExportRequest):
    """Clean Nature-style CSV export"""
    buf = export_utils.create_nature_csv(request)
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={request.filename}.csv"}
    )


@api_router.post("/export/xlsx")
async def export_xlsx(request: ExportRequest):
    """Clean Nature-style Excel export"""
    buf = export_utils.create_nature_excel(request)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={request.filename}.xlsx"}
    )


@api_router.post("/export/pdf")
async def export_pdf(request: ExportRequest):
    """Clean Nature-style PDF export"""
    buf = export_utils.create_nature_pdf(request)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={request.filename}.pdf"}
    )

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
