"""
NEHER Export Utilities - Nature Magazine Style
Clean, professional scientific publication exports
Developed by Kolia H. Badarello
"""
import io
import numpy as np
from datetime import datetime

# Color palette matching NEHER UI
COLORS = {
    'emerald': '#10b981',      # Beats/BF
    'amber': '#f59e0b',        # Light stims
    'purple': '#a855f7',       # Drug perfusion
    'violet_dark': '#7c3aed',  # Drug (darker)
    'silver': '#71717a',       # Intervals/secondary
    'sky': '#0ea5e9',          # Baseline
    'zinc_dark': '#18181b',    # Dark background
    'zinc_light': '#f4f4f5',   # Light background
    'white': '#ffffff',
}

TINTS = {
    'baseline': '#E0F2FE',     # Light blue
    'drug': '#F3E8FF',         # Light purple
    'light': '#FEF3C7',        # Light amber
}


def add_page_footer(fig, page_num, total_pages=None):
    """Add footer to each page"""
    footer_y = 0.02
    fig.text(0.5, footer_y, 'NEHER', ha='center', fontsize=8, fontweight='bold', color='#18181b')
    fig.text(0.95, footer_y, f'Page {page_num}' if not total_pages else f'Page {page_num}/{total_pages}', 
             ha='right', fontsize=8, color='#a1a1aa')
    fig.text(0.05, footer_y, 'Developed by Kolia H. Badarello', ha='left', fontsize=7, color='#a1a1aa', style='italic')


def create_nature_pdf(request):
    """Create a professional PDF export with Optima/Carlito typography - bioptima style"""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    import matplotlib.patches as mpatches
    import matplotlib.font_manager as fm
    
    # Register fonts - Optima for titles, Carlito for body text
    # Fallback to similar fonts if not available
    title_font = 'Optima'
    body_font = 'Carlito'
    
    # Check available fonts and set fallbacks
    available_fonts = [f.name for f in fm.fontManager.ttflist]
    if title_font not in available_fonts:
        title_font = 'DejaVu Sans'  # Similar clean sans-serif
    if body_font not in available_fonts:
        body_font = 'DejaVu Sans'
    
    plt.rcParams.update({
        'font.family': 'sans-serif',
        'font.sans-serif': [body_font, 'Carlito', 'DejaVu Sans', 'Arial'],
        'font.size': 9,
        'axes.labelsize': 9,
        'axes.titlesize': 10,
        'axes.linewidth': 0.8,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'legend.fontsize': 8,
        'figure.titlesize': 12,
        'axes.spines.top': False,
        'axes.spines.right': False,
    })
    
    buf = io.BytesIO()
    page_num = 0
    total_pages = 1  # Will be updated
    
    # Color scheme matching the bioptima reference design
    COLORS = {
        'dark': '#18181b',          # Near black for text
        'header_blue': '#5ba4c9',   # Teal/cyan for table headers (like bioptima)
        'category_blue': '#aed4e6', # Light blue for category column
        'row_blue': '#d5eaf4',      # Very light blue for alternating rows
        'total_red': '#c7254e',     # Red/maroon for total rows
        'purple': '#9333ea',        # Purple for drug sections
        'light_purple': '#e8d5f5',  # Light purple for drug rows
        'amber': '#f59e0b',         # Amber for light sections
        'light_amber': '#fef3c7',   # Light amber tint
        'gray': '#6b7280',          # Gray for secondary text
        'light_gray': '#f3f4f6',    # Light gray background
        'line': '#374151',          # Dark line color
        'emerald': '#10b981',
        'sky': '#0ea5e9',
        'baseline_blue': '#E0F2FE', # Light blue for baseline
    }
    
    # Tints for row backgrounds
    TINTS = {
        'baseline': '#E0F2FE',     # Light blue
        'drug': '#e8d5f5',         # Light purple
        'light': '#FEF3C7',        # Light amber
    }
    
    # Calculate total pages
    has_bf_chart = request.per_beat_data or request.per_minute_data
    has_hrv_charts = request.hrv_windows
    has_light_hrv = request.light_enabled and request.light_metrics_detrended
    has_spont_bf_table = request.per_minute_data
    has_spont_hrv_table = request.hrv_windows
    has_light_hra_table = request.light_enabled and request.light_response
    has_light_hrv_table = request.light_enabled and request.light_metrics_detrended
    
    total_pages = 1  # Summary always
    if has_bf_chart: total_pages += 1
    if has_hrv_charts: total_pages += 1
    if has_light_hrv: total_pages += 1
    if has_spont_bf_table: total_pages += 1
    if has_spont_hrv_table: total_pages += 1
    if has_light_hra_table: total_pages += 1
    if has_light_hrv_table: total_pages += 1
    
    def add_page_header(fig, section_name):
        """Add header in the bioptima style: NEHER section_name"""
        fig.text(0.08, 0.96, 'NEHER', fontsize=12, fontweight='bold', color=COLORS['dark'],
                fontfamily=title_font)
        fig.text(0.16, 0.96, section_name, fontsize=12, fontweight='normal', color=COLORS['gray'],
                fontfamily=title_font)
    
    def add_page_footer(fig, page_num, total_pages=None):
        """Add footer: p. XX | Recording Name | Electrophysiology Analysis Report by NEHER"""
        recording_name = request.recording_name or request.filename or 'Recording'
        fig.text(0.08, 0.025, f'p. {page_num}', fontsize=10, fontweight='bold', color=COLORS['dark'],
                fontfamily=body_font)
        fig.text(0.125, 0.025, f'|  {recording_name}  |  Electrophysiology Analysis Report by NEHER', fontsize=10, color=COLORS['gray'],
                fontfamily=body_font)
    
    def draw_header(fig, x, y, text, color, width=0.38):
        """Draw section header with colored background - bioptima style"""
        # Draw rectangle header
        fig.add_artist(mpatches.Rectangle(
            (x, y - 0.012), width, 0.024,
            facecolor=color, edgecolor='none', transform=fig.transFigure
        ))
        fig.text(x + 0.01, y, text, fontsize=9, fontweight='bold', color='white',
                fontfamily=body_font, va='center')
        return y - 0.028
    
    def draw_separator(fig, x, y, width=0.38, centered=False):
        """Draw horizontal separator line"""
        end_x = x + width
        fig.add_artist(plt.Line2D([x, end_x], [y, y], color=COLORS['line'], 
                      linewidth=0.5, transform=fig.transFigure))
        return y - 0.008
    
    def draw_row(fig, x, y, label, value, bg_color=None, width=0.38, label_width=0.18):
        """Draw a data row with label and value"""
        if bg_color:
            fig.add_artist(mpatches.Rectangle(
                (x, y - 0.008), width, 0.020,
                facecolor=bg_color, edgecolor='none', transform=fig.transFigure
            ))
        fig.text(x + 0.01, y, label, fontsize=8, color=COLORS['gray'], fontfamily=body_font, va='center')
        fig.text(x + label_width, y, str(value), fontsize=8, fontweight='bold', color=COLORS['dark'],
                fontfamily=body_font, va='center')
        return y - 0.020
    
    with PdfPages(buf) as pdf:
        
        # ==================== PAGE 1: SUMMARY ====================
        page_num += 1
        fig1 = plt.figure(figsize=(8.5, 11))
        fig1.patch.set_facecolor('white')
        
        # Header
        add_page_header(fig1, 'summary')
        
        # Main title (large, bold, Optima style - like bioptima) - moved 0.2cm lower
        title = request.recording_name or request.filename or 'Recording Analysis'
        fig1.text(0.08, 0.90, title, ha='left', va='top', fontsize=28, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        
        # Separator line below title (thin black line) - moved 0.2cm lower
        fig1.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig1.transFigure))
        
        # Layout: two columns spanning 0.08 to 0.92 with 1cm (≈0.04) gap between them
        # Total width: 0.84, gap: 0.04, each column: 0.40
        left_x = 0.08
        right_x = 0.52  # 0.08 + 0.40 + 0.04
        col_width = 0.40  # Each column is 0.40 wide (right column ends at 0.92)
        
        # Spacing from title bar - moved 0.2cm lower
        first_section_y = 0.83
        section_gap = 0.025  # Gap between sections
        
        # LEFT COLUMN
        y = draw_header(fig1, left_x, first_section_y, 'RECORDING INFO', COLORS['dark'], width=col_width)
        
        if request.original_filename:
            y = draw_row(fig1, left_x, y, 'Original File:', request.original_filename, width=col_width)
        if request.recording_date:
            y = draw_row(fig1, left_x, y, 'Recording Date:', request.recording_date, width=col_width)
        if request.summary:
            if 'Total Beats' in request.summary:
                y = draw_row(fig1, left_x, y, 'Total Beats:', request.summary['Total Beats'], width=col_width)
            if 'Kept Beats' in request.summary:
                y = draw_row(fig1, left_x, y, 'Kept Beats:', request.summary['Kept Beats'], width=col_width)
            if 'Filter Range' in request.summary:
                y = draw_row(fig1, left_x, y, 'Filter Range:', request.summary['Filter Range'], width=col_width)
        
        # TISSUE INFO (below Recording Info)
        if request.organoid_info:
            y -= section_gap
            y = draw_header(fig1, left_x, y, 'TISSUE INFO', '#6b7280', width=col_width)
            num_organoids = len(request.organoid_info)
            for idx, org in enumerate(request.organoid_info):
                if org.get('cell_type'):
                    cell_type = org.get('other_cell_type') if org.get('cell_type') == 'Other' else org.get('cell_type')
                    # Number the cell types: Cell Type 1, Cell Type 2, etc.
                    cell_type_label = f'Cell Type {idx + 1}:' if num_organoids > 1 else 'Cell Type:'
                    y = draw_row(fig1, left_x, y, cell_type_label, cell_type or '—', width=col_width)
                if org.get('line_name'):
                    y = draw_row(fig1, left_x, y, 'Line:', org.get('line_name'), width=col_width)
                if org.get('passage_number'):
                    y = draw_row(fig1, left_x, y, 'Passage:', org.get('passage_number'), width=col_width)
                if org.get('age_at_recording') is not None:
                    y = draw_row(fig1, left_x, y, 'Age at Recording:', f"{org.get('age_at_recording')} days", width=col_width)
                if org.get('transfection'):
                    trans = org['transfection']
                    if trans.get('name'):
                        y = draw_row(fig1, left_x, y, 'Transfection:', trans.get('name'), width=col_width)
                    if trans.get('days_since_transfection') is not None:
                        y = draw_row(fig1, left_x, y, 'Days Post-Transf.:', trans.get('days_since_transfection'), width=col_width)
                
                # Add separator line between samples (0.2cm = ~0.007 spacing above and below)
                if idx < num_organoids - 1:
                    y -= 0.007  # 0.2cm above line
                    y = draw_separator(fig1, left_x, y, width=col_width)
                    y -= 0.007  # 0.2cm below line
        
        # Add separator line before Days Since Fusion (if there are organoid samples)
        if request.days_since_fusion is not None:
            if request.organoid_info and len(request.organoid_info) > 0:
                y -= 0.007  # 0.2cm above line
                y = draw_separator(fig1, left_x, y, width=col_width)
                y -= 0.007  # 0.2cm below line
            y = draw_row(fig1, left_x, y, 'Days Since Fusion:', request.days_since_fusion, width=col_width)
        
        # DRUG PERFUSION
        if request.all_drugs and len(request.all_drugs) > 0:
            y -= section_gap
            y = draw_header(fig1, left_x, y, 'DRUG PERFUSION', COLORS['purple'], width=col_width)
            for drug in request.all_drugs:
                y = draw_row(fig1, left_x, y, 'Drug:', drug.get('name', 'Drug'), TINTS['drug'], width=col_width)
                if drug.get('concentration'):
                    y = draw_row(fig1, left_x, y, 'Concentration:', f"{drug.get('concentration')}µM", TINTS['drug'], width=col_width)
                y = draw_row(fig1, left_x, y, 'Perf. Start:', f"{drug.get('start', 0)} min", TINTS['drug'], width=col_width)
                y = draw_row(fig1, left_x, y, 'Perf. Delay:', f"{drug.get('delay', 0)} min", TINTS['drug'], width=col_width)
                # Perf. Time = HRV readout minute if available, otherwise start + delay
                perf_start = drug.get('start', 0) or 0
                perf_delay = drug.get('delay', 0) or 0
                perf_time = perf_start + perf_delay  # Default
                # Try to get HRV readout minute from drug_readout_settings
                if request.drug_readout_settings:
                    settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
                    if request.drug_readout_settings.get('enableHrvReadout') and settings_hrv not in (None, ''):
                        try:
                            perf_time = int(float(settings_hrv)) + perf_start + perf_delay
                        except (ValueError, TypeError):
                            pass
                y = draw_row(fig1, left_x, y, 'Perf. Time:', f"{perf_time} min", TINTS['drug'], width=col_width)
                perf_end = drug.get('end')
                y = draw_row(fig1, left_x, y, 'Perf. End:', f"{perf_end} min" if perf_end is not None else '—', TINTS['drug'], width=col_width)
        
        # LIGHT STIMULATION
        if request.light_enabled:
            y -= section_gap
            y = draw_header(fig1, left_x, y, 'LIGHT STIMULATION', COLORS['amber'], width=col_width)
            y = draw_row(fig1, left_x, y, 'Status:', 'Enabled', TINTS['light'], width=col_width)
            if request.light_stim_count and request.light_stim_count > 0:
                y = draw_row(fig1, left_x, y, 'Stims Detected:', str(request.light_stim_count), TINTS['light'], width=col_width)
            # Stims Start - from first pulse (below Stims Detected)
            if request.light_pulses and len(request.light_pulses) > 0:
                first_pulse = request.light_pulses[0]
                light_start = first_pulse.get('start_min')
                if light_start is not None:
                    y = draw_row(fig1, left_x, y, 'Stims Start:', f"{light_start:.2f} min", TINTS['light'], width=col_width)
            if request.light_params:
                if request.light_params.get('pulseDuration') is not None:
                    y = draw_row(fig1, left_x, y, 'Stim Duration:', f"{request.light_params.get('pulseDuration')} sec", TINTS['light'], width=col_width)
                if request.light_params.get('interval'):
                    # Map interval values to display labels
                    interval_val = request.light_params.get('interval')
                    interval_display_map = {
                        'decreasing': '60s-30s-20s-10s',
                        '60': 'Uniform 60s',
                        '30': 'Uniform 30s',
                    }
                    interval_display = interval_display_map.get(str(interval_val), str(interval_val))
                    y = draw_row(fig1, left_x, y, 'Inter-stimuli intervals:', interval_display, TINTS['light'], width=col_width)
        
        # RIGHT COLUMN - READOUTS
        y_right = first_section_y
        
        # BASELINE READOUT
        if request.baseline_enabled and request.baseline:
            y_right = draw_header(fig1, right_x, y_right, 'BASELINE READOUT', COLORS['sky'], width=col_width)
            baseline = request.baseline
            bf_val = baseline.get('baseline_bf')
            y_right = draw_row(fig1, right_x, y_right, 'Mean BF:', f"{bf_val:.1f} bpm" if bf_val else '—', TINTS['baseline'], width=col_width)
            ln_rmssd = baseline.get('baseline_ln_rmssd70')
            y_right = draw_row(fig1, right_x, y_right, 'ln(RMSSD₇₀):', f"{ln_rmssd:.3f}" if ln_rmssd else '—', TINTS['baseline'], width=col_width)
            sdnn = baseline.get('baseline_sdnn')
            ln_sdnn = np.log(sdnn) if sdnn and sdnn > 0 else None
            y_right = draw_row(fig1, right_x, y_right, 'ln(SDNN₇₀):', f"{ln_sdnn:.3f}" if ln_sdnn else '—', TINTS['baseline'], width=col_width)
            pnn50 = baseline.get('baseline_pnn50')
            y_right = draw_row(fig1, right_x, y_right, 'pNN50₇₀:', f"{pnn50:.1f}%" if pnn50 is not None else '—', TINTS['baseline'], width=col_width)
            y_right -= section_gap
        
        # DRUG READOUT - show even without baseline
        if request.drug_readout_enabled and request.all_drugs and len(request.all_drugs) > 0:
            y_right = draw_header(fig1, right_x, y_right, 'DRUG READOUT', COLORS['purple'], width=col_width)
            
            drug_bf = None
            drug_hrv_data = None
            
            # Get drug readout minutes - check both drug_readout and drug_readout_settings
            drug_bf_minute = None
            drug_hrv_minute = None
            
            # First try to get from drug_readout (calculated values)
            if request.drug_readout:
                drug_bf_minute = request.drug_readout.get('bf_minute')
                drug_hrv_minute = request.drug_readout.get('hrv_minute')
            
            # Override with user settings if available (these are stored as strings sometimes)
            if request.drug_readout_settings:
                settings_bf = request.drug_readout_settings.get('bfReadoutMinute')
                settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
                
                # Get perfusion params for calculation
                perf_start = 0
                perf_delay = 0
                if request.all_drugs and len(request.all_drugs) > 0:
                    drug = request.all_drugs[0]
                    perf_start = drug.get('start', 0) or 0
                    perf_delay = drug.get('delay', 0) or 0
                
                # If user has enabled BF readout and provided a minute
                if request.drug_readout_settings.get('enableBfReadout') and settings_bf not in (None, ''):
                    try:
                        drug_bf_minute = int(float(settings_bf)) + perf_start + perf_delay
                    except (ValueError, TypeError):
                        pass
                
                # If user has enabled HRV readout and provided a minute
                if request.drug_readout_settings.get('enableHrvReadout') and settings_hrv not in (None, ''):
                    try:
                        drug_hrv_minute = int(float(settings_hrv)) + perf_start + perf_delay
                    except (ValueError, TypeError):
                        pass
            
            # Convert to int for comparison
            try:
                drug_bf_minute = int(drug_bf_minute) if drug_bf_minute is not None else None
            except (ValueError, TypeError):
                drug_bf_minute = None
                
            try:
                drug_hrv_minute = int(drug_hrv_minute) if drug_hrv_minute is not None else None
            except (ValueError, TypeError):
                drug_hrv_minute = None
            
            # Find BF at drug readout minute - check both per_minute_data and hrv_windows
            if drug_bf_minute is not None and request.per_minute_data:
                for pm in request.per_minute_data:
                    try:
                        minute_str = str(pm.get('minute', ''))
                        minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(float(minute_str))
                        if minute_num == drug_bf_minute:
                            drug_bf = pm.get('mean_bf')
                            break
                    except (ValueError, TypeError):
                        pass
            
            # Find HRV at drug readout minute
            if drug_hrv_minute is not None and request.hrv_windows:
                for w in request.hrv_windows:
                    try:
                        w_minute = w.get('minute')
                        # Handle different minute formats
                        if w_minute is None:
                            continue
                        if isinstance(w_minute, (int, float)):
                            w_minute_num = int(w_minute)
                        else:
                            w_minute_str = str(w_minute)
                            w_minute_num = int(w_minute_str.split('-')[0]) if '-' in w_minute_str else int(float(w_minute_str))
                        
                        if w_minute_num == drug_hrv_minute:
                            drug_hrv_data = w
                            # If BF not found in per_minute_data, get it from HRV window
                            if drug_bf is None and w.get('mean_bf'):
                                drug_bf = w.get('mean_bf')
                            break
                    except (ValueError, TypeError):
                        pass
            
            # Fallback: if BF still not found but HRV minute matches BF minute, use HRV window's BF
            if drug_bf is None and drug_bf_minute is not None and request.hrv_windows:
                for w in request.hrv_windows:
                    try:
                        w_minute = w.get('minute')
                        if w_minute is None:
                            continue
                        if isinstance(w_minute, (int, float)):
                            w_minute_num = int(w_minute)
                        else:
                            w_minute_str = str(w_minute)
                            w_minute_num = int(w_minute_str.split('-')[0]) if '-' in w_minute_str else int(float(w_minute_str))
                        
                        if w_minute_num == drug_bf_minute and w.get('mean_bf'):
                            drug_bf = w.get('mean_bf')
                            break
                    except (ValueError, TypeError):
                        pass
            
            # Always show drug metrics if available
            y_right = draw_row(fig1, right_x, y_right, 'Mean BF:', f"{drug_bf:.1f} bpm" if drug_bf else '—', TINTS['drug'], width=col_width)
            if drug_hrv_data:
                ln_rmssd = drug_hrv_data.get('ln_rmssd70')
                y_right = draw_row(fig1, right_x, y_right, 'ln(RMSSD₇₀):', f"{ln_rmssd:.3f}" if ln_rmssd else '—', TINTS['drug'], width=col_width)
                sdnn = drug_hrv_data.get('sdnn')
                ln_sdnn = np.log(sdnn) if sdnn and sdnn > 0 else None
                y_right = draw_row(fig1, right_x, y_right, 'ln(SDNN₇₀):', f"{ln_sdnn:.3f}" if ln_sdnn else '—', TINTS['drug'], width=col_width)
                pnn50 = drug_hrv_data.get('pnn50')
                y_right = draw_row(fig1, right_x, y_right, 'pNN50₇₀:', f"{pnn50:.1f}%" if pnn50 is not None else '—', TINTS['drug'], width=col_width)
            else:
                y_right = draw_row(fig1, right_x, y_right, 'HRV:', 'No data at readout', TINTS['drug'], width=col_width)
            y_right -= section_gap
        
        # LIGHT STIMULUS READOUT - all HRA metrics
        if request.light_enabled and (request.light_response or request.light_metrics_detrended):
            y_right = draw_header(fig1, right_x, y_right, 'LIGHT STIMULUS READOUT', COLORS['amber'], width=col_width)
            
            if request.light_response:
                valid = [r for r in request.light_response if r]
                if valid:
                    avg_bf = np.mean([r.get('avg_bf', 0) for r in valid if r.get('avg_bf')])
                    peak_bf = np.mean([r.get('peak_bf', 0) for r in valid if r.get('peak_bf')])
                    baseline_bf_vals = [r.get('baseline_bf') for r in valid if r.get('baseline_bf')]
                    baseline_bf = np.mean(baseline_bf_vals) if baseline_bf_vals else None
                    peak_norm_vals = [r.get('peak_norm_pct') for r in valid if r.get('peak_norm_pct') is not None]
                    peak_norm = np.mean(peak_norm_vals) if peak_norm_vals else None
                    ttp_vals = [r.get('time_to_peak_sec') for r in valid if r.get('time_to_peak_sec') is not None]
                    ttp = np.mean(ttp_vals) if ttp_vals else None
                    ttp_1st = valid[0].get('time_to_peak_sec') if valid else None
                    recovery_bf_vals = [r.get('bf_end') for r in valid if r.get('bf_end')]
                    recovery_bf = np.mean(recovery_bf_vals) if recovery_bf_vals else None
                    recovery_pct_vals = [r.get('bf_end_pct') for r in valid if r.get('bf_end_pct')]
                    recovery_pct = np.mean(recovery_pct_vals) if recovery_pct_vals else None
                    amplitude_vals = [r.get('amplitude') for r in valid if r.get('amplitude')]
                    amplitude = np.mean(amplitude_vals) if amplitude_vals else None
                    roc_vals = [r.get('rate_of_change') for r in valid if r.get('rate_of_change') is not None]
                    roc = np.mean(roc_vals) if roc_vals else None
                    
                    # Reorganized order: Baseline BF, Avg BF, Peak BF, Peak (Norm.), TTP (1st Stim), Time to Peak, Recovery BF, Recovery %, Amplitude, Rate of Change
                    y_right = draw_row(fig1, right_x, y_right, 'Baseline BF:', f"{baseline_bf:.1f} bpm" if baseline_bf else '—', TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Avg BF:', f"{avg_bf:.1f} bpm", TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Peak BF:', f"{peak_bf:.1f} bpm", TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Peak (Norm.):', f"{peak_norm:.1f}%" if peak_norm else '—', TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'TTP (1st Stim):', f"{ttp_1st:.1f} s" if ttp_1st is not None else '—', TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Time to Peak:', f"{ttp:.1f} s" if ttp is not None else '—', TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Recovery BF:', f"{recovery_bf:.1f} bpm" if recovery_bf else '—', TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Recovery %:', f"{recovery_pct:.1f}%" if recovery_pct else '—', TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Amplitude:', f"{amplitude:.1f} bpm" if amplitude else '—', TINTS['light'], width=col_width)
                    y_right = draw_row(fig1, right_x, y_right, 'Rate of Change:', f"{roc:.3f} 1/min" if roc else '—', TINTS['light'], width=col_width)
            
            # Corrected HRV
            if request.light_metrics_detrended and request.light_metrics_detrended.get('final'):
                y_right -= 0.008
                fig1.text(right_x, y_right, 'Corrected HRV:', fontsize=7, fontstyle='italic', color='#52525b')
                y_right -= 0.020  # Line height for small text
                final = request.light_metrics_detrended['final']
                ln_rmssd = final.get('ln_rmssd70_detrended')
                y_right = draw_row(fig1, right_x, y_right, 'ln(RMSSD₇₀):', f"{ln_rmssd:.3f}" if ln_rmssd else '—', TINTS['light'], width=col_width)
                ln_sdnn = final.get('ln_sdnn70_detrended')
                y_right = draw_row(fig1, right_x, y_right, 'ln(SDNN₇₀):', f"{ln_sdnn:.3f}" if ln_sdnn else '—', TINTS['light'], width=col_width)
                pnn50 = final.get('pnn50_detrended')
                y_right = draw_row(fig1, right_x, y_right, 'pNN50₇₀:', f"{pnn50:.1f}%" if pnn50 is not None else '—', TINTS['light'], width=col_width)
        
        add_page_footer(fig1, page_num)
        pdf.savefig(fig1)
        plt.close(fig1)
        
        # ==================== PAGE 2: BF TRACES ====================
        if request.per_beat_data or request.per_minute_data:
            page_num += 1
            fig2 = plt.figure(figsize=(8.5, 11))
            fig2.patch.set_facecolor('white')
            
            # Bioptima-style header and title - moved 0.2cm lower
            add_page_header(fig2, 'traces')
            fig2.text(0.08, 0.90, 'BF Evolution', ha='left', va='top', fontsize=28, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            fig2.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig2.transFigure))
            
            # Get baseline or drug BF for normalization
            baseline_bf = None
            norm_source = 'Baseline'
            if request.baseline and request.baseline.get('baseline_bf'):
                baseline_bf = request.baseline.get('baseline_bf')
            elif request.drug_readout and request.per_minute_data:
                # Use drug readout if no baseline
                drug_bf_minute = request.drug_readout.get('bf_minute')
                if drug_bf_minute is not None:
                    try:
                        drug_bf_minute = int(drug_bf_minute)
                    except (ValueError, TypeError):
                        drug_bf_minute = None
                    
                    if drug_bf_minute is not None:
                        for pm in request.per_minute_data:
                            try:
                                minute_str = str(pm.get('minute', ''))
                                minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(float(minute_str))
                                if minute_num == drug_bf_minute:
                                    baseline_bf = pm.get('mean_bf')
                                    norm_source = 'Drug Readout'
                                    break
                            except (ValueError, TypeError):
                                pass
                        
                        # Fallback: try hrv_windows for BF
                        if baseline_bf is None and request.hrv_windows:
                            for w in request.hrv_windows:
                                try:
                                    w_minute = w.get('minute')
                                    if isinstance(w_minute, (int, float)):
                                        w_minute_num = int(w_minute)
                                    else:
                                        w_minute_str = str(w_minute)
                                        w_minute_num = int(w_minute_str.split('-')[0]) if '-' in w_minute_str else int(float(w_minute_str))
                                    
                                    if w_minute_num == drug_bf_minute and w.get('mean_bf'):
                                        baseline_bf = w.get('mean_bf')
                                        norm_source = 'Drug Readout'
                                        break
                                except (ValueError, TypeError):
                                    pass
            
            # Fallback to drug_readout_settings
            if baseline_bf is None and request.drug_readout_settings:
                settings_bf = request.drug_readout_settings.get('bfReadoutMinute')
                if request.drug_readout_settings.get('enableBfReadout') and settings_bf not in (None, ''):
                    try:
                        perf_start = 0
                        perf_delay = 0
                        if request.all_drugs and len(request.all_drugs) > 0:
                            drug = request.all_drugs[0]
                            perf_start = drug.get('start', 0) or 0
                            perf_delay = drug.get('delay', 0) or 0
                        drug_bf_minute = int(float(settings_bf)) + int(perf_start) + int(perf_delay)
                        
                        # Try per_minute_data first
                        if request.per_minute_data:
                            for pm in request.per_minute_data:
                                try:
                                    minute_str = str(pm.get('minute', ''))
                                    minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(float(minute_str))
                                    if minute_num == drug_bf_minute:
                                        baseline_bf = pm.get('mean_bf')
                                        norm_source = 'Drug Readout'
                                        break
                                except (ValueError, TypeError):
                                    pass
                        
                        # Fallback to hrv_windows
                        if baseline_bf is None and request.hrv_windows:
                            for w in request.hrv_windows:
                                try:
                                    w_minute = w.get('minute')
                                    if isinstance(w_minute, (int, float)):
                                        w_minute_num = int(w_minute)
                                    else:
                                        w_minute_str = str(w_minute)
                                        w_minute_num = int(w_minute_str.split('-')[0]) if '-' in w_minute_str else int(float(w_minute_str))
                                    
                                    if w_minute_num == drug_bf_minute and w.get('mean_bf'):
                                        baseline_bf = w.get('mean_bf')
                                        norm_source = 'Drug Readout'
                                        break
                                except (ValueError, TypeError):
                                    pass
                    except (ValueError, TypeError):
                        pass
            
            # Prepare data
            times, bf_values = [], []
            if request.per_beat_data:
                times = [d['time_min'] for d in request.per_beat_data if d.get('status') == 'kept']
                bf_values = [d['bf_bpm'] for d in request.per_beat_data if d.get('status') == 'kept']
            elif request.per_minute_data:
                for pm in request.per_minute_data:
                    minute_str = str(pm.get('minute', ''))
                    try:
                        minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(minute_str)
                        times.append(minute_num + 0.5)
                        bf_values.append(pm.get('mean_bf', 0))
                    except (ValueError, TypeError):
                        pass
            
            if times and bf_values:
                time_max = max(times) if times else 10
                
                # Charts positioned below title bar - first trace at 1cm from bar
                # Bar is at 0.865, 1cm (~0.036) below = 0.83 for top of first chart
                chart_left = 0.15
                chart_width = 0.75
                chart_height = 0.30
                # Top chart: top at 0.83, so y position = 0.83 - 0.30 = 0.53
                top_chart_y = 0.53
                # Bottom chart: moved 1cm higher (add ~0.036)
                bottom_chart_y = 0.16
                
                # Top: BF Filtered
                ax1 = fig2.add_axes([chart_left, top_chart_y, chart_width, chart_height])
                ax1.scatter(times, bf_values, s=3, c=COLORS['emerald'], alpha=0.7, label='Filtered BF')
                ax1.set_ylabel('BF (bpm)', fontsize=9)
                ax1.set_xlabel('Time (min)', fontsize=9)
                ax1.set_title('Beat Frequency (Filtered)', fontsize=10, fontweight='bold', pad=6)
                ax1.set_xlim(0, time_max * 1.05)
                
                # Add drug regions with legend
                if request.all_drugs:
                    for drug in request.all_drugs:
                        start = drug.get('start', 0) + drug.get('delay', 0)
                        end = drug.get('end') if drug.get('end') else time_max * 1.1
                        ax1.axvspan(start, end, alpha=0.15, color=COLORS['purple'])
                
                # Add light stim regions with legend
                if request.light_enabled and request.light_pulses:
                    for pulse in request.light_pulses:
                        start_min = pulse.get('start_min', pulse.get('start_sec', 0) / 60)
                        end_min = pulse.get('end_min', pulse.get('end_sec', 0) / 60)
                        ax1.axvspan(start_min, end_min, alpha=0.2, color=COLORS['amber'])
                
                # Build legend - use Line2D for dot marker
                from matplotlib.lines import Line2D
                handles = [Line2D([0], [0], marker='o', color='w', markerfacecolor=COLORS['emerald'], 
                                  markersize=6, alpha=0.7, label='Filtered BF')]
                if request.all_drugs:
                    handles.append(mpatches.Patch(color=COLORS['purple'], alpha=0.3, label='Drug Perfusion'))
                if request.light_enabled and request.light_pulses:
                    handles.append(mpatches.Patch(color=COLORS['amber'], alpha=0.3, label='Light Stim'))
                ax1.legend(handles=handles, loc='upper right', fontsize=7, framealpha=0.9)
                ax1.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
                
                # Bottom: BF Normalized
                if baseline_bf and baseline_bf > 0:
                    ax2 = fig2.add_axes([chart_left, bottom_chart_y, chart_width, chart_height])
                    bf_norm = [100 * (bf / baseline_bf) for bf in bf_values]
                    ax2.scatter(times, bf_norm, s=3, c=COLORS['emerald'], alpha=0.7)
                    ax2.axhline(y=100, color='#dc2626', linestyle='--', linewidth=1)
                    ax2.set_ylabel('BF (% of Reference)', fontsize=9)
                    ax2.set_xlabel('Time (min)', fontsize=9)
                    ax2.set_title(f'Beat Frequency (Normalized to {norm_source})', fontsize=10, fontweight='bold', pad=8)
                    ax2.set_xlim(0, time_max * 1.05)
                    ax2.set_ylim(0, 200)
                    
                    # Add drug regions
                    if request.all_drugs:
                        for drug in request.all_drugs:
                            start = drug.get('start', 0) + drug.get('delay', 0)
                            end = drug.get('end') if drug.get('end') else time_max * 1.1
                            ax2.axvspan(start, end, alpha=0.15, color=COLORS['purple'])
                    
                    # Add light stim regions
                    if request.light_enabled and request.light_pulses:
                        for pulse in request.light_pulses:
                            start_min = pulse.get('start_min', pulse.get('start_sec', 0) / 60)
                            end_min = pulse.get('end_min', pulse.get('end_sec', 0) / 60)
                            ax2.axvspan(start_min, end_min, alpha=0.2, color=COLORS['amber'])
                    
                    # Build legend with dot, baseline line, drug and light stim
                    legend_label = f'{norm_source} Readout (100%)' if norm_source == 'Baseline' else f'{norm_source} (100%)'
                    handles2 = [
                        Line2D([0], [0], marker='o', color='w', markerfacecolor=COLORS['emerald'], 
                               markersize=6, alpha=0.7, label='Normalized BF'),
                        Line2D([0], [0], color='#dc2626', linestyle='--', linewidth=1, label=legend_label)
                    ]
                    if request.all_drugs:
                        handles2.append(mpatches.Patch(color=COLORS['purple'], alpha=0.3, label='Drug Perfusion'))
                    if request.light_enabled and request.light_pulses:
                        handles2.append(mpatches.Patch(color=COLORS['amber'], alpha=0.3, label='Light Stim'))
                    ax2.legend(handles=handles2, loc='upper right', fontsize=7, framealpha=0.9)
                    ax2.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
            
            add_page_footer(fig2, page_num)
            pdf.savefig(fig2)
            plt.close(fig2)
        
        # ==================== PAGE 3: HRV EVOLUTION ====================
        if request.hrv_windows:
            page_num += 1
            fig3 = plt.figure(figsize=(8.5, 11))
            fig3.patch.set_facecolor('white')
            
            # Bioptima-style header and title - moved 0.2cm lower
            add_page_header(fig3, 'traces')
            fig3.text(0.08, 0.90, 'HRV Evolution', ha='left', va='top', fontsize=28, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            fig3.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig3.transFigure))
            
            minutes = [w.get('minute', i) for i, w in enumerate(request.hrv_windows)]
            time_max = max(minutes) if minutes else 10
            
            ln_rmssd_vals = [w.get('ln_rmssd70') for w in request.hrv_windows]
            sdnn_vals = [w.get('sdnn') for w in request.hrv_windows]
            ln_sdnn_vals = [np.log(s) if s and s > 0 else None for s in sdnn_vals]
            pnn50_vals = [w.get('pnn50') for w in request.hrv_windows]
            
            # Charts - first trace at 1cm from bar
            # Bar is at 0.865, 1cm below = 0.83 for top of first chart
            # All traces and titles in SALMON color
            salmon_color = '#FA8072'  # Salmon color
            chart_left = 0.14
            chart_width = 0.76
            chart_height = 0.19
            
            # Positions: top chart at 0.83 - 0.19 = 0.64
            bottom_y = 0.12
            middle_y = 0.38
            top_y = 0.64
            
            # ln(RMSSD70) - Y: 0 to 8 - SALMON
            ax1 = fig3.add_axes([chart_left, top_y, chart_width, chart_height])
            valid_idx = [i for i, v in enumerate(ln_rmssd_vals) if v is not None]
            if valid_idx:
                ax1.plot([minutes[i] for i in valid_idx], [ln_rmssd_vals[i] for i in valid_idx],
                        'o-', color=salmon_color, markersize=4, linewidth=1.5)
            ax1.set_ylabel('ln(RMSSD₇₀)', fontsize=9)
            ax1.set_xlabel('Time (min)', fontsize=9)
            ax1.set_title('ln(RMSSD₇₀) Evolution', fontsize=10, fontweight='bold', color=salmon_color, pad=6)
            ax1.set_xlim(0, time_max + 1)
            ax1.set_ylim(0, 8)
            ax1.grid(True, alpha=0.3)
            if request.all_drugs:
                for drug in request.all_drugs:
                    start = drug.get('start', 0) + drug.get('delay', 0)
                    end = drug.get('end') if drug.get('end') else time_max + 1
                    ax1.axvspan(start, end, alpha=0.15, color=COLORS['purple'])
            
            # ln(SDNN70) - Y: 0 to 8 - SALMON
            ax2 = fig3.add_axes([chart_left, middle_y, chart_width, chart_height])
            valid_idx = [i for i, v in enumerate(ln_sdnn_vals) if v is not None]
            if valid_idx:
                ax2.plot([minutes[i] for i in valid_idx], [ln_sdnn_vals[i] for i in valid_idx],
                        'o-', color=salmon_color, markersize=4, linewidth=1.5)
            ax2.set_ylabel('ln(SDNN₇₀)', fontsize=9)
            ax2.set_xlabel('Time (min)', fontsize=9)
            ax2.set_title('ln(SDNN₇₀) Evolution', fontsize=10, fontweight='bold', color=salmon_color, pad=6)
            ax2.set_xlim(0, time_max + 1)
            ax2.set_ylim(0, 8)
            ax2.grid(True, alpha=0.3)
            if request.all_drugs:
                for drug in request.all_drugs:
                    start = drug.get('start', 0) + drug.get('delay', 0)
                    end = drug.get('end') if drug.get('end') else time_max + 1
                    ax2.axvspan(start, end, alpha=0.15, color=COLORS['purple'])
            
            # pNN50 - Y: 0 to 100 - SALMON
            ax3 = fig3.add_axes([chart_left, bottom_y, chart_width, chart_height])
            valid_idx = [i for i, v in enumerate(pnn50_vals) if v is not None]
            if valid_idx:
                ax3.plot([minutes[i] for i in valid_idx], [pnn50_vals[i] for i in valid_idx],
                        'o-', color=salmon_color, markersize=4, linewidth=1.5)
            ax3.set_ylabel('pNN50₇₀ (%)', fontsize=9)
            ax3.set_xlabel('Time (min)', fontsize=9)
            ax3.set_title('pNN50₇₀ Evolution', fontsize=10, fontweight='bold', color=salmon_color, pad=6)
            ax3.set_xlim(0, time_max + 1)
            ax3.set_ylim(0, 100)
            ax3.grid(True, alpha=0.3)
            if request.all_drugs:
                for drug in request.all_drugs:
                    start = drug.get('start', 0) + drug.get('delay', 0)
                    end = drug.get('end') if drug.get('end') else time_max + 1
                    ax3.axvspan(start, end, alpha=0.15, color=COLORS['purple'])
            
            add_page_footer(fig3, page_num)
            pdf.savefig(fig3)
            plt.close(fig3)
        
        # ==================== PAGE 3b: LIGHT-INDUCED CORRECTED HRV ANALYSIS ====================
        if request.light_enabled and request.light_metrics_detrended:
            # Support both 'per_stim' and 'per_pulse' keys (backend uses 'per_pulse')
            per_stim = request.light_metrics_detrended.get('per_stim') or request.light_metrics_detrended.get('per_pulse', [])
            # Filter for stims that have visualization data
            valid_stims = [(i, s) for i, s in enumerate(per_stim) if s and s.get('viz')]
            
            if valid_stims:
                page_num += 1
                n_stims = len(valid_stims)
                fig3b = plt.figure(figsize=(8.5, 11))
                fig3b.patch.set_facecolor('white')
                
                # Bioptima-style header and title - moved 0.2cm lower
                add_page_header(fig3b, 'traces')
                fig3b.text(0.08, 0.90, 'Light-induced Corrected HRV', ha='left', va='top', fontsize=28, fontweight='bold', 
                         color=COLORS['dark'], fontfamily=title_font)
                fig3b.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig3b.transFigure))
                
                # Charts positioned below title bar
                # Moved 0.3cm lower (subtract ~0.011)
                chart_left = 0.13
                chart_total_width = 0.77
                col3_offset = 0.02
                
                # Calculate row height based on number of stims
                available_height = 0.73
                row_height = min(0.14, available_height / max(n_stims, 1))
                top_margin = 0.86  # Moved 0.3cm lower from 0.87
                
                for row_idx, (stim_idx, stim_data) in enumerate(valid_stims):
                    viz = stim_data.get('viz', {})
                    time_rel = viz.get('time_rel', [])  # Time in seconds
                    nn_70 = viz.get('nn_70', [])
                    trend = viz.get('trend', [])
                    residual = viz.get('residual', [])
                    
                    if not time_rel or not nn_70:
                        continue
                    
                    y_pos = top_margin - (row_idx + 1) * row_height
                    col_width = chart_total_width / 3 - 0.03  # Three columns with gaps
                    
                    # Calculate shared Y-axis limits for columns 1 and 2 (NN values)
                    all_nn_values = list(nn_70) + (list(trend) if trend else [])
                    nn_min = min(all_nn_values) if all_nn_values else 700
                    nn_max = max(all_nn_values) if all_nn_values else 1400
                    nn_margin = (nn_max - nn_min) * 0.1
                    nn_ylim = (nn_min - nn_margin, nn_max + nn_margin)
                    
                    # Calculate Y-axis limits for column 3 (Residuals - centered at 0)
                    if residual:
                        res_max = max(abs(min(residual)), abs(max(residual)))
                        res_ylim = (-res_max * 1.2, res_max * 1.2)
                    else:
                        res_ylim = (-100, 100)
                    
                    # Column 1: NN₇₀ (emerald) - moved 0.2cm left
                    col1_x = chart_left
                    ax1 = fig3b.add_axes([col1_x, y_pos, col_width, row_height * 0.80])
                    ax1.plot(time_rel, nn_70, color=COLORS['emerald'], linewidth=1)
                    ax1.set_facecolor('white')
                    ax1.set_ylim(nn_ylim)
                    # Stim label vertical on the left side
                    ax1.text(-0.12, 0.5, f'Stim {stim_idx + 1}', transform=ax1.transAxes, 
                            fontsize=7, fontweight='bold', va='center', ha='right', rotation=90)
                    if row_idx == 0:
                        ax1.set_title('NN₇₀ (ms)', fontsize=8, fontweight='bold', pad=4)
                    if row_idx == n_stims - 1:
                        ax1.set_xlabel('Time (s)', fontsize=7)
                    else:
                        ax1.set_xticklabels([])
                    ax1.tick_params(axis='both', labelsize=5)
                    ax1.grid(True, alpha=0.3)
                    
                    # Column 2: NN₇₀ + Trend (emerald + grey)
                    col2_x = chart_left + chart_total_width / 3
                    ax2 = fig3b.add_axes([col2_x, y_pos, col_width, row_height * 0.80])
                    ax2.plot(time_rel, nn_70, color=COLORS['emerald'], linewidth=1, alpha=0.7)
                    if trend:
                        ax2.plot(time_rel, trend, color='#6b7280', linewidth=1.5)  # Grey color
                    ax2.set_facecolor('white')
                    ax2.set_ylim(nn_ylim)  # Same Y-axis as column 1
                    if row_idx == 0:
                        ax2.set_title('NN₇₀ + Trend', fontsize=8, fontweight='bold', pad=4)
                    if row_idx == n_stims - 1:
                        ax2.set_xlabel('Time (s)', fontsize=7)
                    else:
                        ax2.set_xticklabels([])
                    ax2.set_yticklabels([])
                    ax2.tick_params(axis='both', labelsize=5)
                    ax2.grid(True, alpha=0.3)
                    
                    # Column 3: Residuals (amber) - moved 0.2cm right
                    col3_x = chart_left + 2 * chart_total_width / 3 + col3_offset
                    ax3 = fig3b.add_axes([col3_x, y_pos, col_width, row_height * 0.80])
                    if residual:
                        ax3.plot(time_rel, residual, color=COLORS['amber'], linewidth=1)
                        ax3.axhline(y=0, color='gray', linestyle='--', linewidth=0.5, alpha=0.7)
                    ax3.set_facecolor('white')
                    ax3.set_ylim(res_ylim)  # Different Y-axis, centered at 0
                    if row_idx == 0:
                        ax3.set_title('Residual (ms)', fontsize=8, fontweight='bold', pad=4)
                    if row_idx == n_stims - 1:
                        ax3.set_xlabel('Time (s)', fontsize=7)
                    else:
                        ax3.set_xticklabels([])
                    ax3.tick_params(axis='both', labelsize=5)
                    ax3.grid(True, alpha=0.3)
                
                add_page_footer(fig3b, page_num)
                pdf.savefig(fig3b)
                plt.close(fig3b)
        
        # ==================== PAGE 4: SPONTANEOUS ACTIVITY BF DATA TABLE ====================
        if request.per_minute_data:
            page_num += 1
            fig4 = plt.figure(figsize=(8.5, 11))
            fig4.patch.set_facecolor('white')
            
            # Bioptima-style header and title - moved 0.2cm lower
            add_page_header(fig4, 'spontaneous activity')
            fig4.text(0.08, 0.90, 'Beat Frequency', ha='left', va='top', fontsize=28, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            fig4.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig4.transFigure))
            
            # Table section title - moved 0.7cm higher
            fig4.text(0.08, 0.84, 'Table 1 | Per-Minute Beat Frequency Data', fontsize=11, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            fig4.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig4.transFigure))
            
            ax = fig4.add_axes([0.08, 0.10, 0.84, 0.72])
            ax.axis('off')
            
            # Get baseline and drug readout windows for highlighting
            baseline_window = None
            drug_window = None
            
            # Get baseline window if baseline is enabled
            if request.baseline_enabled and request.baseline:
                baseline_range = request.baseline.get('baseline_bf_range')
                if baseline_range is not None:
                    # Normalize baseline range to "X-Y" format (remove " min" suffix if present)
                    try:
                        baseline_range_str = str(baseline_range).replace(' min', '').strip()
                        if '-' in baseline_range_str:
                            baseline_window = baseline_range_str
                        else:
                            bmin = int(float(baseline_range_str))
                            baseline_window = f"{bmin}-{bmin+1}"
                    except (ValueError, TypeError):
                        baseline_window = str(baseline_range).replace(' min', '').strip()
                
                # Also check baseline_bf_minute as fallback
                if baseline_window is None:
                    bf_min = request.baseline.get('baseline_bf_minute')
                    if bf_min is not None:
                        try:
                            bf_min = int(float(bf_min))
                            baseline_window = f"{bf_min}-{bf_min+1}"
                        except (ValueError, TypeError):
                            pass
            
            # Get drug readout window from drug_readout_settings (user override) or drug_readout (calculated)
            if request.drug_readout_enabled or request.drug_readout or request.drug_readout_settings:
                # First try user-specified override from drug_readout_settings
                if request.drug_readout_settings:
                    settings_bf = request.drug_readout_settings.get('bfReadoutMinute')
                    if settings_bf not in (None, ''):
                        try:
                            perf_start = 0
                            perf_delay = 0
                            if request.all_drugs and len(request.all_drugs) > 0:
                                drug = request.all_drugs[0]
                                perf_start = int(float(drug.get('start', 0) or 0))
                                perf_delay = int(float(drug.get('delay', 0) or 0))
                            bf_min = int(float(settings_bf)) + perf_start + perf_delay
                            drug_window = f"{bf_min}-{bf_min+1}"
                        except (ValueError, TypeError):
                            pass
                
                # Fallback to calculated drug_readout if no user override
                if drug_window is None and request.drug_readout:
                    bf_min = request.drug_readout.get('bf_minute')
                    if bf_min is not None:
                        try:
                            bf_min = int(float(bf_min))
                            drug_window = f"{bf_min}-{bf_min+1}"
                        except (ValueError, TypeError):
                            pass
            
            headers = ['Window (min)', 'Mean BF (bpm)', 'Mean NN (ms)']
            table_data = []
            row_colors = []
            
            for pm in request.per_minute_data:
                minute_val = pm.get('minute', '')
                # Format window as "X-X+1"
                try:
                    minute_num = int(str(minute_val).split('-')[0]) if '-' in str(minute_val) else int(minute_val)
                    window_str = f"{minute_num}-{minute_num+1}"
                except (ValueError, TypeError):
                    window_str = str(minute_val)
                
                # Get BF and NN values - support both naming conventions
                bf_val = pm.get('mean_bf') or pm.get('avg_bf')
                nn_val = pm.get('mean_nn') or pm.get('avg_nn')
                
                table_data.append([
                    window_str,
                    f"{bf_val:.1f}" if bf_val else '—',
                    f"{nn_val:.1f}" if nn_val else '—',
                ])
                
                # Determine row color
                if baseline_window and window_str == baseline_window:
                    row_colors.append(TINTS['baseline'])
                elif drug_window and window_str == drug_window:
                    row_colors.append(TINTS['drug'])
                else:
                    row_colors.append(None)
            
            if table_data:
                table = ax.table(cellText=table_data, colLabels=headers, loc='upper center', cellLoc='center')
                table.auto_set_font_size(False)
                table.set_fontsize(9)
                table.scale(1.0, 1.8)
                
                for (row, col), cell in table.get_celld().items():
                    cell.set_edgecolor('#e5e7eb')
                    if row == 0:
                        cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                        cell.set_facecolor(COLORS['emerald'])
                    elif row > 0 and row <= len(row_colors) and row_colors[row-1]:
                        cell.set_facecolor(row_colors[row-1])
                        cell.set_text_props(fontweight='bold', fontfamily=body_font)
                    else:
                        cell.set_facecolor('#d1fae5' if row % 2 == 0 else 'white')
                        cell.set_text_props(fontfamily=body_font)
            
            add_page_footer(fig4, page_num)
            pdf.savefig(fig4)
            plt.close(fig4)
        
        # ==================== PAGE 5: SPONTANEOUS ACTIVITY HRV DATA TABLE ====================
        if request.hrv_windows:
            page_num += 1
            fig5 = plt.figure(figsize=(8.5, 11))
            fig5.patch.set_facecolor('white')
            
            # Bioptima-style header and title - moved 0.2cm lower
            add_page_header(fig5, 'spontaneous activity')
            fig5.text(0.08, 0.90, 'Heart Rate Variability', ha='left', va='top', fontsize=28, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            fig5.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig5.transFigure))
            
            # Table section title - moved 0.7cm higher
            fig5.text(0.08, 0.84, 'Table 2 | Per-Three Minutes HRV Data', fontsize=11, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            fig5.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig5.transFigure))
            
            ax = fig5.add_axes([0.08, 0.10, 0.84, 0.72])
            ax.axis('off')
            
            # Get baseline and drug readout windows for highlighting
            baseline_minute = None
            drug_minute = None
            
            # Get baseline minute if baseline is enabled
            if request.baseline_enabled and request.baseline:
                baseline_minute = request.baseline.get('baseline_hrv_minute')
                if baseline_minute is not None:
                    try:
                        baseline_minute = int(float(baseline_minute))
                    except (ValueError, TypeError):
                        baseline_minute = None
            
            # Get drug readout minute from drug_readout_settings (user override) or drug_readout (calculated)
            if request.drug_readout_enabled or request.drug_readout or request.drug_readout_settings:
                # First try user-specified override from drug_readout_settings
                if request.drug_readout_settings:
                    settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
                    if settings_hrv not in (None, ''):
                        try:
                            perf_start = 0
                            perf_delay = 0
                            if request.all_drugs and len(request.all_drugs) > 0:
                                drug = request.all_drugs[0]
                                perf_start = int(float(drug.get('start', 0) or 0))
                                perf_delay = int(float(drug.get('delay', 0) or 0))
                            drug_minute = int(float(settings_hrv)) + perf_start + perf_delay
                        except (ValueError, TypeError):
                            pass
                
                # Fallback to calculated drug_readout if no user override
                if drug_minute is None and request.drug_readout:
                    drug_minute = request.drug_readout.get('hrv_minute')
                    if drug_minute is not None:
                        try:
                            drug_minute = int(float(drug_minute))
                        except (ValueError, TypeError):
                            drug_minute = None
            
            headers = ['Window', 'ln(RMSSD₇₀)', 'RMSSD₇₀', 'ln(SDNN₇₀)', 'SDNN', 'pNN50₇₀', 'BF']
            table_data = []
            row_colors = []
            
            for w in request.hrv_windows:
                sdnn = w.get('sdnn')
                ln_sdnn = np.log(sdnn) if sdnn and sdnn > 0 else None
                minute = w.get('minute', 0)
                
                # Convert minute to int for comparison
                try:
                    if isinstance(minute, (int, float)):
                        minute_num = int(minute)
                    else:
                        minute_str = str(minute)
                        minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(float(minute_str))
                except (ValueError, TypeError):
                    minute_num = None
                
                table_data.append([
                    w.get('window', ''),
                    f"{w.get('ln_rmssd70', 0):.3f}" if w.get('ln_rmssd70') else '—',
                    f"{w.get('rmssd70', 0):.1f}" if w.get('rmssd70') else '—',
                    f"{ln_sdnn:.3f}" if ln_sdnn else '—',
                    f"{sdnn:.1f}" if sdnn else '—',
                    f"{w.get('pnn50', 0):.1f}" if w.get('pnn50') is not None else '0.0',
                    f"{w.get('mean_bf', 0):.1f}" if w.get('mean_bf') else '—',
                ])
                
                # Determine row color
                if baseline_minute is not None and minute_num == baseline_minute:
                    row_colors.append(TINTS['baseline'])
                elif drug_minute is not None and minute_num == drug_minute:
                    row_colors.append(TINTS['drug'])
                else:
                    row_colors.append(None)
            
            if table_data:
                table = ax.table(cellText=table_data, colLabels=headers, loc='upper center', cellLoc='center')
                table.auto_set_font_size(False)
                table.set_fontsize(8)
                table.scale(1.0, 1.6)
                
                for (row, col), cell in table.get_celld().items():
                    cell.set_edgecolor('#e5e7eb')
                    if row == 0:
                        cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                        cell.set_facecolor(COLORS['emerald'])
                    elif row > 0 and row <= len(row_colors) and row_colors[row-1]:
                        cell.set_facecolor(row_colors[row-1])
                        cell.set_text_props(fontweight='bold', fontfamily=body_font)
                    else:
                        cell.set_facecolor('#d1fae5' if row % 2 == 0 else 'white')
                        cell.set_text_props(fontfamily=body_font)
            
            add_page_footer(fig5, page_num)
            pdf.savefig(fig5)
            plt.close(fig5)
        
        # ==================== PAGE 6: LIGHT-INDUCED HRA DATA TABLE ====================
        if request.light_enabled and request.light_response:
            valid = [r for r in request.light_response if r]
            if valid:
                page_num += 1
                fig6 = plt.figure(figsize=(8.5, 11))
                fig6.patch.set_facecolor('white')
                
                # Bioptima-style header and title - moved 0.2cm lower
                add_page_header(fig6, 'light stimulus')
                fig6.text(0.08, 0.90, 'Heart Rate Adaptation', ha='left', va='top', fontsize=28, fontweight='bold', 
                         color=COLORS['dark'], fontfamily=title_font)
                fig6.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig6.transFigure))
                
                # Table section title - moved 0.7cm higher
                fig6.text(0.08, 0.84, 'Table 3 | Per-Stimulus HRA Data', fontsize=11, fontweight='bold', 
                         color=COLORS['dark'], fontfamily=title_font)
                fig6.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig6.transFigure))
                
                ax = fig6.add_axes([0.08, 0.10, 0.84, 0.72])
                ax.axis('off')
                
                # All HRA metrics - reordered with 1st TTP
                headers = ['Stim', 'Baseline BF', 'Avg BF', 'Peak BF', 'Peak %', '1st TTP (s)', 'TTP (s)', 'BF Rec', 'Rec %', 'Amp. BF', 'RoC (1/min)']
                table_data = []
                
                for i, r in enumerate(valid):
                    # For 1st TTP, show 0.0 instead of dash
                    first_ttp_val = r.get('first_ttp_sec')
                    first_ttp_str = f"{first_ttp_val:.1f}" if first_ttp_val is not None else "0.0"
                    
                    table_data.append([
                        str(i + 1),
                        f"{r.get('baseline_bf', 0):.1f}" if r.get('baseline_bf') else '—',
                        f"{r.get('avg_bf', 0):.1f}" if r.get('avg_bf') else '—',
                        f"{r.get('peak_bf', 0):.1f}" if r.get('peak_bf') else '—',
                        f"{r.get('peak_norm_pct', 0):.1f}" if r.get('peak_norm_pct') else '—',
                        first_ttp_str,
                        f"{r.get('time_to_peak_sec', 0):.1f}" if r.get('time_to_peak_sec') is not None else '—',
                        f"{r.get('bf_end', 0):.1f}" if r.get('bf_end') else '—',
                        f"{r.get('bf_end_pct', 0):.1f}" if r.get('bf_end_pct') else '—',
                        f"{r.get('amplitude', 0):.1f}" if r.get('amplitude') is not None else '—',
                        f"{r.get('rate_of_change', 0):.3f}" if r.get('rate_of_change') is not None else '—',
                    ])
                
                # Add average row
                if len(valid) > 1:
                    def safe_avg(key):
                        vals = [r.get(key) for r in valid if r.get(key) is not None]
                        return np.mean(vals) if vals else None
                    
                    # For 1st TTP avg, include 0 values
                    first_ttp_vals = [r.get('first_ttp_sec', 0) for r in valid]
                    first_ttp_avg = np.mean(first_ttp_vals) if first_ttp_vals else 0
                    
                    avg_row = [
                        'Avg',
                        f"{safe_avg('baseline_bf'):.1f}" if safe_avg('baseline_bf') else '—',
                        f"{safe_avg('avg_bf'):.1f}" if safe_avg('avg_bf') else '—',
                        f"{safe_avg('peak_bf'):.1f}" if safe_avg('peak_bf') else '—',
                        f"{safe_avg('peak_norm_pct'):.1f}" if safe_avg('peak_norm_pct') else '—',
                        f"{first_ttp_avg:.1f}",
                        f"{safe_avg('time_to_peak_sec'):.1f}" if safe_avg('time_to_peak_sec') is not None else '—',
                        f"{safe_avg('bf_end'):.1f}" if safe_avg('bf_end') else '—',
                        f"{safe_avg('bf_end_pct'):.1f}" if safe_avg('bf_end_pct') else '—',
                        f"{safe_avg('amplitude'):.1f}" if safe_avg('amplitude') is not None else '—',
                        f"{safe_avg('rate_of_change'):.3f}" if safe_avg('rate_of_change') is not None else '—',
                    ]
                    table_data.append(avg_row)
                
                table = ax.table(cellText=table_data, colLabels=headers, loc='upper center', cellLoc='center')
                table.auto_set_font_size(False)
                table.set_fontsize(7)
                table.scale(1.0, 1.8)
                
                # Lighter red for average row
                light_red = '#f87171'
                
                for (row, col), cell in table.get_celld().items():
                    cell.set_edgecolor('#e5e7eb')
                    if row == 0:
                        cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                        cell.set_facecolor(COLORS['amber'])
                    elif row == len(table_data):
                        cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                        cell.set_facecolor(light_red)
                    else:
                        cell.set_facecolor('#fef3c7' if row % 2 == 0 else 'white')
                        cell.set_text_props(fontfamily=body_font)
                
                add_page_footer(fig6, page_num)
                pdf.savefig(fig6)
                plt.close(fig6)
        
        # ==================== PAGE 7: LIGHT-INDUCED CORRECTED HRV DATA TABLE ====================
        if request.light_enabled and request.light_metrics_detrended:
            # Support both 'per_stim' and 'per_pulse' keys (backend uses 'per_pulse')
            per_stim = request.light_metrics_detrended.get('per_stim') or request.light_metrics_detrended.get('per_pulse', [])
            final = request.light_metrics_detrended.get('final', {})
            
            if per_stim or final:
                page_num += 1
                fig7 = plt.figure(figsize=(8.5, 11))
                fig7.patch.set_facecolor('white')
                
                # Bioptima-style header and title - moved 0.2cm lower
                add_page_header(fig7, 'light stimulus')
                fig7.text(0.08, 0.90, 'Corrected HRV', ha='left', va='top', fontsize=28, fontweight='bold', 
                         color=COLORS['dark'], fontfamily=title_font)
                fig7.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig7.transFigure))
                
                # Table section title - moved 0.7cm higher
                fig7.text(0.08, 0.84, 'Table 4 | Per-Stimulus Detrended HRV Data', fontsize=11, fontweight='bold', 
                         color=COLORS['dark'], fontfamily=title_font)
                fig7.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig7.transFigure))
                
                ax = fig7.add_axes([0.08, 0.10, 0.84, 0.72])
                ax.axis('off')
                
                headers = ['Stim', 'ln(RMSSD₇₀)', 'RMSSD₇₀', 'ln(SDNN₇₀)', 'SDNN', 'pNN50₇₀']
                table_data = []
                
                # Always show 5 stims before the median
                num_stims = max(5, len(per_stim))
                for i in range(num_stims):
                    s = per_stim[i] if i < len(per_stim) else None
                    # Always show the stim number, check if we have actual HRV data
                    has_data = s and (s.get('ln_rmssd70_detrended') is not None or 
                                     s.get('rmssd70_detrended') is not None or
                                     s.get('ln_sdnn70_detrended') is not None)
                    
                    if has_data:
                        table_data.append([
                            str(i + 1),
                            f"{s.get('ln_rmssd70_detrended', 0):.3f}" if s.get('ln_rmssd70_detrended') is not None else '—',
                            f"{s.get('rmssd70_detrended', 0):.3f}" if s.get('rmssd70_detrended') is not None else '—',
                            f"{s.get('ln_sdnn70_detrended', 0):.3f}" if s.get('ln_sdnn70_detrended') is not None else '—',
                            f"{s.get('sdnn_detrended', 0):.3f}" if s.get('sdnn_detrended') is not None else '—',
                            f"{s.get('pnn50_detrended', 0):.1f}" if s.get('pnn50_detrended') is not None else '—',
                        ])
                    else:
                        # Show stim number even if data is empty/missing
                        table_data.append([str(i + 1), '—', '—', '—', '—', '—'])
                
                # Add median row at the end
                if final:
                    table_data.append([
                        'Median',
                        f"{final.get('ln_rmssd70_detrended', 0):.3f}" if final.get('ln_rmssd70_detrended') else '—',
                        f"{final.get('rmssd70_detrended', 0):.3f}" if final.get('rmssd70_detrended') else '—',
                        f"{final.get('ln_sdnn70_detrended', 0):.3f}" if final.get('ln_sdnn70_detrended') else '—',
                        f"{final.get('sdnn_detrended', 0):.3f}" if final.get('sdnn_detrended') else '—',
                        f"{final.get('pnn50_detrended', 0):.1f}" if final.get('pnn50_detrended') is not None else '—',
                    ])
                
                if table_data:
                    table = ax.table(cellText=table_data, colLabels=headers, loc='upper center', cellLoc='center')
                    table.auto_set_font_size(False)
                    table.set_fontsize(9)
                    table.scale(1.0, 1.8)
                    
                    # Lighter red for median row
                    light_red = '#f87171'
                    
                    for (row, col), cell in table.get_celld().items():
                        cell.set_edgecolor('#e5e7eb')
                        if row == 0:
                            cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                            cell.set_facecolor(COLORS['amber'])
                        elif row == len(table_data):
                            cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                            cell.set_facecolor(light_red)
                        else:
                            cell.set_facecolor('#fef3c7' if row % 2 == 0 else 'white')
                            cell.set_text_props(fontfamily=body_font)
                
                add_page_footer(fig7, page_num)
                pdf.savefig(fig7)
                plt.close(fig7)
    
    buf.seek(0)
    return buf


def create_nature_excel(request):
    """Create a clean Excel export matching PDF structure with multiple sheets"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    section_font = Font(bold=True, size=11)
    data_font = Font(size=9)
    bold_data_font = Font(bold=True, size=9)
    avg_font = Font(bold=True, color='FFFFFF', size=9)
    
    emerald_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    purple_fill = PatternFill(start_color='A855F7', end_color='A855F7', fill_type='solid')
    amber_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    sky_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
    dark_fill = PatternFill(start_color='18181B', end_color='18181B', fill_type='solid')
    gray_fill = PatternFill(start_color='6B7280', end_color='6B7280', fill_type='solid')
    avg_fill = PatternFill(start_color='F87171', end_color='F87171', fill_type='solid')
    
    baseline_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
    drug_fill = PatternFill(start_color='F3E8FF', end_color='F3E8FF', fill_type='solid')
    light_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def fmt(val, dec=2):
        if val is None:
            return '—'
        try:
            if dec == 0:
                return f"{float(val):.0f}"
            elif dec == 1:
                return f"{float(val):.1f}"
            elif dec == 3:
                return f"{float(val):.3f}"
            elif dec == 4:
                return f"{float(val):.4f}"
            return f"{float(val):.{dec}f}"
        except:
            return '—'
    
    # ==================== SHEET 1: SUMMARY ====================
    ws = wb.active
    ws.title = 'Summary'
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 3  # Empty separator column
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 25
    
    row = 1
    title = request.recording_name or request.filename or 'Recording Analysis'
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ws.cell(row=row, column=1, value='Electrophysiology Analysis Report by NEHER').font = Font(size=10, color='71717A')
    row += 1
    if request.recording_date:
        ws.cell(row=row, column=1, value=f'Recording Date: {request.recording_date}').font = Font(size=9, color='A1A1AA')
    row += 2
    
    # LEFT COLUMN - RECORDING INFO
    left_col = 1
    right_col = 4  # Column D (after empty separator column C)
    left_row = row
    right_row = row
    
    # RECORDING INFO
    ws.cell(row=left_row, column=left_col, value='RECORDING INFO').font = header_font
    ws.cell(row=left_row, column=left_col).fill = dark_fill
    ws.merge_cells(start_row=left_row, start_column=left_col, end_row=left_row, end_column=left_col+1)
    left_row += 1
    
    if request.original_filename:
        ws.cell(row=left_row, column=left_col, value='Original File:').font = data_font
        ws.cell(row=left_row, column=left_col+1, value=request.original_filename).font = bold_data_font
        left_row += 1
    if request.recording_date:
        ws.cell(row=left_row, column=left_col, value='Recording Date:').font = data_font
        ws.cell(row=left_row, column=left_col+1, value=request.recording_date).font = bold_data_font
        left_row += 1
    if request.summary:
        if 'Total Beats' in request.summary:
            ws.cell(row=left_row, column=left_col, value='Total Beats:').font = data_font
            ws.cell(row=left_row, column=left_col+1, value=request.summary['Total Beats']).font = bold_data_font
            left_row += 1
        if 'Kept Beats' in request.summary:
            ws.cell(row=left_row, column=left_col, value='Kept Beats:').font = data_font
            ws.cell(row=left_row, column=left_col+1, value=request.summary['Kept Beats']).font = bold_data_font
            left_row += 1
        if 'Filter Range' in request.summary:
            ws.cell(row=left_row, column=left_col, value='Filter Range:').font = data_font
            ws.cell(row=left_row, column=left_col+1, value=request.summary['Filter Range']).font = bold_data_font
            left_row += 1
    left_row += 1
    
    # TISSUE INFO
    if request.organoid_info:
        ws.cell(row=left_row, column=left_col, value='TISSUE INFO').font = header_font
        ws.cell(row=left_row, column=left_col).fill = gray_fill
        ws.merge_cells(start_row=left_row, start_column=left_col, end_row=left_row, end_column=left_col+1)
        left_row += 1
        
        for idx, org in enumerate(request.organoid_info):
            if org.get('cell_type'):
                cell_type = org.get('other_cell_type') if org.get('cell_type') == 'Other' else org.get('cell_type')
                label = f'Cell Type {idx + 1}:' if len(request.organoid_info) > 1 else 'Cell Type:'
                ws.cell(row=left_row, column=left_col, value=label).font = data_font
                ws.cell(row=left_row, column=left_col+1, value=cell_type or '—').font = bold_data_font
                left_row += 1
            if org.get('line_name'):
                ws.cell(row=left_row, column=left_col, value='Line:').font = data_font
                ws.cell(row=left_row, column=left_col+1, value=org.get('line_name')).font = bold_data_font
                left_row += 1
            if org.get('passage_number'):
                ws.cell(row=left_row, column=left_col, value='Passage:').font = data_font
                ws.cell(row=left_row, column=left_col+1, value=org.get('passage_number')).font = bold_data_font
                left_row += 1
            if org.get('age_at_recording') is not None:
                ws.cell(row=left_row, column=left_col, value='Age at Recording:').font = data_font
                ws.cell(row=left_row, column=left_col+1, value=f"{org.get('age_at_recording')} days").font = bold_data_font
                left_row += 1
            if org.get('transfection'):
                trans = org['transfection']
                if trans.get('name'):
                    ws.cell(row=left_row, column=left_col, value='Transfection:').font = data_font
                    ws.cell(row=left_row, column=left_col+1, value=trans.get('name')).font = bold_data_font
                    left_row += 1
                if trans.get('days_since_transfection') is not None:
                    ws.cell(row=left_row, column=left_col, value='Days Post-Transf.:').font = data_font
                    ws.cell(row=left_row, column=left_col+1, value=trans.get('days_since_transfection')).font = bold_data_font
                    left_row += 1
        
        if request.days_since_fusion is not None:
            ws.cell(row=left_row, column=left_col, value='Days Since Fusion:').font = data_font
            ws.cell(row=left_row, column=left_col+1, value=request.days_since_fusion).font = bold_data_font
            left_row += 1
        left_row += 1
    
    # DRUG PERFUSION
    if request.all_drugs and len(request.all_drugs) > 0:
        ws.cell(row=left_row, column=left_col, value='DRUG PERFUSION').font = header_font
        ws.cell(row=left_row, column=left_col).fill = purple_fill
        ws.merge_cells(start_row=left_row, start_column=left_col, end_row=left_row, end_column=left_col+1)
        left_row += 1
        
        for drug in request.all_drugs:
            ws.cell(row=left_row, column=left_col, value='Drug:').font = data_font
            ws.cell(row=left_row, column=left_col).fill = drug_fill
            ws.cell(row=left_row, column=left_col+1, value=drug.get('name', 'Drug')).font = bold_data_font
            ws.cell(row=left_row, column=left_col+1).fill = drug_fill
            left_row += 1
            if drug.get('concentration'):
                ws.cell(row=left_row, column=left_col, value='Concentration:').font = data_font
                ws.cell(row=left_row, column=left_col).fill = drug_fill
                ws.cell(row=left_row, column=left_col+1, value=f"{drug.get('concentration')}µM").font = bold_data_font
                ws.cell(row=left_row, column=left_col+1).fill = drug_fill
                left_row += 1
            ws.cell(row=left_row, column=left_col, value='Perf. Start:').font = data_font
            ws.cell(row=left_row, column=left_col).fill = drug_fill
            ws.cell(row=left_row, column=left_col+1, value=f"{drug.get('start', 0)} min").font = bold_data_font
            ws.cell(row=left_row, column=left_col+1).fill = drug_fill
            left_row += 1
            ws.cell(row=left_row, column=left_col, value='Perf. Delay:').font = data_font
            ws.cell(row=left_row, column=left_col).fill = drug_fill
            ws.cell(row=left_row, column=left_col+1, value=f"{drug.get('delay', 0)} min").font = bold_data_font
            ws.cell(row=left_row, column=left_col+1).fill = drug_fill
            left_row += 1
            # Perf. Time = HRV readout minute if available, otherwise start + delay
            perf_start = drug.get('start', 0) or 0
            perf_delay = drug.get('delay', 0) or 0
            perf_time = perf_start + perf_delay  # Default
            if request.drug_readout_settings:
                settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
                if request.drug_readout_settings.get('enableHrvReadout') and settings_hrv not in (None, ''):
                    try:
                        perf_time = int(float(settings_hrv)) + perf_start + perf_delay
                    except (ValueError, TypeError):
                        pass
            ws.cell(row=left_row, column=left_col, value='Perf. Time:').font = data_font
            ws.cell(row=left_row, column=left_col).fill = drug_fill
            ws.cell(row=left_row, column=left_col+1, value=f"{perf_time} min").font = bold_data_font
            ws.cell(row=left_row, column=left_col+1).fill = drug_fill
            left_row += 1
            perf_end = drug.get('end')
            ws.cell(row=left_row, column=left_col, value='Perf. End:').font = data_font
            ws.cell(row=left_row, column=left_col).fill = drug_fill
            ws.cell(row=left_row, column=left_col+1, value=f"{perf_end} min" if perf_end is not None else '—').font = bold_data_font
            ws.cell(row=left_row, column=left_col+1).fill = drug_fill
            left_row += 1
        left_row += 1
    
    # LIGHT STIMULATION
    if request.light_enabled:
        ws.cell(row=left_row, column=left_col, value='LIGHT STIMULATION').font = header_font
        ws.cell(row=left_row, column=left_col).fill = amber_fill
        ws.merge_cells(start_row=left_row, start_column=left_col, end_row=left_row, end_column=left_col+1)
        left_row += 1
        
        ws.cell(row=left_row, column=left_col, value='Status:').font = data_font
        ws.cell(row=left_row, column=left_col).fill = light_fill
        ws.cell(row=left_row, column=left_col+1, value='Enabled').font = bold_data_font
        ws.cell(row=left_row, column=left_col+1).fill = light_fill
        left_row += 1
        
        if request.light_stim_count and request.light_stim_count > 0:
            ws.cell(row=left_row, column=left_col, value='Stims Detected:').font = data_font
            ws.cell(row=left_row, column=left_col).fill = light_fill
            ws.cell(row=left_row, column=left_col+1, value=str(request.light_stim_count)).font = bold_data_font
            ws.cell(row=left_row, column=left_col+1).fill = light_fill
            left_row += 1
        
        if request.light_pulses and len(request.light_pulses) > 0:
            first_pulse = request.light_pulses[0]
            light_start = first_pulse.get('start_min')
            if light_start is not None:
                ws.cell(row=left_row, column=left_col, value='Stims Start:').font = data_font
                ws.cell(row=left_row, column=left_col).fill = light_fill
                ws.cell(row=left_row, column=left_col+1, value=f"{light_start:.2f} min").font = bold_data_font
                ws.cell(row=left_row, column=left_col+1).fill = light_fill
                left_row += 1
        
        if request.light_params:
            if request.light_params.get('pulseDuration') is not None:
                ws.cell(row=left_row, column=left_col, value='Stim Duration:').font = data_font
                ws.cell(row=left_row, column=left_col).fill = light_fill
                ws.cell(row=left_row, column=left_col+1, value=f"{request.light_params.get('pulseDuration')} sec").font = bold_data_font
                ws.cell(row=left_row, column=left_col+1).fill = light_fill
                left_row += 1
            if request.light_params.get('interval'):
                interval_val = request.light_params.get('interval')
                interval_map = {'decreasing': '60s-30s-20s-10s', '60': 'Uniform 60s', '30': 'Uniform 30s'}
                interval_display = interval_map.get(str(interval_val), str(interval_val))
                ws.cell(row=left_row, column=left_col, value='Inter-stim Intervals:').font = data_font
                ws.cell(row=left_row, column=left_col).fill = light_fill
                ws.cell(row=left_row, column=left_col+1, value=interval_display).font = bold_data_font
                ws.cell(row=left_row, column=left_col+1).fill = light_fill
                left_row += 1
    
    # RIGHT COLUMN - READOUTS
    # BASELINE READOUT
    if request.baseline_enabled and request.baseline:
        ws.cell(row=right_row, column=right_col, value='BASELINE READOUT').font = header_font
        ws.cell(row=right_row, column=right_col).fill = sky_fill
        ws.merge_cells(start_row=right_row, start_column=right_col, end_row=right_row, end_column=right_col+1)
        right_row += 1
        
        baseline = request.baseline
        bf_val = baseline.get('baseline_bf')
        ws.cell(row=right_row, column=right_col, value='Mean BF:').font = data_font
        ws.cell(row=right_row, column=right_col).fill = baseline_fill
        ws.cell(row=right_row, column=right_col+1, value=f"{bf_val:.1f} bpm" if bf_val else '—').font = bold_data_font
        ws.cell(row=right_row, column=right_col+1).fill = baseline_fill
        right_row += 1
        
        ln_rmssd = baseline.get('baseline_ln_rmssd70')
        ws.cell(row=right_row, column=right_col, value='ln(RMSSD₇₀):').font = data_font
        ws.cell(row=right_row, column=right_col).fill = baseline_fill
        ws.cell(row=right_row, column=right_col+1, value=f"{ln_rmssd:.3f}" if ln_rmssd else '—').font = bold_data_font
        ws.cell(row=right_row, column=right_col+1).fill = baseline_fill
        right_row += 1
        
        sdnn = baseline.get('baseline_sdnn')
        ln_sdnn = np.log(sdnn) if sdnn and sdnn > 0 else None
        ws.cell(row=right_row, column=right_col, value='ln(SDNN₇₀):').font = data_font
        ws.cell(row=right_row, column=right_col).fill = baseline_fill
        ws.cell(row=right_row, column=right_col+1, value=f"{ln_sdnn:.3f}" if ln_sdnn else '—').font = bold_data_font
        ws.cell(row=right_row, column=right_col+1).fill = baseline_fill
        right_row += 1
        
        pnn50 = baseline.get('baseline_pnn50')
        ws.cell(row=right_row, column=right_col, value='pNN50₇₀:').font = data_font
        ws.cell(row=right_row, column=right_col).fill = baseline_fill
        ws.cell(row=right_row, column=right_col+1, value=f"{pnn50:.1f}%" if pnn50 is not None else '—').font = bold_data_font
        ws.cell(row=right_row, column=right_col+1).fill = baseline_fill
        right_row += 2
    
    # DRUG READOUT
    if request.drug_readout_enabled and request.all_drugs and len(request.all_drugs) > 0:
        ws.cell(row=right_row, column=right_col, value='DRUG READOUT').font = header_font
        ws.cell(row=right_row, column=right_col).fill = purple_fill
        ws.merge_cells(start_row=right_row, start_column=right_col, end_row=right_row, end_column=right_col+1)
        right_row += 1
        
        drug_bf = None
        drug_hrv_data = None
        drug_bf_minute = None
        drug_hrv_minute = None
        
        if request.drug_readout:
            drug_bf_minute = request.drug_readout.get('bf_minute')
            drug_hrv_minute = request.drug_readout.get('hrv_minute')
        
        if request.drug_readout_settings:
            settings_bf = request.drug_readout_settings.get('bfReadoutMinute')
            settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
            perf_start = 0
            perf_delay = 0
            if request.all_drugs and len(request.all_drugs) > 0:
                drug = request.all_drugs[0]
                perf_start = drug.get('start', 0) or 0
                perf_delay = drug.get('delay', 0) or 0
            
            if request.drug_readout_settings.get('enableBfReadout') and settings_bf not in (None, ''):
                try:
                    drug_bf_minute = int(float(settings_bf)) + perf_start + perf_delay
                except (ValueError, TypeError):
                    pass
            
            if request.drug_readout_settings.get('enableHrvReadout') and settings_hrv not in (None, ''):
                try:
                    drug_hrv_minute = int(float(settings_hrv)) + perf_start + perf_delay
                except (ValueError, TypeError):
                    pass
        
        # Find BF at drug readout minute
        if drug_bf_minute is not None and request.per_minute_data:
            for pm in request.per_minute_data:
                try:
                    minute_str = str(pm.get('minute', ''))
                    minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(float(minute_str))
                    if minute_num == int(drug_bf_minute):
                        drug_bf = pm.get('mean_bf')
                        break
                except (ValueError, TypeError):
                    pass
        
        # Find HRV at drug readout minute
        if drug_hrv_minute is not None and request.hrv_windows:
            for w in request.hrv_windows:
                try:
                    w_minute = w.get('minute')
                    if w_minute is None:
                        continue
                    w_minute_num = int(w_minute) if isinstance(w_minute, (int, float)) else int(str(w_minute).split('-')[0])
                    if w_minute_num == int(drug_hrv_minute):
                        drug_hrv_data = w
                        if drug_bf is None and w.get('mean_bf'):
                            drug_bf = w.get('mean_bf')
                        break
                except (ValueError, TypeError):
                    pass
        
        ws.cell(row=right_row, column=right_col, value='Mean BF:').font = data_font
        ws.cell(row=right_row, column=right_col).fill = drug_fill
        ws.cell(row=right_row, column=right_col+1, value=f"{drug_bf:.1f} bpm" if drug_bf else '—').font = bold_data_font
        ws.cell(row=right_row, column=right_col+1).fill = drug_fill
        right_row += 1
        
        if drug_hrv_data:
            ln_rmssd = drug_hrv_data.get('ln_rmssd70')
            ws.cell(row=right_row, column=right_col, value='ln(RMSSD₇₀):').font = data_font
            ws.cell(row=right_row, column=right_col).fill = drug_fill
            ws.cell(row=right_row, column=right_col+1, value=f"{ln_rmssd:.3f}" if ln_rmssd else '—').font = bold_data_font
            ws.cell(row=right_row, column=right_col+1).fill = drug_fill
            right_row += 1
            
            sdnn = drug_hrv_data.get('sdnn')
            ln_sdnn = np.log(sdnn) if sdnn and sdnn > 0 else None
            ws.cell(row=right_row, column=right_col, value='ln(SDNN₇₀):').font = data_font
            ws.cell(row=right_row, column=right_col).fill = drug_fill
            ws.cell(row=right_row, column=right_col+1, value=f"{ln_sdnn:.3f}" if ln_sdnn else '—').font = bold_data_font
            ws.cell(row=right_row, column=right_col+1).fill = drug_fill
            right_row += 1
            
            pnn50 = drug_hrv_data.get('pnn50')
            ws.cell(row=right_row, column=right_col, value='pNN50₇₀:').font = data_font
            ws.cell(row=right_row, column=right_col).fill = drug_fill
            ws.cell(row=right_row, column=right_col+1, value=f"{pnn50:.1f}%" if pnn50 is not None else '—').font = bold_data_font
            ws.cell(row=right_row, column=right_col+1).fill = drug_fill
            right_row += 1
        right_row += 1
    
    # LIGHT STIMULUS READOUT
    if request.light_enabled and (request.light_response or request.light_metrics_detrended):
        ws.cell(row=right_row, column=right_col, value='LIGHT STIMULUS READOUT').font = header_font
        ws.cell(row=right_row, column=right_col).fill = amber_fill
        ws.merge_cells(start_row=right_row, start_column=right_col, end_row=right_row, end_column=right_col+1)
        right_row += 1
        
        if request.light_response:
            valid = [r for r in request.light_response if r]
            if valid:
                baseline_bf_vals = [r.get('baseline_bf') for r in valid if r.get('baseline_bf')]
                baseline_bf = np.mean(baseline_bf_vals) if baseline_bf_vals else None
                avg_bf = np.mean([r.get('avg_bf', 0) for r in valid if r.get('avg_bf')])
                peak_bf = np.mean([r.get('peak_bf', 0) for r in valid if r.get('peak_bf')])
                peak_norm_vals = [r.get('peak_norm_pct') for r in valid if r.get('peak_norm_pct') is not None]
                peak_norm = np.mean(peak_norm_vals) if peak_norm_vals else None
                ttp_1st = valid[0].get('time_to_peak_sec') if valid else None
                ttp_vals = [r.get('time_to_peak_sec') for r in valid if r.get('time_to_peak_sec') is not None]
                ttp = np.mean(ttp_vals) if ttp_vals else None
                recovery_bf_vals = [r.get('bf_end') for r in valid if r.get('bf_end')]
                recovery_bf = np.mean(recovery_bf_vals) if recovery_bf_vals else None
                recovery_pct_vals = [r.get('bf_end_pct') for r in valid if r.get('bf_end_pct')]
                recovery_pct = np.mean(recovery_pct_vals) if recovery_pct_vals else None
                amplitude_vals = [r.get('amplitude') for r in valid if r.get('amplitude')]
                amplitude = np.mean(amplitude_vals) if amplitude_vals else None
                roc_vals = [r.get('rate_of_change') for r in valid if r.get('rate_of_change') is not None]
                roc = np.mean(roc_vals) if roc_vals else None
                
                hra_data = [
                    ('Baseline BF:', f"{baseline_bf:.1f} bpm" if baseline_bf else '—'),
                    ('Avg BF:', f"{avg_bf:.1f} bpm"),
                    ('Peak BF:', f"{peak_bf:.1f} bpm"),
                    ('Peak (Norm.):', f"{peak_norm:.1f}%" if peak_norm else '—'),
                    ('TTP (1st Stim):', f"{ttp_1st:.1f} s" if ttp_1st is not None else '—'),
                    ('Time to Peak:', f"{ttp:.1f} s" if ttp is not None else '—'),
                    ('Recovery BF:', f"{recovery_bf:.1f} bpm" if recovery_bf else '—'),
                    ('Recovery %:', f"{recovery_pct:.1f}%" if recovery_pct else '—'),
                    ('Amplitude:', f"{amplitude:.1f} bpm" if amplitude else '—'),
                    ('Rate of Change:', f"{roc:.3f} 1/min" if roc else '—'),
                ]
                
                for label, value in hra_data:
                    ws.cell(row=right_row, column=right_col, value=label).font = data_font
                    ws.cell(row=right_row, column=right_col).fill = light_fill
                    ws.cell(row=right_row, column=right_col+1, value=value).font = bold_data_font
                    ws.cell(row=right_row, column=right_col+1).fill = light_fill
                    right_row += 1
        
        # Corrected HRV
        if request.light_metrics_detrended and request.light_metrics_detrended.get('final'):
            right_row += 1
            ws.cell(row=right_row, column=right_col, value='Corrected HRV:').font = Font(size=8, italic=True, color='52525B')
            right_row += 1
            
            final = request.light_metrics_detrended['final']
            hrv_data = [
                ('ln(RMSSD₇₀):', f"{final.get('ln_rmssd70_detrended'):.3f}" if final.get('ln_rmssd70_detrended') else '—'),
                ('ln(SDNN₇₀):', f"{final.get('ln_sdnn70_detrended'):.3f}" if final.get('ln_sdnn70_detrended') else '—'),
                ('pNN50₇₀:', f"{final.get('pnn50_detrended'):.1f}%" if final.get('pnn50_detrended') is not None else '—'),
            ]
            
            for label, value in hrv_data:
                ws.cell(row=right_row, column=right_col, value=label).font = data_font
                ws.cell(row=right_row, column=right_col).fill = light_fill
                ws.cell(row=right_row, column=right_col+1, value=value).font = bold_data_font
                ws.cell(row=right_row, column=right_col+1).fill = light_fill
                right_row += 1
    
    # ==================== SHEET 2: SPONTANEOUS BF (Table 1) ====================
    if request.per_minute_data:
        ws_bf = wb.create_sheet('Spontaneous BF')
        
        ws_bf['A1'] = 'Spontaneous Activity - Beat Frequency'
        ws_bf['A1'].font = Font(bold=True, size=12)
        
        ws_bf['A2'] = 'Table 1 | Per-Minute Beat Frequency Data'
        ws_bf['A2'].font = Font(bold=True, size=10)
        
        # Match PDF columns exactly
        headers = ['Window (min)', 'Mean BF (bpm)', 'Mean NN (ms)']
        for col, header in enumerate(headers, 1):
            cell = ws_bf.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = emerald_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Get baseline and drug readout windows for highlighting
        baseline_window = None
        drug_window = None
        
        # Get baseline window
        if request.baseline_enabled and request.baseline:
            baseline_range = request.baseline.get('baseline_bf_range')
            if baseline_range is not None:
                try:
                    baseline_range_str = str(baseline_range).replace(' min', '').strip()
                    if '-' in baseline_range_str:
                        baseline_window = baseline_range_str
                    else:
                        bmin = int(float(baseline_range_str))
                        baseline_window = f"{bmin}-{bmin+1}"
                except (ValueError, TypeError):
                    baseline_window = str(baseline_range).replace(' min', '').strip()
            
            if baseline_window is None:
                bf_min = request.baseline.get('baseline_bf_minute')
                if bf_min is not None:
                    try:
                        bf_min = int(float(bf_min))
                        baseline_window = f"{bf_min}-{bf_min+1}"
                    except (ValueError, TypeError):
                        pass
        
        # Get drug readout window
        if request.drug_readout_enabled or request.drug_readout or request.drug_readout_settings:
            if request.drug_readout_settings:
                settings_bf = request.drug_readout_settings.get('bfReadoutMinute')
                if settings_bf not in (None, ''):
                    try:
                        perf_start = 0
                        perf_delay = 0
                        if request.all_drugs and len(request.all_drugs) > 0:
                            drug = request.all_drugs[0]
                            perf_start = int(float(drug.get('start', 0) or 0))
                            perf_delay = int(float(drug.get('delay', 0) or 0))
                        bf_min = int(float(settings_bf)) + perf_start + perf_delay
                        drug_window = f"{bf_min}-{bf_min+1}"
                    except (ValueError, TypeError):
                        pass
            
            if drug_window is None and request.drug_readout:
                bf_min = request.drug_readout.get('bf_minute')
                if bf_min is not None:
                    try:
                        bf_min = int(float(bf_min))
                        drug_window = f"{bf_min}-{bf_min+1}"
                    except (ValueError, TypeError):
                        pass
        
        row = 5
        for pm in request.per_minute_data:
            minute_val = pm.get('minute', '')
            # Format window as "X-X+1"
            try:
                minute_num = int(str(minute_val).split('-')[0]) if '-' in str(minute_val) else int(float(minute_val)) if minute_val != '' else None
                window_str = f"{minute_num}-{minute_num+1}" if minute_num is not None else str(minute_val)
            except (ValueError, TypeError):
                window_str = str(minute_val)
            
            bf_val = pm.get('mean_bf') or pm.get('avg_bf')
            nn_val = pm.get('mean_nn') or pm.get('avg_nn')
            
            # Convert to float
            try:
                bf_float = float(bf_val) if bf_val is not None else None
            except (ValueError, TypeError):
                bf_float = None
            try:
                nn_float = float(nn_val) if nn_val is not None else None
            except (ValueError, TypeError):
                nn_float = None
            
            data_row = [window_str, fmt(bf_float, 1), fmt(nn_float, 1)]
            
            # Determine if this row should be highlighted
            is_baseline = baseline_window and window_str == baseline_window
            is_drug = drug_window and window_str == drug_window
            
            for col, value in enumerate(data_row, 1):
                cell = ws_bf.cell(row=row, column=col, value=value)
                cell.font = bold_data_font if (is_baseline or is_drug) else data_font
                cell.border = thin_border
                cell.alignment = center_align
                if is_baseline:
                    cell.fill = baseline_fill
                elif is_drug:
                    cell.fill = drug_fill
            row += 1
        
        for col in range(1, 4):
            ws_bf.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== SHEET 3: SPONTANEOUS HRV (Table 2) ====================
    if request.hrv_windows:
        ws_hrv = wb.create_sheet('Spontaneous HRV')
        
        ws_hrv['A1'] = 'Spontaneous Activity - Heart Rate Variability'
        ws_hrv['A1'].font = Font(bold=True, size=12)
        
        ws_hrv['A2'] = 'Table 2 | Per-Three Minutes HRV Data'
        ws_hrv['A2'].font = Font(bold=True, size=10)
        
        # Match PDF columns exactly
        headers = ['Window', 'ln(RMSSD₇₀)', 'RMSSD₇₀', 'ln(SDNN₇₀)', 'SDNN', 'pNN50₇₀', 'BF']
        for col, header in enumerate(headers, 1):
            cell = ws_hrv.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = emerald_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Get baseline and drug readout minutes for highlighting
        baseline_minute = None
        drug_minute = None
        
        if request.baseline_enabled and request.baseline:
            baseline_minute = request.baseline.get('baseline_hrv_minute')
            if baseline_minute is not None:
                try:
                    baseline_minute = int(float(baseline_minute))
                except (ValueError, TypeError):
                    baseline_minute = None
        
        if request.drug_readout_enabled or request.drug_readout or request.drug_readout_settings:
            if request.drug_readout_settings:
                settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
                if settings_hrv not in (None, ''):
                    try:
                        perf_start = 0
                        perf_delay = 0
                        if request.all_drugs and len(request.all_drugs) > 0:
                            drug = request.all_drugs[0]
                            perf_start = int(float(drug.get('start', 0) or 0))
                            perf_delay = int(float(drug.get('delay', 0) or 0))
                        drug_minute = int(float(settings_hrv)) + perf_start + perf_delay
                    except (ValueError, TypeError):
                        pass
            
            if drug_minute is None and request.drug_readout:
                drug_minute = request.drug_readout.get('hrv_minute')
                if drug_minute is not None:
                    try:
                        drug_minute = int(float(drug_minute))
                    except (ValueError, TypeError):
                        drug_minute = None
        
        row = 5
        for w in request.hrv_windows:
            window = w.get('window', '')
            minute = w.get('minute', 0)
            
            # Convert minute to int for comparison
            try:
                if isinstance(minute, (int, float)):
                    minute_num = int(minute)
                else:
                    minute_str = str(minute)
                    minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(float(minute_str))
            except (ValueError, TypeError):
                minute_num = None
            
            sdnn = w.get('sdnn')
            try:
                sdnn_float = float(sdnn) if sdnn is not None else None
                ln_sdnn = np.log(sdnn_float) if sdnn_float and sdnn_float > 0 else None
            except (ValueError, TypeError):
                sdnn_float = None
                ln_sdnn = None
            
            data_row = [
                window,
                fmt(w.get('ln_rmssd70'), 3),
                fmt(w.get('rmssd70'), 1),
                fmt(ln_sdnn, 3),
                fmt(sdnn_float, 1),
                fmt(w.get('pnn50'), 1) if w.get('pnn50') is not None else '0.0',
                fmt(w.get('mean_bf'), 1),
            ]
            
            # Determine if this row should be highlighted
            is_baseline = baseline_minute is not None and minute_num == baseline_minute
            is_drug = drug_minute is not None and minute_num == drug_minute
            
            for col, value in enumerate(data_row, 1):
                cell = ws_hrv.cell(row=row, column=col, value=value)
                cell.font = bold_data_font if (is_baseline or is_drug) else data_font
                cell.border = thin_border
                cell.alignment = center_align
                if is_baseline:
                    cell.fill = baseline_fill
                elif is_drug:
                    cell.fill = drug_fill
            row += 1
        
        for col in range(1, 8):
            ws_hrv.column_dimensions[get_column_letter(col)].width = 13
    
    # ==================== SHEET 4: LIGHT HRA (Table 3) ====================
    if request.light_enabled and request.light_response:
        valid_responses = [r for r in request.light_response if r]
        if valid_responses:
            ws_hra = wb.create_sheet('Light HRA')
            
            ws_hra['A1'] = 'Light Stimulus - Heart Rate Adaptation'
            ws_hra['A1'].font = Font(bold=True, size=12)
            
            ws_hra['A2'] = 'Table 3 | Per-Stimulus HRA Data'
            ws_hra['A2'].font = Font(bold=True, size=10)
            
            # Match PDF columns exactly
            headers = ['Stim', 'Baseline BF', 'Avg BF', 'Peak BF', 'Peak %', '1st TTP (s)', 'TTP (s)', 'BF Rec', 'Rec %', 'Amp. BF', 'RoC (1/min)']
            for col, header in enumerate(headers, 1):
                cell = ws_hra.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = amber_fill
                cell.border = thin_border
                cell.alignment = center_align
            
            row = 5
            for idx, resp in enumerate(valid_responses, 1):
                # For 1st TTP, show value for stim 1, 0.0 for others
                first_ttp_val = resp.get('first_ttp_sec')
                first_ttp_str = f"{first_ttp_val:.1f}" if first_ttp_val is not None else "0.0"
                
                data_row = [
                    str(idx),
                    fmt(resp.get('baseline_bf'), 1),
                    fmt(resp.get('avg_bf'), 1),
                    fmt(resp.get('peak_bf'), 1),
                    fmt(resp.get('peak_norm_pct'), 1),
                    first_ttp_str,
                    fmt(resp.get('time_to_peak_sec'), 1),
                    fmt(resp.get('bf_end'), 1),
                    fmt(resp.get('bf_end_pct'), 1),
                    fmt(resp.get('amplitude'), 1),
                    fmt(resp.get('rate_of_change'), 3),
                ]
                for col, value in enumerate(data_row, 1):
                    cell = ws_hra.cell(row=row, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = center_align
                    if col >= 2:
                        cell.fill = light_fill
                row += 1
            
            # Average row
            if len(valid_responses) > 1:
                def safe_avg(key):
                    vals = [r.get(key) for r in valid_responses if r.get(key) is not None]
                    return np.mean(vals) if vals else None
                
                first_ttp_vals = [r.get('first_ttp_sec', 0) for r in valid_responses]
                first_ttp_avg = np.mean(first_ttp_vals) if first_ttp_vals else 0
                
                avg_row = [
                    'Avg',
                    fmt(safe_avg('baseline_bf'), 1),
                    fmt(safe_avg('avg_bf'), 1),
                    fmt(safe_avg('peak_bf'), 1),
                    fmt(safe_avg('peak_norm_pct'), 1),
                    f"{first_ttp_avg:.1f}",
                    fmt(safe_avg('time_to_peak_sec'), 1),
                    fmt(safe_avg('bf_end'), 1),
                    fmt(safe_avg('bf_end_pct'), 1),
                    fmt(safe_avg('amplitude'), 1),
                    fmt(safe_avg('rate_of_change'), 3),
                ]
                for col, value in enumerate(avg_row, 1):
                    cell = ws_hra.cell(row=row, column=col, value=value)
                    cell.font = avg_font
                    cell.fill = avg_fill
                    cell.border = thin_border
                    cell.alignment = center_align
            
            for col in range(1, 12):
                ws_hra.column_dimensions[get_column_letter(col)].width = 10
    
    # ==================== SHEET 5: LIGHT CORRECTED HRV (Table 4) ====================
    if request.light_enabled and request.light_metrics_detrended:
        # Support both 'per_stim' and 'per_pulse' keys
        per_stim = request.light_metrics_detrended.get('per_stim') or request.light_metrics_detrended.get('per_pulse', [])
        final = request.light_metrics_detrended.get('final', {})
        
        if per_stim or final:
            ws_corr = wb.create_sheet('Light Corrected HRV')
            
            ws_corr['A1'] = 'Light Stimulus - Corrected HRV'
            ws_corr['A1'].font = Font(bold=True, size=12)
            
            ws_corr['A2'] = 'Table 4 | Per-Stimulus Detrended HRV Data'
            ws_corr['A2'].font = Font(bold=True, size=10)
            
            # Match PDF columns exactly
            headers = ['Stim', 'ln(RMSSD₇₀)', 'RMSSD₇₀', 'ln(SDNN₇₀)', 'SDNN', 'pNN50₇₀']
            for col, header in enumerate(headers, 1):
                cell = ws_corr.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = amber_fill
                cell.border = thin_border
                cell.alignment = center_align
            
            row = 5
            # Always show at least 5 stims before the median (matching PDF)
            num_stims = max(5, len(per_stim))
            for i in range(num_stims):
                s = per_stim[i] if i < len(per_stim) else None
                has_data = s and (s.get('ln_rmssd70_detrended') is not None or 
                                 s.get('rmssd70_detrended') is not None or
                                 s.get('ln_sdnn70_detrended') is not None)
                
                if has_data:
                    data_row = [
                        str(i + 1),
                        fmt(s.get('ln_rmssd70_detrended'), 3),
                        fmt(s.get('rmssd70_detrended'), 3),
                        fmt(s.get('ln_sdnn70_detrended'), 3),
                        fmt(s.get('sdnn_detrended'), 3),
                        fmt(s.get('pnn50_detrended'), 1),
                    ]
                else:
                    data_row = [str(i + 1), '—', '—', '—', '—', '—']
                
                for col, value in enumerate(data_row, 1):
                    cell = ws_corr.cell(row=row, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = center_align
                    if col >= 2:
                        cell.fill = light_fill
                row += 1
            
            # Median row
            if final:
                median_row = [
                    'Median',
                    fmt(final.get('ln_rmssd70_detrended'), 3),
                    fmt(final.get('rmssd70_detrended'), 3),
                    fmt(final.get('ln_sdnn70_detrended'), 3),
                    fmt(final.get('sdnn_detrended'), 3),
                    fmt(final.get('pnn50_detrended'), 1),
                ]
                for col, value in enumerate(median_row, 1):
                    cell = ws_corr.cell(row=row, column=col, value=value)
                    cell.font = avg_font
                    cell.fill = avg_fill
                    cell.border = thin_border
                    cell.alignment = center_align
            
            for col in range(1, 7):
                ws_corr.column_dimensions[get_column_letter(col)].width = 13
    
    # ==================== SHEET 6: PER-BEAT DATA (kept beats only) ====================
    if request.per_beat_data:
        kept_beats = [b for b in request.per_beat_data if b.get('status') == 'kept']
        if kept_beats:
            ws_beat = wb.create_sheet('Per-Beat')
            
            ws_beat['A1'] = 'Per-Beat Data (Kept Beats Only)'
            ws_beat['A1'].font = Font(bold=True, size=12)
            
            headers = ['Beat #', 'Time (min)', 'BF Filtered (bpm)', 'NN Filtered (ms)']
            for col, header in enumerate(headers, 1):
                cell = ws_beat.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = emerald_fill
                cell.border = thin_border
                cell.alignment = center_align
            
            row = 4
            for i, beat in enumerate(kept_beats, 1):
                data_row = [
                    i,
                    round(beat.get('time_min', 0), 4) if beat.get('time_min') is not None else None,
                    round(beat.get('bf_bpm', 0), 1) if beat.get('bf_bpm') is not None else None,
                    round(beat.get('nn_ms', 0), 1) if beat.get('nn_ms') is not None else None,
                ]
                for col, value in enumerate(data_row, 1):
                    cell = ws_beat.cell(row=row, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = center_align
                row += 1
            
            for col in range(1, 5):
                ws_beat.column_dimensions[get_column_letter(col)].width = 18
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
def create_nature_csv(request):
    """Create a clean CSV export matching PDF structure"""
    import csv
    
    buf = io.StringIO()
    writer = csv.writer(buf)
    
    # ==================== HEADER ====================
    title = request.recording_name or request.filename or 'Recording Analysis'
    writer.writerow([title])
    writer.writerow(['Electrophysiology Analysis Report by NEHER'])
    if request.recording_date:
        writer.writerow([f'Recording Date: {request.recording_date}'])
    writer.writerow([])
    
    # ==================== RECORDING INFO ====================
    writer.writerow(['=== RECORDING INFO ==='])
    if request.original_filename:
        writer.writerow(['Original File', request.original_filename])
    if request.recording_date:
        writer.writerow(['Recording Date', request.recording_date])
    if request.summary:
        if 'Total Beats' in request.summary:
            writer.writerow(['Total Beats', request.summary['Total Beats']])
        if 'Kept Beats' in request.summary:
            writer.writerow(['Kept Beats', request.summary['Kept Beats']])
        if 'Filter Range' in request.summary:
            writer.writerow(['Filter Range', request.summary['Filter Range']])
    writer.writerow([])
    
    # ==================== TISSUE INFO ====================
    if request.organoid_info:
        writer.writerow(['=== TISSUE INFO ==='])
        for idx, org in enumerate(request.organoid_info):
            if len(request.organoid_info) > 1:
                writer.writerow([f'--- Sample {idx + 1} ---'])
            if org.get('cell_type'):
                cell_type = org.get('other_cell_type') if org.get('cell_type') == 'Other' else org.get('cell_type')
                writer.writerow(['Cell Type', cell_type or ''])
            if org.get('line_name'):
                writer.writerow(['Line', org.get('line_name')])
            if org.get('passage_number'):
                writer.writerow(['Passage', org.get('passage_number')])
            if org.get('age_at_recording') is not None:
                writer.writerow(['Age at Recording', f"{org.get('age_at_recording')} days"])
            if org.get('transfection'):
                trans = org['transfection']
                if trans.get('name'):
                    writer.writerow(['Transfection', trans.get('name')])
                if trans.get('days_since_transfection') is not None:
                    writer.writerow(['Days Post-Transfection', trans.get('days_since_transfection')])
        if request.days_since_fusion is not None:
            writer.writerow(['Days Since Fusion', request.days_since_fusion])
        writer.writerow([])
    
    # ==================== DRUG PERFUSION ====================
    if request.all_drugs and len(request.all_drugs) > 0:
        writer.writerow(['=== DRUG PERFUSION ==='])
        for drug in request.all_drugs:
            writer.writerow(['Drug', drug.get('name', 'Drug')])
            if drug.get('concentration'):
                writer.writerow(['Concentration', f"{drug.get('concentration')}µM"])
            writer.writerow(['Perf. Start', f"{drug.get('start', 0)} min"])
            writer.writerow(['Perf. Delay', f"{drug.get('delay', 0)} min"])
            # Perf. Time = HRV readout minute if available, otherwise start + delay
            perf_start = drug.get('start', 0) or 0
            perf_delay = drug.get('delay', 0) or 0
            perf_time = perf_start + perf_delay  # Default
            # Try to get HRV readout minute from drug_readout_settings
            if request.drug_readout_settings:
                settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
                if request.drug_readout_settings.get('enableHrvReadout') and settings_hrv not in (None, ''):
                    try:
                        perf_time = int(float(settings_hrv)) + perf_start + perf_delay
                    except (ValueError, TypeError):
                        pass
            writer.writerow(['Perf. Time', f"{perf_time} min"])
            perf_end = drug.get('end')
            writer.writerow(['Perf. End', f"{perf_end} min" if perf_end is not None else '—'])
        writer.writerow([])
    
    # ==================== LIGHT STIMULATION ====================
    if request.light_enabled:
        writer.writerow(['=== LIGHT STIMULATION ==='])
        writer.writerow(['Status', 'Enabled'])
        if request.light_stim_count and request.light_stim_count > 0:
            writer.writerow(['Stims Detected', request.light_stim_count])
        if request.light_pulses and len(request.light_pulses) > 0:
            first_pulse = request.light_pulses[0]
            light_start = first_pulse.get('start_min')
            if light_start is not None:
                writer.writerow(['Stims Start', f"{light_start:.2f} min"])
        if request.light_params:
            if request.light_params.get('pulseDuration') is not None:
                writer.writerow(['Stim Duration', f"{request.light_params.get('pulseDuration')} sec"])
            if request.light_params.get('interval'):
                interval_val = request.light_params.get('interval')
                interval_display_map = {
                    'decreasing': '60s-30s-20s-10s',
                    '60': 'Uniform 60s',
                    '30': 'Uniform 30s',
                }
                interval_display = interval_display_map.get(str(interval_val), str(interval_val))
                writer.writerow(['Inter-stimuli intervals', interval_display])
        writer.writerow([])
    
    # ==================== BASELINE READOUT ====================
    if request.baseline_enabled and request.baseline:
        writer.writerow(['=== BASELINE READOUT ==='])
        baseline = request.baseline
        bf_val = baseline.get('baseline_bf')
        writer.writerow(['Mean BF', f"{bf_val:.1f} bpm" if bf_val else '—'])
        ln_rmssd = baseline.get('baseline_ln_rmssd70')
        writer.writerow(['ln(RMSSD70)', f"{ln_rmssd:.3f}" if ln_rmssd else '—'])
        sdnn = baseline.get('baseline_sdnn')
        ln_sdnn = np.log(sdnn) if sdnn and sdnn > 0 else None
        writer.writerow(['ln(SDNN70)', f"{ln_sdnn:.3f}" if ln_sdnn else '—'])
        pnn50 = baseline.get('baseline_pnn50')
        writer.writerow(['pNN50_70', f"{pnn50:.1f}%" if pnn50 is not None else '—'])
        writer.writerow([])
    
    # ==================== DRUG READOUT ====================
    if request.drug_readout_enabled and request.all_drugs and len(request.all_drugs) > 0:
        writer.writerow(['=== DRUG READOUT ==='])
        
        # Get drug readout data (same logic as PDF/Excel)
        drug_bf = None
        drug_hrv_data = None
        drug_bf_minute = None
        drug_hrv_minute = None
        
        if request.drug_readout:
            drug_bf_minute = request.drug_readout.get('bf_minute')
            drug_hrv_minute = request.drug_readout.get('hrv_minute')
        
        if request.drug_readout_settings:
            settings_bf = request.drug_readout_settings.get('bfReadoutMinute')
            settings_hrv = request.drug_readout_settings.get('hrvReadoutMinute')
            perf_start = 0
            perf_delay = 0
            if request.all_drugs and len(request.all_drugs) > 0:
                drug = request.all_drugs[0]
                perf_start = drug.get('start', 0) or 0
                perf_delay = drug.get('delay', 0) or 0
            
            if request.drug_readout_settings.get('enableBfReadout') and settings_bf not in (None, ''):
                try:
                    drug_bf_minute = int(float(settings_bf)) + perf_start + perf_delay
                except (ValueError, TypeError):
                    pass
            if request.drug_readout_settings.get('enableHrvReadout') and settings_hrv not in (None, ''):
                try:
                    drug_hrv_minute = int(float(settings_hrv)) + perf_start + perf_delay
                except (ValueError, TypeError):
                    pass
        
        if drug_bf_minute is not None and request.per_minute_data:
            for pm in request.per_minute_data:
                try:
                    minute_str = str(pm.get('minute', ''))
                    minute_num = int(minute_str.split('-')[0]) if '-' in minute_str else int(float(minute_str))
                    if minute_num == drug_bf_minute:
                        drug_bf = pm.get('mean_bf')
                        break
                except (ValueError, TypeError):
                    pass
        
        if drug_hrv_minute is not None and request.hrv_windows:
            for w in request.hrv_windows:
                try:
                    w_minute = w.get('minute')
                    if isinstance(w_minute, (int, float)):
                        w_minute_num = int(w_minute)
                    else:
                        w_minute_str = str(w_minute)
                        w_minute_num = int(w_minute_str.split('-')[0]) if '-' in w_minute_str else int(float(w_minute_str))
                    if w_minute_num == drug_hrv_minute:
                        drug_hrv_data = w
                        if drug_bf is None and w.get('mean_bf'):
                            drug_bf = w.get('mean_bf')
                        break
                except (ValueError, TypeError):
                    pass
        
        writer.writerow(['Mean BF', f"{drug_bf:.1f} bpm" if drug_bf else '—'])
        if drug_hrv_data:
            ln_rmssd = drug_hrv_data.get('ln_rmssd70')
            writer.writerow(['ln(RMSSD70)', f"{ln_rmssd:.3f}" if ln_rmssd else '—'])
            sdnn = drug_hrv_data.get('sdnn')
            ln_sdnn = np.log(sdnn) if sdnn and sdnn > 0 else None
            writer.writerow(['ln(SDNN70)', f"{ln_sdnn:.3f}" if ln_sdnn else '—'])
            pnn50 = drug_hrv_data.get('pnn50')
            writer.writerow(['pNN50_70', f"{pnn50:.1f}%" if pnn50 is not None else '—'])
        writer.writerow([])
    
    # ==================== LIGHT READOUT ====================
    if request.light_enabled and (request.light_response or request.light_metrics_detrended):
        writer.writerow(['=== LIGHT READOUT ==='])
        
        if request.light_response:
            valid = [r for r in request.light_response if r]
            if valid:
                baseline_bf_vals = [r.get('baseline_bf') for r in valid if r.get('baseline_bf')]
                baseline_bf = np.mean(baseline_bf_vals) if baseline_bf_vals else None
                avg_bf = np.mean([r.get('avg_bf', 0) for r in valid if r.get('avg_bf')])
                peak_bf = np.mean([r.get('peak_bf', 0) for r in valid if r.get('peak_bf')])
                peak_norm_vals = [r.get('peak_norm_pct') for r in valid if r.get('peak_norm_pct') is not None]
                peak_norm = np.mean(peak_norm_vals) if peak_norm_vals else None
                amplitude_vals = [r.get('amplitude') for r in valid if r.get('amplitude')]
                amplitude = np.mean(amplitude_vals) if amplitude_vals else None
                ttp_vals = [r.get('time_to_peak_sec') for r in valid if r.get('time_to_peak_sec') is not None]
                ttp = np.mean(ttp_vals) if ttp_vals else None
                ttp_1st = valid[0].get('time_to_peak_sec') if valid else None
                roc_vals = [r.get('rate_of_change') for r in valid if r.get('rate_of_change') is not None]
                roc = np.mean(roc_vals) if roc_vals else None
                recovery_bf_vals = [r.get('bf_end') for r in valid if r.get('bf_end')]
                recovery_bf = np.mean(recovery_bf_vals) if recovery_bf_vals else None
                recovery_pct_vals = [r.get('bf_end_pct') for r in valid if r.get('bf_end_pct')]
                recovery_pct = np.mean(recovery_pct_vals) if recovery_pct_vals else None
                
                writer.writerow(['Baseline BF', f"{baseline_bf:.1f} bpm" if baseline_bf else '—'])
                writer.writerow(['Avg BF', f"{avg_bf:.1f} bpm"])
                writer.writerow(['Peak BF', f"{peak_bf:.1f} bpm"])
                writer.writerow(['Peak (Norm.)', f"{peak_norm:.1f}%" if peak_norm else '—'])
                writer.writerow(['Amplitude', f"{amplitude:.1f} bpm" if amplitude else '—'])
                writer.writerow(['Time to Peak (AVG)', f"{ttp:.1f} s" if ttp is not None else '—'])
                writer.writerow(['TTP (1st Stim)', f"{ttp_1st:.1f} s" if ttp_1st is not None else '—'])
                writer.writerow(['Rate of Change', f"{roc:.3f} 1/min" if roc else '—'])
                writer.writerow(['Recovery BF', f"{recovery_bf:.1f} bpm" if recovery_bf else '—'])
                writer.writerow(['Recovery %', f"{recovery_pct:.1f}%" if recovery_pct else '—'])
        
        # Corrected HRV
        if request.light_metrics_detrended and request.light_metrics_detrended.get('final'):
            writer.writerow([])
            writer.writerow(['--- Corrected HRV ---'])
            final = request.light_metrics_detrended['final']
            ln_rmssd = final.get('ln_rmssd70_detrended')
            writer.writerow(['ln(RMSSD70)', f"{ln_rmssd:.3f}" if ln_rmssd else '—'])
            ln_sdnn = final.get('ln_sdnn70_detrended')
            writer.writerow(['ln(SDNN70)', f"{ln_sdnn:.3f}" if ln_sdnn else '—'])
            pnn50 = final.get('pnn50_detrended')
            writer.writerow(['pNN50_70', f"{pnn50:.1f}%" if pnn50 is not None else '—'])
        writer.writerow([])
    
    # ==================== SPONTANEOUS BF TABLE ====================
    if request.per_minute_data:
        writer.writerow(['=== SPONTANEOUS BF DATA ==='])
        writer.writerow(['Window (min)', 'Mean BF (bpm)', 'Mean NN (ms)'])
        for pm in request.per_minute_data:
            minute_val = pm.get('minute', '')
            try:
                minute_num = int(str(minute_val).split('-')[0]) if '-' in str(minute_val) else int(minute_val)
                window_str = f"{minute_num}-{minute_num+1}"
            except (ValueError, TypeError):
                window_str = str(minute_val)
            
            bf_val = pm.get('mean_bf') or pm.get('avg_bf')
            nn_val = pm.get('mean_nn') or pm.get('avg_nn')
            
            writer.writerow([
                window_str,
                round(bf_val, 1) if bf_val else '',
                round(nn_val, 1) if nn_val else '',
            ])
        writer.writerow([])
    
    # ==================== SPONTANEOUS HRV TABLE ====================
    if request.hrv_windows:
        writer.writerow(['=== SPONTANEOUS HRV DATA ==='])
        writer.writerow(['Window', 'ln(RMSSD70)', 'RMSSD70', 'ln(SDNN70)', 'SDNN', 'pNN50_70', 'BF'])
        for w in request.hrv_windows:
            sdnn = w.get('sdnn')
            ln_sdnn = round(np.log(sdnn), 3) if sdnn and sdnn > 0 else ''
            writer.writerow([
                w.get('window', ''),
                round(w.get('ln_rmssd70', 0), 3) if w.get('ln_rmssd70') else '',
                round(w.get('rmssd70', 0), 1) if w.get('rmssd70') else '',
                ln_sdnn,
                round(sdnn, 1) if sdnn else '',
                round(w.get('pnn50', 0), 1) if w.get('pnn50') is not None else '0',
                round(w.get('mean_bf', 0), 1) if w.get('mean_bf') else '',
            ])
        writer.writerow([])
    
    # ==================== LIGHT HRA TABLE ====================
    if request.light_enabled and request.light_response:
        valid = [r for r in request.light_response if r]
        if valid:
            writer.writerow(['=== LIGHT-INDUCED HRA ==='])
            writer.writerow(['Stim', 'Baseline BF', 'Avg BF', 'Peak BF', 'Peak %', 'Amplitude', 'BF End', 'Recovery %', 'TTP (s)', 'RoC (1/min)'])
            for i, r in enumerate(valid):
                writer.writerow([
                    i + 1,
                    round(r.get('baseline_bf', 0), 1) if r.get('baseline_bf') else '',
                    round(r.get('avg_bf', 0), 1) if r.get('avg_bf') else '',
                    round(r.get('peak_bf', 0), 1) if r.get('peak_bf') else '',
                    round(r.get('peak_norm_pct', 0), 1) if r.get('peak_norm_pct') else '',
                    round(r.get('amplitude', 0), 1) if r.get('amplitude') is not None else '',
                    round(r.get('bf_end', 0), 1) if r.get('bf_end') else '',
                    round(r.get('bf_end_pct', 0), 1) if r.get('bf_end_pct') else '',
                    round(r.get('time_to_peak_sec', 0), 1) if r.get('time_to_peak_sec') is not None else '',
                    round(r.get('rate_of_change', 0), 3) if r.get('rate_of_change') is not None else '',
                ])
            
            # Add average row
            if len(valid) > 1:
                def safe_avg(key):
                    vals = [r.get(key) for r in valid if r.get(key) is not None]
                    return np.mean(vals) if vals else None
                
                writer.writerow([
                    'Avg',
                    round(safe_avg('baseline_bf'), 1) if safe_avg('baseline_bf') else '',
                    round(safe_avg('avg_bf'), 1) if safe_avg('avg_bf') else '',
                    round(safe_avg('peak_bf'), 1) if safe_avg('peak_bf') else '',
                    round(safe_avg('peak_norm_pct'), 1) if safe_avg('peak_norm_pct') else '',
                    round(safe_avg('amplitude'), 1) if safe_avg('amplitude') is not None else '',
                    round(safe_avg('bf_end'), 1) if safe_avg('bf_end') else '',
                    round(safe_avg('bf_end_pct'), 1) if safe_avg('bf_end_pct') else '',
                    round(safe_avg('time_to_peak_sec'), 1) if safe_avg('time_to_peak_sec') is not None else '',
                    round(safe_avg('rate_of_change'), 3) if safe_avg('rate_of_change') is not None else '',
                ])
            writer.writerow([])
    
    # ==================== CORRECTED HRV TABLE ====================
    if request.light_enabled and request.light_metrics_detrended:
        per_stim = request.light_metrics_detrended.get('per_stim') or request.light_metrics_detrended.get('per_pulse', [])
        final = request.light_metrics_detrended.get('final', {})
        
        if per_stim or final:
            writer.writerow(['=== LIGHT-INDUCED CORRECTED HRV ==='])
            writer.writerow(['Stim', 'ln(RMSSD70)', 'RMSSD70', 'ln(SDNN70)', 'SDNN', 'pNN50_70'])
            
            num_stims = max(5, len(per_stim))
            for i in range(num_stims):
                s = per_stim[i] if i < len(per_stim) else None
                has_data = s and (s.get('ln_rmssd70_detrended') is not None or 
                                 s.get('rmssd70_detrended') is not None)
                
                if has_data:
                    writer.writerow([
                        i + 1,
                        round(s.get('ln_rmssd70_detrended', 0), 3) if s.get('ln_rmssd70_detrended') is not None else '',
                        round(s.get('rmssd70_detrended', 0), 3) if s.get('rmssd70_detrended') is not None else '',
                        round(s.get('ln_sdnn70_detrended', 0), 3) if s.get('ln_sdnn70_detrended') is not None else '',
                        round(s.get('sdnn_detrended', 0), 3) if s.get('sdnn_detrended') is not None else '',
                        round(s.get('pnn50_detrended', 0), 1) if s.get('pnn50_detrended') is not None else '',
                    ])
                else:
                    writer.writerow([i + 1, '', '', '', '', ''])
            
            if final:
                writer.writerow([
                    'Median',
                    round(final.get('ln_rmssd70_detrended', 0), 3) if final.get('ln_rmssd70_detrended') else '',
                    round(final.get('rmssd70_detrended', 0), 3) if final.get('rmssd70_detrended') else '',
                    round(final.get('ln_sdnn70_detrended', 0), 3) if final.get('ln_sdnn70_detrended') else '',
                    round(final.get('sdnn_detrended', 0), 3) if final.get('sdnn_detrended') else '',
                    round(final.get('pnn50_detrended', 0), 1) if final.get('pnn50_detrended') is not None else '',
                ])
            writer.writerow([])
    
    # ==================== PER-BEAT DATA (kept beats only) ====================
    if request.per_beat_data:
        kept_beats = [b for b in request.per_beat_data if b.get('status') == 'kept']
        if kept_beats:
            writer.writerow(['=== PER-BEAT DATA ==='])
            writer.writerow(['Beat #', 'Time (min)', 'BF Filtered (bpm)', 'NN Filtered (ms)'])
            for i, beat in enumerate(kept_beats):
                writer.writerow([
                    i + 1,
                    round(beat.get('time_min', 0), 4) if beat.get('time_min') is not None else '',
                    round(beat.get('bf_bpm', 0), 1) if beat.get('bf_bpm') is not None else '',
                    round(beat.get('nn_ms', 0), 1) if beat.get('nn_ms') is not None else '',
                ])
    
    output = io.BytesIO()
    output.write(buf.getvalue().encode('utf-8'))
    output.seek(0)
    return output



def create_comparison_pdf(folder_name, comparison_data):
    """Create a comparison PDF using the bioptima style from single recording PDF - 7 pages"""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    import matplotlib.patches as mpatches
    import matplotlib.font_manager as fm
    
    # Register fonts - Optima for titles, Carlito for body text
    title_font = 'Optima'
    body_font = 'Carlito'
    
    available_fonts = [f.name for f in fm.fontManager.ttflist]
    if title_font not in available_fonts:
        title_font = 'DejaVu Sans'
    if body_font not in available_fonts:
        body_font = 'DejaVu Sans'
    
    plt.rcParams.update({
        'font.family': 'sans-serif',
        'font.sans-serif': [body_font, 'Carlito', 'DejaVu Sans', 'Arial'],
        'font.size': 9,
        'axes.labelsize': 9,
        'axes.titlesize': 10,
        'axes.linewidth': 0.8,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'legend.fontsize': 8,
        'figure.titlesize': 12,
        'axes.spines.top': False,
        'axes.spines.right': False,
    })
    
    data = comparison_data
    recordings = data.get('recordings', [])
    summary = data.get('summary', {})
    spont_averages = data.get('spontaneous_averages', {}).get('averages', {})
    hra_averages = data.get('light_hra_averages', {}).get('averages', {})
    hrv_averages = data.get('light_hrv_averages', {}).get('averages', {})
    
    # Sort recordings alphabetically
    recordings = sorted(recordings, key=lambda x: x.get('name', '').lower())
    
    buf = io.BytesIO()
    
    # Color scheme matching bioptima style
    COLORS = {
        'dark': '#18181b',
        'header_blue': '#5ba4c9',
        'gray': '#6b7280',
        'line': '#374151',
        'emerald': '#10b981',
        'purple': '#a855f7',
        'amber': '#f59e0b',
        'sky': '#0ea5e9',
        'cyan': '#06b6d4',
    }
    
    TINTS = {
        'baseline': '#E0F2FE',
        'drug': '#e8d5f5',
        'light': '#FEF3C7',
        'avg': '#d1fae5',
    }
    
    def fmt(val, dec=2):
        if val is None:
            return '—'
        try:
            if dec == 0:
                return f"{float(val):.0f}"
            elif dec == 1:
                return f"{float(val):.1f}"
            elif dec == 3:
                return f"{float(val):.3f}"
            elif dec == 4:
                return f"{float(val):.4f}"
            return f"{float(val):.{dec}f}"
        except:
            return '—'
    
    def fmt_age_range(age_dict):
        if not age_dict or age_dict.get('min') is None:
            return '—'
        return f"{age_dict.get('min')} - {age_dict.get('max')} days"
    
    def add_page_header(fig, section_name):
        """Add header in bioptima style: NEHER section_name"""
        fig.text(0.08, 0.96, 'NEHER', fontsize=12, fontweight='bold', color=COLORS['dark'],
                fontfamily=title_font)
        fig.text(0.16, 0.96, section_name, fontsize=12, fontweight='normal', color=COLORS['gray'],
                fontfamily=title_font)
    
    def add_page_footer(fig, page_num, total_pages):
        """Add footer: p. XX | Folder Name | Folder Comparison Report by NEHER"""
        fig.text(0.08, 0.025, f'p. {page_num}', fontsize=10, fontweight='bold', color=COLORS['dark'],
                fontfamily=body_font)
        fig.text(0.125, 0.025, f'|  {folder_name}  |  Comparison Report by NEHER', fontsize=10, color=COLORS['gray'],
                fontfamily=body_font)
    
    def draw_header(fig, x, y, text, color, width=0.38):
        """Draw section header with colored background - bioptima style"""
        fig.add_artist(mpatches.Rectangle(
            (x, y - 0.012), width, 0.024,
            facecolor=color, edgecolor='none', transform=fig.transFigure
        ))
        fig.text(x + 0.01, y, text, fontsize=9, fontweight='bold', color='white',
                fontfamily=body_font, va='center')
        return y - 0.028
    
    def draw_separator(fig, x, y, width=0.38):
        """Draw horizontal separator line"""
        fig.add_artist(plt.Line2D([x, x + width], [y, y], color=COLORS['line'], 
                      linewidth=0.5, transform=fig.transFigure))
        return y - 0.008
    
    def draw_row(fig, x, y, label, value, bg_color=None, width=0.38, label_width=0.18):
        """Draw a data row with label and value"""
        if bg_color:
            fig.add_artist(mpatches.Rectangle(
                (x, y - 0.008), width, 0.020,
                facecolor=bg_color, edgecolor='none', transform=fig.transFigure
            ))
        fig.text(x + 0.01, y, label, fontsize=8, color=COLORS['gray'], fontfamily=body_font, va='center')
        fig.text(x + label_width, y, str(value), fontsize=8, fontweight='bold', color=COLORS['dark'],
                fontfamily=body_font, va='center')
        return y - 0.020
    
    def norm_val(val, avg):
        if val is None or avg == 0:
            return None
        return 100 * val / avg
    
    def extract_short_name(full_name):
        """Extract FX or CX from the end of the recording name"""
        if not full_name:
            return '—'
        # Look for pattern like F1, F2, C1, C2 at the end
        import re
        match = re.search(r'[FC]\d+$', full_name)
        if match:
            return match.group()
        # If no match, try to get last part after dash or slash
        parts = full_name.replace('/', '-').split('-')
        return parts[-1] if parts else full_name[:10]
    
    def parse_drug_info(drug_info):
        """Parse drug info to extract name and concentration"""
        if not drug_info:
            return None
        if isinstance(drug_info, dict):
            name = drug_info.get('name', '')
            conc = drug_info.get('concentration', '')
            if name:
                return f"{name} ({conc})" if conc else name
            return None
        if isinstance(drug_info, list):
            results = []
            for d in drug_info:
                parsed = parse_drug_info(d)
                if parsed:
                    results.append(parsed)
            return results if results else None
        # String type
        drug_str = str(drug_info).strip()
        if drug_str.lower() in ['no drug', 'no', 'none', '—', '-', '']:
            return None
        return drug_str
    
    total_pages = 5  # Summary, Metadata, Spont+Norm, HRA+Norm, HRV
    
    with PdfPages(buf) as pdf:
        
        # ==================== PAGE 1: SUMMARY ====================
        fig1 = plt.figure(figsize=(8.5, 11))  # US Letter Portrait
        fig1.patch.set_facecolor('white')
        
        add_page_header(fig1, 'summary')
        
        # Main title - wrap if needed
        title_text = folder_name
        title_fontsize = 24 if len(title_text) > 30 else 28
        if len(title_text) > 40:
            # Split into two lines
            mid = len(title_text) // 2
            # Find a good split point near middle
            split_idx = title_text.rfind(' ', 0, mid + 10)
            if split_idx == -1:
                split_idx = mid
            line1 = title_text[:split_idx]
            line2 = title_text[split_idx:].strip()
            fig1.text(0.08, 0.91, line1, ha='left', va='top', fontsize=22, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            fig1.text(0.08, 0.875, line2, ha='left', va='top', fontsize=22, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            separator_y = 0.845
            first_section_y = 0.81
        else:
            fig1.text(0.08, 0.90, title_text, ha='left', va='top', fontsize=title_fontsize, fontweight='bold', 
                     color=COLORS['dark'], fontfamily=title_font)
            separator_y = 0.865
            first_section_y = 0.83
        
        fig1.add_artist(plt.Line2D([0.08, 0.92], [separator_y, separator_y], color=COLORS['dark'], linewidth=1.0, transform=fig1.transFigure))
        
        # Two columns layout
        left_x = 0.08
        right_x = 0.52
        col_width = 0.40
        
        # LEFT COLUMN: Folder Overview, Age Ranges, Parameters
        y = draw_header(fig1, left_x, first_section_y, 'FOLDER OVERVIEW', COLORS['dark'], width=col_width)
        y = draw_row(fig1, left_x, y, 'Recordings:', str(summary.get('recording_count', len(recordings))), width=col_width)
        y = draw_row(fig1, left_x, y, 'Date Created:', datetime.now().strftime('%Y-%m-%d'), width=col_width)
        y -= 0.015
        
        y = draw_header(fig1, left_x, y, 'AGE RANGES', COLORS['gray'], width=col_width)
        hspo_range = summary.get('hspo_age_range', {})
        hco_range = summary.get('hco_age_range', {})
        fusion_range = summary.get('fusion_age_range', {})
        y = draw_row(fig1, left_x, y, 'hSpOs:', f"{fmt_age_range(hspo_range)} (n={hspo_range.get('n', 0)})", width=col_width)
        y = draw_row(fig1, left_x, y, 'hCOs:', f"{fmt_age_range(hco_range)} (n={hco_range.get('n', 0)})", width=col_width)
        y = draw_row(fig1, left_x, y, 'Fusion:', f"{fmt_age_range(fusion_range)} (n={fusion_range.get('n', 0)})", width=col_width)
        y -= 0.015
        
        # Parameters section
        y = draw_header(fig1, left_x, y, 'PARAMETERS', COLORS['gray'], width=col_width)
        
        # Drug Used: Parse properly and include concentration unit in parentheses
        all_drugs = []
        drug_concentration_unit = ''
        for r in recordings:
            drug_info_raw = r.get('drug_info')
            if isinstance(drug_info_raw, dict):
                conc_unit = drug_info_raw.get('concentration_unit', '') or drug_info_raw.get('unit', '') or drug_info_raw.get('conc_unit', '')
                if conc_unit and not drug_concentration_unit:
                    drug_concentration_unit = conc_unit
            # Also check if drug_info is a string with concentration like "Nepi 100nM"
            elif isinstance(drug_info_raw, str) and drug_info_raw:
                # Try to extract unit from string like "100nM", "10uM", etc.
                import re
                unit_match = re.search(r'\d+\s*(nM|µM|uM|mM|pM)', drug_info_raw, re.IGNORECASE)
                if unit_match and not drug_concentration_unit:
                    drug_concentration_unit = unit_match.group(1)
            parsed = parse_drug_info(drug_info_raw)
            if parsed:
                if isinstance(parsed, list):
                    all_drugs.extend(parsed)
                else:
                    all_drugs.append(parsed)
        
        # Default to µM if we have drugs but no unit found
        if all_drugs and not drug_concentration_unit:
            drug_concentration_unit = 'µM'
        
        if all_drugs:
            unique_drugs = list(dict.fromkeys(all_drugs))  # Preserve order, remove duplicates
            drug_text = unique_drugs[0]
            # Combine concentration and unit - e.g., "ruxolitinib (2)" + "µM" -> "ruxolitinib (2µM)"
            if drug_concentration_unit:
                # Check if drug_text already has a concentration in parentheses like "drug (2)"
                import re
                match = re.search(r'\(([^)]+)\)$', drug_text)
                if match:
                    # Replace "(2)" with "(2µM)"
                    conc_value = match.group(1)
                    drug_text = re.sub(r'\([^)]+\)$', f'({conc_value}{drug_concentration_unit})', drug_text)
                else:
                    drug_text = f"{drug_text} ({drug_concentration_unit})"
        else:
            drug_text = '—'
        
        y = draw_row(fig1, left_x, y, 'Drug Used:', drug_text, width=col_width)
        
        # Add additional drug lines if more than one
        if all_drugs and len(set(all_drugs)) > 1:
            unique_drugs = list(dict.fromkeys(all_drugs))
            for drug in unique_drugs[1:3]:  # Show up to 3 drugs total
                y = draw_row(fig1, left_x, y, '', drug, width=col_width)
        
        # Light Stim: Check if any recording has stim_duration and isi_structure
        light_used = False
        for r in recordings:
            # Check for stim_duration and isi_structure fields
            stim_dur = r.get('stim_duration')
            isi_struct = r.get('isi_structure')
            has_light = r.get('has_light_stim')
            
            if has_light or (stim_dur and isi_struct):
                light_used = True
                break
            
            # Fallback: check light_stim_info for time patterns
            light_info = r.get('light_stim_info', '')
            if isinstance(light_info, list):
                light_info = ' '.join(str(i) for i in light_info)
            light_info = str(light_info).strip()
            if light_info:
                light_lower = light_info.lower()
                if (re.search(r'\d+s', light_info) or 
                    'stim' in light_lower or 
                    'isi' in light_lower):
                    if 'no light' not in light_lower:
                        light_used = True
                        break
        
        light_text = 'Yes' if light_used else 'No'
        y = draw_row(fig1, left_x, y, 'Light Stim:', light_text, width=col_width)
        
        # RIGHT COLUMN: Spontaneous Activity & Light Stimulus readouts
        y_right = first_section_y
        
        y_right = draw_header(fig1, right_x, y_right, 'SPONTANEOUS ACTIVITY', COLORS['emerald'], width=col_width)
        
        fig1.text(right_x + 0.01, y_right, 'Baseline Readout', fontsize=7, fontstyle='italic', color='#71717a')
        y_right -= 0.018
        y_right = draw_row(fig1, right_x, y_right, 'Mean BF:', f"{fmt(spont_averages.get('baseline_bf'), 1)} bpm", TINTS['baseline'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'ln(RMSSD₇₀):', fmt(spont_averages.get('baseline_ln_rmssd70'), 3), TINTS['baseline'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'ln(SDNN₇₀):', fmt(spont_averages.get('baseline_ln_sdnn70'), 3), TINTS['baseline'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'pNN50₇₀:', f"{fmt(spont_averages.get('baseline_pnn50'), 1)}%", TINTS['baseline'], width=col_width)
        y_right -= 0.008
        
        fig1.text(right_x + 0.01, y_right, 'Drug Readout', fontsize=7, fontstyle='italic', color='#71717a')
        y_right -= 0.018
        y_right = draw_row(fig1, right_x, y_right, 'Mean BF:', f"{fmt(spont_averages.get('drug_bf'), 1)} bpm", TINTS['drug'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'ln(RMSSD₇₀):', fmt(spont_averages.get('drug_ln_rmssd70'), 3), TINTS['drug'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'ln(SDNN₇₀):', fmt(spont_averages.get('drug_ln_sdnn70'), 3), TINTS['drug'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'pNN50₇₀:', f"{fmt(spont_averages.get('drug_pnn50'), 1)}%", TINTS['drug'], width=col_width)
        y_right -= 0.015
        
        y_right = draw_header(fig1, right_x, y_right, 'LIGHT STIMULUS', COLORS['amber'], width=col_width)
        
        fig1.text(right_x + 0.01, y_right, 'Heart Rate Adaptation (HRA)', fontsize=7, fontstyle='italic', color='#71717a')
        y_right -= 0.018
        y_right = draw_row(fig1, right_x, y_right, 'Baseline BF:', f"{fmt(hra_averages.get('light_baseline_bf'), 1)} bpm", TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'Peak BF:', f"{fmt(hra_averages.get('light_peak_bf'), 1)} bpm", TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'Peak (Norm.):', f"{fmt(hra_averages.get('light_peak_norm'), 1)}%", TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'Amplitude:', f"{fmt(hra_averages.get('light_amplitude'), 1)} bpm", TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'TTP (Avg):', f"{fmt(hra_averages.get('light_ttp_avg'), 1)} s", TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'Rate of Change:', fmt(hra_averages.get('light_roc'), 4), TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'Recovery %:', f"{fmt(hra_averages.get('light_recovery_pct'), 1)}%", TINTS['light'], width=col_width)
        y_right -= 0.008
        
        fig1.text(right_x + 0.01, y_right, 'Corrected HRV', fontsize=7, fontstyle='italic', color='#71717a')
        y_right -= 0.018
        y_right = draw_row(fig1, right_x, y_right, 'ln(RMSSD₇₀):', fmt(hrv_averages.get('light_hrv_ln_rmssd70'), 3), TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'ln(SDNN₇₀):', fmt(hrv_averages.get('light_hrv_ln_sdnn70'), 3), TINTS['light'], width=col_width)
        y_right = draw_row(fig1, right_x, y_right, 'pNN50₇₀:', f"{fmt(hrv_averages.get('light_hrv_pnn50'), 1)}%", TINTS['light'], width=col_width)
        
        add_page_footer(fig1, 1, total_pages)
        pdf.savefig(fig1)
        plt.close(fig1)
        
        # ==================== PAGE 2: METADATA TABLE ====================
        fig2 = plt.figure(figsize=(8.5, 11))  # Portrait
        fig2.patch.set_facecolor('white')
        
        add_page_header(fig2, 'metadata')
        fig2.text(0.08, 0.90, 'Recording Metadata', ha='left', va='top', fontsize=28, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig2.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig2.transFigure))
        
        fig2.text(0.08, 0.84, 'Table 1 | Recording Information', fontsize=11, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig2.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig2.transFigure))
        
        ax2 = fig2.add_axes([0.08, 0.06, 0.84, 0.76])
        ax2.axis('off')
        
        # Full metadata columns - ALL information, use line breaks
        meta_headers = ['Recording', 'Date', 'hSpO Info', 'hCO Info', 'Fusion', 'Drug Info', 'Light Stim', 'Notes']
        meta_data = []
        
        for rec in recordings:
            # Get recording name with file
            rec_name = rec.get('name', '')
            abf_file = rec.get('abf_filename', '') or rec.get('filename', '') or rec.get('abf_file', '')
            rec_display = f"{rec_name}\n{abf_file}" if abf_file else rec_name
            
            # Parse hSpO Info - from hspo_info dict
            hspo_info_parts = []
            hspo = rec.get('hspo_info') or {}
            if isinstance(hspo, dict):
                if hspo.get('line_name'):
                    hspo_info_parts.append(str(hspo.get('line_name')))
                if hspo.get('passage'):
                    hspo_info_parts.append(f"P{hspo.get('passage')}")
                if hspo.get('age'):
                    hspo_info_parts.append(f"D{hspo.get('age')}")
                if hspo.get('has_transduction'):
                    hspo_info_parts.append('Transduced')
            hspo_info = '\n'.join(hspo_info_parts) if hspo_info_parts else '—'
            
            # Parse hCO Info - from hco_info dict
            hco_info_parts = []
            hco = rec.get('hco_info') or {}
            if isinstance(hco, dict):
                if hco.get('line_name'):
                    hco_info_parts.append(str(hco.get('line_name')))
                if hco.get('passage'):
                    hco_info_parts.append(f"P{hco.get('passage')}")
                if hco.get('age'):
                    hco_info_parts.append(f"D{hco.get('age')}")
            hco_info = '\n'.join(hco_info_parts) if hco_info_parts else '—'
            
            # Fusion date
            fusion = rec.get('fusion_date', '') or '—'
            
            # Drug info - from drug_info list of dicts - show ALL information
            drug_info_raw = rec.get('drug_info', [])
            has_drug = rec.get('has_drug', False)
            drug_hrv_readout = rec.get('drug_hrv_readout_minute')  # Get HRV readout minute (can be 0)
            drug_parts = []
            
            if has_drug and isinstance(drug_info_raw, list) and drug_info_raw:
                for d in drug_info_raw:
                    if isinstance(d, dict):
                        name = d.get('name', '')
                        conc = d.get('concentration', '')
                        unit = d.get('concentration_unit', '') or 'µM'
                        # Use HRV readout minute from comparison data (handles 0 correctly)
                        if drug_hrv_readout is not None:
                            perf_time = drug_hrv_readout
                        else:
                            perf_time = d.get('perfusion_time', '') or d.get('bf_readout_time', '') or d.get('perfusionTime', '')
                        if name:
                            # Format: "DrugName\nConc: 2µM\nPerf. Time: 5min"
                            drug_str = name
                            if conc:
                                drug_str += f"\n{conc}{unit}"
                            if perf_time is not None and perf_time != '':
                                drug_str += f"\nPerf. Time: {perf_time}min"
                            drug_parts.append(drug_str)
            elif isinstance(drug_info_raw, dict):
                name = drug_info_raw.get('name', '')
                conc = drug_info_raw.get('concentration', '')
                unit = drug_info_raw.get('concentration_unit', '') or 'µM'
                if drug_hrv_readout is not None:
                    perf_time = drug_hrv_readout
                else:
                    perf_time = drug_info_raw.get('perfusion_time', '') or drug_info_raw.get('bf_readout_time', '') or drug_info_raw.get('perfusionTime', '')
                if name:
                    drug_str = name
                    if conc:
                        drug_str += f"\n{conc}{unit}"
                    if perf_time is not None and perf_time != '':
                        drug_str += f"\nPerf. Time: {perf_time}min"
                    drug_parts.append(drug_str)
            elif drug_info_raw and str(drug_info_raw).lower() not in ['no', 'none', 'no drug', '—', '-', '']:
                drug_parts.append(str(drug_info_raw))
            
            # Also check for separate drug fields
            if not drug_parts:
                drug_names = rec.get('drug_names', '')
                drug_concs = rec.get('drug_concentrations', '')
                if drug_names and drug_names != '—':
                    drug_str = drug_names
                    if drug_concs and drug_concs != '—':
                        drug_str += f"\n{drug_concs}"
                    drug_parts.append(drug_str)
            
            drug_info = '\n'.join(drug_parts) if drug_parts else 'No drug'
            
            # Light stim info - show all details: stim count, stim_duration and isi_structure
            light_parts = []
            if rec.get('has_light_stim'):
                stim_count = rec.get('light_stim_count', '') or rec.get('stim_count', '')
                stim_dur = rec.get('stim_duration', '')
                isi_struct = rec.get('isi_structure', '')
                
                # Add stim count first
                if stim_count:
                    light_parts.append(f"{stim_count} stim")
                
                # Then duration
                if stim_dur:
                    light_parts.append(f"{stim_dur}s")
                
                # Then ISI
                if isi_struct:
                    light_parts.append(f"ISI: {isi_struct}")
                
                # Also check light_stim_info if present
                light_info_raw = rec.get('light_stim_info', '')
                if isinstance(light_info_raw, dict):
                    if not stim_count and light_info_raw.get('stim_count'):
                        light_parts.insert(0, f"{light_info_raw.get('stim_count')} stim")
                    if not stim_dur and light_info_raw.get('stim_duration'):
                        light_parts.append(f"{light_info_raw.get('stim_duration')}s")
                    if not isi_struct and light_info_raw.get('isi'):
                        light_parts.append(f"ISI: {light_info_raw.get('isi')}")
                elif isinstance(light_info_raw, list):
                    for item in light_info_raw:
                        if item and str(item) not in ' '.join(light_parts):
                            light_parts.append(str(item))
                elif light_info_raw and str(light_info_raw) not in ' '.join(light_parts):
                    light_parts.append(str(light_info_raw))
            light_info = '\n'.join(light_parts) if light_parts else '—'
            
            # Notes
            notes = rec.get('notes', '') or rec.get('recording_description', '') or '—'
            
            meta_data.append([
                rec_display,  # Full name
                rec.get('recording_date', '—') or '—',
                hspo_info,  # Full info
                hco_info,   # Full info
                fusion,
                drug_info,  # Full info
                light_info, # Full info
                notes,      # Full info
            ])
        
        if meta_data:
            table2 = ax2.table(cellText=meta_data, colLabels=meta_headers, loc='upper center', cellLoc='center',
                              colWidths=[0.17, 0.10, 0.13, 0.10, 0.07, 0.11, 0.17, 0.15])
            table2.auto_set_font_size(False)
            
            # Dynamic sizing based on number of recordings
            n_recs = len(recordings)
            font_size = 5 if n_recs <= 10 else 4
            row_scale = 3.5 if n_recs <= 6 else 2.8 if n_recs <= 10 else 2.2 if n_recs <= 15 else 1.8
            table2.set_fontsize(font_size)
            table2.scale(1.0, row_scale)
            
            for (row, col), cell in table2.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor(COLORS['dark'])
                else:
                    cell.set_facecolor('white' if row % 2 == 0 else '#f9fafb')
                    cell.set_text_props(fontfamily=body_font)
        
        add_page_footer(fig2, 2, total_pages)
        pdf.savefig(fig2)
        plt.close(fig2)
        
        # ==================== PAGE 3: SPONTANEOUS ACTIVITY + NORMALIZED ====================
        fig3 = plt.figure(figsize=(8.5, 11))
        fig3.patch.set_facecolor('white')
        
        add_page_header(fig3, 'spontaneous activity')
        fig3.text(0.08, 0.90, 'Spontaneous Activity', ha='left', va='top', fontsize=28, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig3.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig3.transFigure))
        
        # Table 2: Drug-induced BF and HRV Data
        fig3.text(0.08, 0.84, 'Table 2 | Drug-induced BF and HRV Data', fontsize=10, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig3.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig3.transFigure))
        
        # First table area (upper half) - within page borders
        ax3a = fig3.add_axes([0.08, 0.48, 0.84, 0.34])
        ax3a.axis('off')
        
        spont_headers = ['Rec', 'Base BF', 'Base\nRMSSD', 'Base\nSDNN', 'Base\npNN50', 
                        'Drug BF', 'Drug\nRMSSD', 'Drug\nSDNN', 'Drug\npNN50']
        spont_data = []
        
        for rec in recordings:
            spont_data.append([
                extract_short_name(rec.get('name', '')),
                fmt(rec.get('baseline_bf'), 1),
                fmt(rec.get('baseline_ln_rmssd70'), 3),
                fmt(rec.get('baseline_ln_sdnn70'), 3),
                fmt(rec.get('baseline_pnn50'), 1),
                fmt(rec.get('drug_bf'), 1),
                fmt(rec.get('drug_ln_rmssd70'), 3),
                fmt(rec.get('drug_ln_sdnn70'), 3),
                fmt(rec.get('drug_pnn50'), 1),
            ])
        
        # Add average row
        spont_data.append([
            f'Avg',
            fmt(spont_averages.get('baseline_bf'), 1),
            fmt(spont_averages.get('baseline_ln_rmssd70'), 3),
            fmt(spont_averages.get('baseline_ln_sdnn70'), 3),
            fmt(spont_averages.get('baseline_pnn50'), 1),
            fmt(spont_averages.get('drug_bf'), 1),
            fmt(spont_averages.get('drug_ln_rmssd70'), 3),
            fmt(spont_averages.get('drug_ln_sdnn70'), 3),
            fmt(spont_averages.get('drug_pnn50'), 1),
        ])
        
        if spont_data:
            table3a = ax3a.table(cellText=spont_data, colLabels=spont_headers, loc='upper center', cellLoc='center',
                                colWidths=[0.08, 0.115, 0.115, 0.115, 0.115, 0.115, 0.115, 0.115, 0.115])
            table3a.auto_set_font_size(False)
            font_size = 6 if len(recordings) <= 10 else 5
            table3a.set_fontsize(font_size)
            row_scale = 1.6 if len(recordings) <= 8 else 1.3
            table3a.scale(1.0, row_scale)
            
            for (row, col), cell in table3a.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor(COLORS['emerald'])
                elif row == len(spont_data):
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor('#f87171')
                else:
                    if 1 <= col <= 4:
                        cell.set_facecolor(TINTS['baseline'])
                    elif 5 <= col <= 8:
                        cell.set_facecolor(TINTS['drug'])
                    else:
                        cell.set_facecolor('white')
                    cell.set_text_props(fontfamily=body_font)
        
        # Table 3: Drug-induced BF and HRV Normalized Data
        fig3.text(0.08, 0.44, 'Table 3 | Drug-induced BF and HRV Normalized Data', fontsize=10, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig3.add_artist(plt.Line2D([0.08, 0.92], [0.425, 0.425], color=COLORS['line'], linewidth=0.5, transform=fig3.transFigure))
        
        # Second table area (lower half) - top closer to title bar at 0.425
        ax3b = fig3.add_axes([0.08, 0.06, 0.84, 0.36])
        ax3b.axis('off')
        
        # Calculate cohort baseline averages for normalization
        baseline_bfs = [r.get('baseline_bf') for r in recordings if r.get('baseline_bf') is not None]
        baseline_rmssds = [r.get('baseline_ln_rmssd70') for r in recordings if r.get('baseline_ln_rmssd70') is not None]
        baseline_sdnns = [r.get('baseline_ln_sdnn70') for r in recordings if r.get('baseline_ln_sdnn70') is not None]
        baseline_pnn50s = [r.get('baseline_pnn50') for r in recordings if r.get('baseline_pnn50') is not None]
        
        avg_bf = sum(baseline_bfs) / len(baseline_bfs) if baseline_bfs else 1
        avg_rmssd = sum(baseline_rmssds) / len(baseline_rmssds) if baseline_rmssds else 1
        avg_sdnn = sum(baseline_sdnns) / len(baseline_sdnns) if baseline_sdnns else 1
        avg_pnn50 = sum(baseline_pnn50s) / len(baseline_pnn50s) if baseline_pnn50s else 1
        
        norm_headers = ['Rec', 'Base\nBF%', 'Base\nRMSSD%', 'Base\nSDNN%', 'Base\npNN50%',
                       'Drug\nBF%', 'Drug\nRMSSD%', 'Drug\nSDNN%', 'Drug\npNN50%']
        norm_data = []
        norm_sums = {'base_bf': [], 'base_rmssd': [], 'base_sdnn': [], 'base_pnn50': [],
                    'drug_bf': [], 'drug_rmssd': [], 'drug_sdnn': [], 'drug_pnn50': []}
        
        for rec in recordings:
            # Check if this recording has baseline data
            has_baseline = rec.get('baseline_bf') is not None
            
            if has_baseline:
                n_base_bf = norm_val(rec.get('baseline_bf'), avg_bf)
                n_base_rmssd = norm_val(rec.get('baseline_ln_rmssd70'), avg_rmssd)
                n_base_sdnn = norm_val(rec.get('baseline_ln_sdnn70'), avg_sdnn)
                n_base_pnn50 = norm_val(rec.get('baseline_pnn50'), avg_pnn50)
                n_drug_bf = norm_val(rec.get('drug_bf'), avg_bf)
                n_drug_rmssd = norm_val(rec.get('drug_ln_rmssd70'), avg_rmssd)
                n_drug_sdnn = norm_val(rec.get('drug_ln_sdnn70'), avg_sdnn)
                n_drug_pnn50 = norm_val(rec.get('drug_pnn50'), avg_pnn50)
            else:
                # No baseline, show dashes for all normalized values
                n_base_bf = n_base_rmssd = n_base_sdnn = n_base_pnn50 = None
                n_drug_bf = n_drug_rmssd = n_drug_sdnn = n_drug_pnn50 = None
            
            norm_data.append([
                extract_short_name(rec.get('name', '')),
                fmt(n_base_bf, 1), fmt(n_base_rmssd, 1), fmt(n_base_sdnn, 1), fmt(n_base_pnn50, 1),
                fmt(n_drug_bf, 1), fmt(n_drug_rmssd, 1), fmt(n_drug_sdnn, 1), fmt(n_drug_pnn50, 1),
            ])
            
            if has_baseline:
                for key, val in [('base_bf', n_base_bf), ('base_rmssd', n_base_rmssd), 
                                ('base_sdnn', n_base_sdnn), ('base_pnn50', n_base_pnn50),
                                ('drug_bf', n_drug_bf), ('drug_rmssd', n_drug_rmssd),
                                ('drug_sdnn', n_drug_sdnn), ('drug_pnn50', n_drug_pnn50)]:
                    if val is not None:
                        norm_sums[key].append(val)
        
        # Add average row
        norm_data.append([
            'Avg',
            fmt(sum(norm_sums['base_bf']) / len(norm_sums['base_bf']) if norm_sums['base_bf'] else None, 1),
            fmt(sum(norm_sums['base_rmssd']) / len(norm_sums['base_rmssd']) if norm_sums['base_rmssd'] else None, 1),
            fmt(sum(norm_sums['base_sdnn']) / len(norm_sums['base_sdnn']) if norm_sums['base_sdnn'] else None, 1),
            fmt(sum(norm_sums['base_pnn50']) / len(norm_sums['base_pnn50']) if norm_sums['base_pnn50'] else None, 1),
            fmt(sum(norm_sums['drug_bf']) / len(norm_sums['drug_bf']) if norm_sums['drug_bf'] else None, 1),
            fmt(sum(norm_sums['drug_rmssd']) / len(norm_sums['drug_rmssd']) if norm_sums['drug_rmssd'] else None, 1),
            fmt(sum(norm_sums['drug_sdnn']) / len(norm_sums['drug_sdnn']) if norm_sums['drug_sdnn'] else None, 1),
            fmt(sum(norm_sums['drug_pnn50']) / len(norm_sums['drug_pnn50']) if norm_sums['drug_pnn50'] else None, 1),
        ])
        
        if norm_data:
            table3b = ax3b.table(cellText=norm_data, colLabels=norm_headers, loc='upper center', cellLoc='center',
                                colWidths=[0.08, 0.115, 0.115, 0.115, 0.115, 0.115, 0.115, 0.115, 0.115])
            table3b.auto_set_font_size(False)
            table3b.set_fontsize(font_size)
            table3b.scale(1.0, row_scale)
            
            for (row, col), cell in table3b.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor(COLORS['emerald'])
                elif row == len(norm_data):
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor('#f87171')
                else:
                    if 1 <= col <= 4:
                        cell.set_facecolor(TINTS['baseline'])
                    elif 5 <= col <= 8:
                        cell.set_facecolor(TINTS['drug'])
                    else:
                        cell.set_facecolor('white')
                    cell.set_text_props(fontfamily=body_font)
        
        add_page_footer(fig3, 3, total_pages)
        pdf.savefig(fig3)
        plt.close(fig3)
        
        # ==================== PAGE 4: LIGHT HRA + NORMALIZED ====================
        fig4 = plt.figure(figsize=(8.5, 11))
        fig4.patch.set_facecolor('white')
        
        add_page_header(fig4, 'light stimulus')
        fig4.text(0.08, 0.90, 'Heart Rate Adaptation', ha='left', va='top', fontsize=28, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig4.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig4.transFigure))
        
        # Table 4: Light-Induced HRA Data
        fig4.text(0.08, 0.84, 'Table 4 | Light-Induced HRA Data', fontsize=10, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig4.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig4.transFigure))
        
        # First table area - within page borders, increased width by 0.05cm each side
        ax4a = fig4.add_axes([0.065, 0.48, 0.87, 0.34])
        ax4a.axis('off')
        
        # Reordered columns with 1st TTP
        hra_headers = ['Stim', 'Baseline\nBF', 'Avg\nBF', 'Peak\nBF', 'Peak\n%', '1st TTP\n(s)', 'TTP\n(s)', 'BF\nRec', 'Rec\n%', 'Amp.\nBF', 'RoC\n(1/min)']
        hra_data = []
        
        def fmt_ttp(val):
            """Format TTP - show 0.0 even if value is 0 or None"""
            if val is None:
                return '0.0'
            return f"{val:.1f}"
        
        for rec in recordings:
            hra_data.append([
                extract_short_name(rec.get('name', '')),
                fmt(rec.get('light_baseline_bf'), 1),
                fmt(rec.get('light_avg_bf'), 1),
                fmt(rec.get('light_peak_bf'), 1),
                fmt(rec.get('light_peak_norm'), 1),
                fmt_ttp(rec.get('light_ttp_first')),       # 1st TTP (s)
                fmt(rec.get('light_ttp_avg'), 1),          # TTP (s)
                fmt(rec.get('light_recovery_bf'), 1),      # BF Rec
                fmt(rec.get('light_recovery_pct'), 1),     # Rec %
                fmt(rec.get('light_amplitude'), 1),        # Amp. BF
                fmt(rec.get('light_roc'), 4),              # RoC (1/min)
            ])
        
        hra_data.append([
            'Avg',
            fmt(hra_averages.get('light_baseline_bf'), 1),
            fmt(hra_averages.get('light_avg_bf'), 1),
            fmt(hra_averages.get('light_peak_bf'), 1),
            fmt(hra_averages.get('light_peak_norm'), 1),
            fmt_ttp(hra_averages.get('light_ttp_first')),
            fmt(hra_averages.get('light_ttp_avg'), 1),
            fmt(hra_averages.get('light_recovery_bf'), 1),
            fmt(hra_averages.get('light_recovery_pct'), 1),
            fmt(hra_averages.get('light_amplitude'), 1),
            fmt(hra_averages.get('light_roc'), 4),
        ])
        
        if hra_data:
            table4a = ax4a.table(cellText=hra_data, colLabels=hra_headers, loc='upper center', cellLoc='center',
                                colWidths=[0.07, 0.09, 0.09, 0.09, 0.09, 0.09, 0.09, 0.09, 0.09, 0.09, 0.09])
            table4a.auto_set_font_size(False)
            table4a.set_fontsize(5)
            table4a.scale(1.0, row_scale)
            
            for (row, col), cell in table4a.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor(COLORS['amber'])
                elif row == len(hra_data):
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor('#f87171')
                else:
                    if col >= 1:
                        cell.set_facecolor(TINTS['light'])
                    else:
                        cell.set_facecolor('white')
                    cell.set_text_props(fontfamily=body_font)
        
        # Table 5: Light-Induced HRA Normalized Data
        fig4.text(0.08, 0.44, 'Table 5 | Light-Induced HRA Normalized Data', fontsize=10, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig4.add_artist(plt.Line2D([0.08, 0.92], [0.425, 0.425], color=COLORS['line'], linewidth=0.5, transform=fig4.transFigure))
        
        # Second table area - top closer to title bar at 0.425
        ax4b = fig4.add_axes([0.08, 0.06, 0.84, 0.36])
        ax4b.axis('off')
        
        light_baseline_bfs = [r.get('light_baseline_bf') for r in recordings if r.get('light_baseline_bf') is not None]
        avg_light_bf = sum(light_baseline_bfs) / len(light_baseline_bfs) if light_baseline_bfs else 1
        
        norm_hra_headers = ['Rec', 'Baseline BF%', 'Avg BF%', 'Peak BF%', 'Recovery BF%']
        norm_hra_data = []
        norm_hra_sums = {'base': [], 'avg': [], 'peak': [], 'rec': []}
        
        for rec in recordings:
            n_base = norm_val(rec.get('light_baseline_bf'), avg_light_bf)
            n_avg = norm_val(rec.get('light_avg_bf'), avg_light_bf)
            n_peak = norm_val(rec.get('light_peak_bf'), avg_light_bf)
            n_rec = norm_val(rec.get('light_recovery_bf'), avg_light_bf)
            
            norm_hra_data.append([
                extract_short_name(rec.get('name', '')),
                fmt(n_base, 1), fmt(n_avg, 1), fmt(n_peak, 1), fmt(n_rec, 1),
            ])
            
            for key, val in [('base', n_base), ('avg', n_avg), ('peak', n_peak), ('rec', n_rec)]:
                if val is not None:
                    norm_hra_sums[key].append(val)
        
        norm_hra_data.append([
            'Avg',
            fmt(sum(norm_hra_sums['base']) / len(norm_hra_sums['base']) if norm_hra_sums['base'] else None, 1),
            fmt(sum(norm_hra_sums['avg']) / len(norm_hra_sums['avg']) if norm_hra_sums['avg'] else None, 1),
            fmt(sum(norm_hra_sums['peak']) / len(norm_hra_sums['peak']) if norm_hra_sums['peak'] else None, 1),
            fmt(sum(norm_hra_sums['rec']) / len(norm_hra_sums['rec']) if norm_hra_sums['rec'] else None, 1),
        ])
        
        if norm_hra_data:
            table4b = ax4b.table(cellText=norm_hra_data, colLabels=norm_hra_headers, loc='upper center', cellLoc='center',
                                colWidths=[0.16, 0.21, 0.21, 0.21, 0.21])
            table4b.auto_set_font_size(False)
            table4b.set_fontsize(7)
            table4b.scale(1.0, row_scale)
            
            for (row, col), cell in table4b.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor(COLORS['amber'])
                elif row == len(norm_hra_data):
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor('#f87171')
                else:
                    if col >= 1:
                        cell.set_facecolor(TINTS['light'])
                    else:
                        cell.set_facecolor('white')
                    cell.set_text_props(fontfamily=body_font)
        
        add_page_footer(fig4, 4, total_pages)
        pdf.savefig(fig4)
        plt.close(fig4)
        
        # ==================== PAGE 5: DETRENDED HRV ====================
        fig5 = plt.figure(figsize=(8.5, 11))
        fig5.patch.set_facecolor('white')
        
        add_page_header(fig5, 'light stimulus')
        fig5.text(0.08, 0.90, 'Detrended HRV', ha='left', va='top', fontsize=28, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig5.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig5.transFigure))
        
        fig5.text(0.08, 0.84, 'Table 6 | Light-Induced Detrended HRV Data', fontsize=11, fontweight='bold', 
                 color=COLORS['dark'], fontfamily=title_font)
        fig5.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig5.transFigure))
        
        ax5 = fig5.add_axes([0.08, 0.05, 0.84, 0.77])
        ax5.axis('off')
        
        # Only ln(RMSSD), ln(SDNN), pNN50 - remove raw RMSSD and SDNN
        hrv_headers = ['Recording', 'ln(RMSSD₇₀)', 'ln(SDNN₇₀)', 'pNN50₇₀']
        hrv_data = []
        
        for rec in recordings:
            hrv_data.append([
                extract_short_name(rec.get('name', '')),
                fmt(rec.get('light_hrv_ln_rmssd70'), 3),
                fmt(rec.get('light_hrv_ln_sdnn70'), 3),
                fmt(rec.get('light_hrv_pnn50'), 1),
            ])
        
        hrv_data.append([
            'Median',
            fmt(hrv_averages.get('light_hrv_ln_rmssd70'), 3),
            fmt(hrv_averages.get('light_hrv_ln_sdnn70'), 3),
            fmt(hrv_averages.get('light_hrv_pnn50'), 1),
        ])
        
        if hrv_data:
            table5 = ax5.table(cellText=hrv_data, colLabels=hrv_headers, loc='upper center', cellLoc='center',
                              colWidths=[0.25, 0.25, 0.25, 0.25])
            table5.auto_set_font_size(False)
            font_size = 8 if len(recordings) <= 15 else 7
            table5.set_fontsize(font_size)
            row_scale = 1.8 if len(recordings) <= 15 else 1.5 if len(recordings) <= 20 else 1.3
            table5.scale(1.0, row_scale)
            
            for (row, col), cell in table5.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor(COLORS['amber'])
                elif row == len(hrv_data):
                    cell.set_text_props(fontweight='bold', color='white', fontfamily=body_font)
                    cell.set_facecolor('#f87171')
                else:
                    if col >= 1:
                        cell.set_facecolor(TINTS['light'])
                    else:
                        cell.set_facecolor('white')
                    cell.set_text_props(fontfamily=body_font)
        
        add_page_footer(fig5, 5, total_pages)
        pdf.savefig(fig5)
        plt.close(fig5)
    
    buf.seek(0)
    return buf



def create_comparison_xlsx(folder_name, comparison_data):
    """Create comparison Excel export matching the PDF structure with multiple sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import re
    
    data = comparison_data
    recordings = data.get('recordings', [])
    summary = data.get('summary', {})
    spont_averages = data.get('spontaneous_averages', {}).get('averages', {})
    hra_averages = data.get('light_hra_averages', {}).get('averages', {})
    hrv_averages = data.get('light_hrv_averages', {}).get('averages', {})
    
    # Sort recordings alphabetically
    recordings = sorted(recordings, key=lambda x: x.get('name', '').lower())
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    data_font = Font(size=9)
    bold_data_font = Font(bold=True, size=9)
    avg_font = Font(bold=True, color='FFFFFF', size=9)
    
    dark_fill = PatternFill(start_color='18181B', end_color='18181B', fill_type='solid')
    emerald_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    amber_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    avg_fill = PatternFill(start_color='F87171', end_color='F87171', fill_type='solid')
    baseline_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
    drug_fill = PatternFill(start_color='F3E8FF', end_color='F3E8FF', fill_type='solid')
    light_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def fmt(val, dec=2):
        if val is None:
            return '—'
        try:
            if dec == 0:
                return f"{float(val):.0f}"
            elif dec == 1:
                return f"{float(val):.1f}"
            elif dec == 3:
                return f"{float(val):.3f}"
            elif dec == 4:
                return f"{float(val):.4f}"
            return f"{float(val):.{dec}f}"
        except:
            return '—'
    
    def fmt_age_range(age_dict):
        if not age_dict or age_dict.get('min') is None:
            return '—'
        return f"{age_dict.get('min')} - {age_dict.get('max')} days"
    
    def extract_short_name(full_name):
        if not full_name:
            return '—'
        match = re.search(r'[FC]\d+$', full_name)
        if match:
            return match.group()
        parts = full_name.replace('/', '-').split('-')
        return parts[-1] if parts else full_name[:10]
    
    def norm_val(val, avg):
        if val is None or avg == 0:
            return None
        return 100 * val / avg
    
    def parse_drug_info(drug_info):
        if not drug_info:
            return None
        if isinstance(drug_info, dict):
            name = drug_info.get('name', '')
            conc = drug_info.get('concentration', '')
            if name:
                return f"{name} ({conc})" if conc else name
            return None
        if isinstance(drug_info, list):
            results = []
            for d in drug_info:
                parsed = parse_drug_info(d)
                if parsed:
                    results.append(parsed)
            return results if results else None
        drug_str = str(drug_info).strip()
        if drug_str.lower() in ['no drug', 'no', 'none', '—', '-', '']:
            return None
        return drug_str
    
    # ==================== SHEET 1: SUMMARY ====================
    ws1 = wb.active
    ws1.title = "Summary"
    
    # Title
    ws1.merge_cells('A1:D1')
    ws1['A1'] = folder_name
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = left_align
    
    row = 3
    # Folder Overview
    ws1[f'A{row}'] = 'FOLDER OVERVIEW'
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = dark_fill
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1
    ws1[f'A{row}'] = 'Recordings:'
    ws1[f'B{row}'] = summary.get('recording_count', len(recordings))
    ws1[f'B{row}'].font = bold_data_font
    row += 1
    ws1[f'A{row}'] = 'Date Created:'
    ws1[f'B{row}'] = datetime.now().strftime('%Y-%m-%d')
    ws1[f'B{row}'].font = bold_data_font
    row += 2
    
    # Age Ranges
    ws1[f'A{row}'] = 'AGE RANGES'
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = PatternFill(start_color='6B7280', end_color='6B7280', fill_type='solid')
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1
    hspo_range = summary.get('hspo_age_range', {})
    hco_range = summary.get('hco_age_range', {})
    fusion_range = summary.get('fusion_age_range', {})
    ws1[f'A{row}'] = 'hSpOs:'
    ws1[f'B{row}'] = f"{fmt_age_range(hspo_range)} (n={hspo_range.get('n', 0)})"
    row += 1
    ws1[f'A{row}'] = 'hCOs:'
    ws1[f'B{row}'] = f"{fmt_age_range(hco_range)} (n={hco_range.get('n', 0)})"
    row += 1
    ws1[f'A{row}'] = 'Fusion:'
    ws1[f'B{row}'] = f"{fmt_age_range(fusion_range)} (n={fusion_range.get('n', 0)})"
    row += 2
    
    # Parameters
    ws1[f'A{row}'] = 'PARAMETERS'
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = PatternFill(start_color='6B7280', end_color='6B7280', fill_type='solid')
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Drug Used
    all_drugs = []
    drug_concentration_unit = ''
    for r in recordings:
        drug_info_raw = r.get('drug_info')
        if isinstance(drug_info_raw, dict):
            conc_unit = drug_info_raw.get('concentration_unit', '') or drug_info_raw.get('unit', '')
            if conc_unit and not drug_concentration_unit:
                drug_concentration_unit = conc_unit
        parsed = parse_drug_info(drug_info_raw)
        if parsed:
            if isinstance(parsed, list):
                all_drugs.extend(parsed)
            else:
                all_drugs.append(parsed)
    
    if all_drugs and not drug_concentration_unit:
        drug_concentration_unit = 'µM'
    
    if all_drugs:
        unique_drugs = list(dict.fromkeys(all_drugs))
        drug_text = unique_drugs[0]
        if drug_concentration_unit:
            match = re.search(r'\(([^)]+)\)$', drug_text)
            if match:
                conc_value = match.group(1)
                drug_text = re.sub(r'\([^)]+\)$', f'({conc_value}{drug_concentration_unit})', drug_text)
    else:
        drug_text = '—'
    
    ws1[f'A{row}'] = 'Drug Used:'
    ws1[f'B{row}'] = drug_text
    row += 1
    
    # Light Stim
    light_used = any(r.get('has_light_stim') or r.get('stim_duration') for r in recordings)
    ws1[f'A{row}'] = 'Light Stim:'
    ws1[f'B{row}'] = 'Yes' if light_used else 'No'
    row += 2
    
    # Right side - Spontaneous Activity Averages
    ws1[f'D3'] = 'SPONTANEOUS ACTIVITY'
    ws1[f'D3'].font = header_font
    ws1[f'D3'].fill = emerald_fill
    ws1.merge_cells('D3:E3')
    
    ws1['D4'] = 'Baseline Readout'
    ws1['D4'].font = Font(italic=True, size=8, color='71717A')
    ws1['D5'] = 'Mean BF:'
    ws1['E5'] = f"{fmt(spont_averages.get('baseline_bf'), 1)} bpm"
    ws1['E5'].fill = baseline_fill
    ws1['D6'] = 'ln(RMSSD₇₀):'
    ws1['E6'] = fmt(spont_averages.get('baseline_ln_rmssd70'), 3)
    ws1['E6'].fill = baseline_fill
    ws1['D7'] = 'ln(SDNN₇₀):'
    ws1['E7'] = fmt(spont_averages.get('baseline_ln_sdnn70'), 3)
    ws1['E7'].fill = baseline_fill
    ws1['D8'] = 'pNN50₇₀:'
    ws1['E8'] = f"{fmt(spont_averages.get('baseline_pnn50'), 1)}%"
    ws1['E8'].fill = baseline_fill
    
    ws1['D10'] = 'Drug Readout'
    ws1['D10'].font = Font(italic=True, size=8, color='71717A')
    ws1['D11'] = 'Mean BF:'
    ws1['E11'] = f"{fmt(spont_averages.get('drug_bf'), 1)} bpm"
    ws1['E11'].fill = drug_fill
    ws1['D12'] = 'ln(RMSSD₇₀):'
    ws1['E12'] = fmt(spont_averages.get('drug_ln_rmssd70'), 3)
    ws1['E12'].fill = drug_fill
    ws1['D13'] = 'ln(SDNN₇₀):'
    ws1['E13'] = fmt(spont_averages.get('drug_ln_sdnn70'), 3)
    ws1['E13'].fill = drug_fill
    ws1['D14'] = 'pNN50₇₀:'
    ws1['E14'] = f"{fmt(spont_averages.get('drug_pnn50'), 1)}%"
    ws1['E14'].fill = drug_fill
    
    # Light Stimulus Averages
    ws1[f'D16'] = 'LIGHT STIMULUS'
    ws1[f'D16'].font = header_font
    ws1[f'D16'].fill = amber_fill
    ws1.merge_cells('D16:E16')
    
    ws1['D17'] = 'Heart Rate Adaptation (HRA)'
    ws1['D17'].font = Font(italic=True, size=8, color='71717A')
    ws1['D18'] = 'Baseline BF:'
    ws1['E18'] = f"{fmt(hra_averages.get('light_baseline_bf'), 1)} bpm"
    ws1['E18'].fill = light_fill
    ws1['D19'] = 'Peak BF:'
    ws1['E19'] = f"{fmt(hra_averages.get('light_peak_bf'), 1)} bpm"
    ws1['E19'].fill = light_fill
    ws1['D20'] = 'Peak (Norm.):'
    ws1['E20'] = f"{fmt(hra_averages.get('light_peak_norm'), 1)}%"
    ws1['E20'].fill = light_fill
    ws1['D21'] = 'Amplitude:'
    ws1['E21'] = f"{fmt(hra_averages.get('light_amplitude'), 1)} bpm"
    ws1['E21'].fill = light_fill
    ws1['D22'] = 'TTP (Avg):'
    ws1['E22'] = f"{fmt(hra_averages.get('light_ttp_avg'), 1)} s"
    ws1['E22'].fill = light_fill
    ws1['D23'] = 'Rate of Change:'
    ws1['E23'] = fmt(hra_averages.get('light_roc'), 4)
    ws1['E23'].fill = light_fill
    ws1['D24'] = 'Recovery %:'
    ws1['E24'] = f"{fmt(hra_averages.get('light_recovery_pct'), 1)}%"
    ws1['E24'].fill = light_fill
    
    ws1['D26'] = 'Corrected HRV'
    ws1['D26'].font = Font(italic=True, size=8, color='71717A')
    ws1['D27'] = 'ln(RMSSD₇₀):'
    ws1['E27'] = fmt(hrv_averages.get('light_hrv_ln_rmssd70'), 3)
    ws1['E27'].fill = light_fill
    ws1['D28'] = 'ln(SDNN₇₀):'
    ws1['E28'] = fmt(hrv_averages.get('light_hrv_ln_sdnn70'), 3)
    ws1['E28'].fill = light_fill
    ws1['D29'] = 'pNN50₇₀:'
    ws1['E29'] = f"{fmt(hrv_averages.get('light_hrv_pnn50'), 1)}%"
    ws1['E29'].fill = light_fill
    
    # Set column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 5
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 15
    
    # ==================== SHEET 2: METADATA ====================
    ws2 = wb.create_sheet("Metadata")
    
    # Table 1: Recording Information
    ws2['A1'] = 'Table 1 | Recording Information'
    ws2['A1'].font = Font(bold=True, size=12)
    
    meta_headers = ['Recording', 'Date', 'hSpO Info', 'hCO Info', 'Fusion', 'Drug Info', 'Light Stim', 'Notes']
    for col, header in enumerate(meta_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row = 4
    for idx, rec in enumerate(recordings):
        # Recording name
        rec_name = rec.get('name', '')
        abf_file = rec.get('abf_filename', '') or rec.get('filename', '')
        rec_display = f"{rec_name}\n{abf_file}" if abf_file else rec_name
        
        # hSpO Info
        hspo_info_parts = []
        hspo = rec.get('hspo_info') or {}
        if isinstance(hspo, dict):
            if hspo.get('line_name'):
                hspo_info_parts.append(str(hspo.get('line_name')))
            if hspo.get('passage'):
                hspo_info_parts.append(f"P{hspo.get('passage')}")
            if hspo.get('age'):
                hspo_info_parts.append(f"D{hspo.get('age')}")
        hspo_info = '\n'.join(hspo_info_parts) if hspo_info_parts else '—'
        
        # hCO Info
        hco_info_parts = []
        hco = rec.get('hco_info') or {}
        if isinstance(hco, dict):
            if hco.get('line_name'):
                hco_info_parts.append(str(hco.get('line_name')))
            if hco.get('passage'):
                hco_info_parts.append(f"P{hco.get('passage')}")
            if hco.get('age'):
                hco_info_parts.append(f"D{hco.get('age')}")
        hco_info = '\n'.join(hco_info_parts) if hco_info_parts else '—'
        
        # Drug info
        drug_info_raw = rec.get('drug_info', [])
        has_drug = rec.get('has_drug', False)
        drug_hrv_readout = rec.get('drug_hrv_readout_minute')
        drug_parts = []
        
        if has_drug and isinstance(drug_info_raw, list) and drug_info_raw:
            for d in drug_info_raw:
                if isinstance(d, dict):
                    name = d.get('name', '')
                    conc = d.get('concentration', '')
                    unit = d.get('concentration_unit', '') or 'µM'
                    if drug_hrv_readout is not None:
                        perf_time = drug_hrv_readout
                    else:
                        perf_time = d.get('perfusion_time', '') or d.get('bf_readout_time', '')
                    if name:
                        drug_str = name
                        if conc:
                            drug_str += f"\n{conc}{unit}"
                        if perf_time is not None and perf_time != '':
                            drug_str += f"\nPerf. Time: {perf_time}min"
                        drug_parts.append(drug_str)
        drug_info = '\n'.join(drug_parts) if drug_parts else 'No drug'
        
        # Light stim info
        light_parts = []
        if rec.get('has_light_stim'):
            stim_count = rec.get('light_stim_count', '')
            stim_dur = rec.get('stim_duration', '')
            isi_struct = rec.get('isi_structure', '')
            if stim_count:
                light_parts.append(f"{stim_count} stim")
            if stim_dur:
                light_parts.append(f"{stim_dur}s")
            if isi_struct:
                light_parts.append(f"ISI: {isi_struct}")
        light_info = '\n'.join(light_parts) if light_parts else '—'
        
        # Notes
        notes = rec.get('notes', '') or '—'
        
        data_row = [rec_display, rec.get('recording_date', '—'), hspo_info, hco_info,
                    rec.get('fusion_date', '—'), drug_info, light_info, notes]
        
        for col, value in enumerate(data_row, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = alt_fill
        row += 1
    
    # Set column widths
    for col, width in enumerate([25, 12, 15, 12, 10, 18, 20, 20], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    
    # ==================== SHEET 3: SPONTANEOUS ACTIVITY ====================
    ws3 = wb.create_sheet("Spontaneous Activity")
    
    # Table 2: Drug-induced BF and HRV Data
    ws3['A1'] = 'Table 2 | Drug-induced BF and HRV Data'
    ws3['A1'].font = Font(bold=True, size=12)
    
    spont_headers = ['Rec', 'Base BF', 'Base RMSSD', 'Base SDNN', 'Base pNN50', 
                    'Drug BF', 'Drug RMSSD', 'Drug SDNN', 'Drug pNN50']
    for col, header in enumerate(spont_headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = emerald_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row = 4
    for rec in recordings:
        data_row = [
            extract_short_name(rec.get('name', '')),
            fmt(rec.get('baseline_bf'), 1),
            fmt(rec.get('baseline_ln_rmssd70'), 3),
            fmt(rec.get('baseline_ln_sdnn70'), 3),
            fmt(rec.get('baseline_pnn50'), 1),
            fmt(rec.get('drug_bf'), 1),
            fmt(rec.get('drug_ln_rmssd70'), 3),
            fmt(rec.get('drug_ln_sdnn70'), 3),
            fmt(rec.get('drug_pnn50'), 1),
        ]
        for col, value in enumerate(data_row, 1):
            cell = ws3.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if 2 <= col <= 5:
                cell.fill = baseline_fill
            elif 6 <= col <= 9:
                cell.fill = drug_fill
        row += 1
    
    # Average row
    avg_row = [
        'Avg',
        fmt(spont_averages.get('baseline_bf'), 1),
        fmt(spont_averages.get('baseline_ln_rmssd70'), 3),
        fmt(spont_averages.get('baseline_ln_sdnn70'), 3),
        fmt(spont_averages.get('baseline_pnn50'), 1),
        fmt(spont_averages.get('drug_bf'), 1),
        fmt(spont_averages.get('drug_ln_rmssd70'), 3),
        fmt(spont_averages.get('drug_ln_sdnn70'), 3),
        fmt(spont_averages.get('drug_pnn50'), 1),
    ]
    for col, value in enumerate(avg_row, 1):
        cell = ws3.cell(row=row, column=col, value=value)
        cell.font = avg_font
        cell.fill = avg_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 2
    
    # Table 3: Drug-induced BF and HRV Normalized Data
    ws3[f'A{row}'] = 'Table 3 | Drug-induced BF and HRV Normalized Data'
    ws3[f'A{row}'].font = Font(bold=True, size=12)
    row += 2
    
    norm_headers = ['Rec', 'Base BF%', 'Base RMSSD%', 'Base SDNN%', 'Base pNN50%',
                   'Drug BF%', 'Drug RMSSD%', 'Drug SDNN%', 'Drug pNN50%']
    for col, header in enumerate(norm_headers, 1):
        cell = ws3.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = emerald_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    # Calculate cohort averages for normalization
    baseline_bfs = [r.get('baseline_bf') for r in recordings if r.get('baseline_bf') is not None]
    baseline_rmssds = [r.get('baseline_ln_rmssd70') for r in recordings if r.get('baseline_ln_rmssd70') is not None]
    baseline_sdnns = [r.get('baseline_ln_sdnn70') for r in recordings if r.get('baseline_ln_sdnn70') is not None]
    baseline_pnn50s = [r.get('baseline_pnn50') for r in recordings if r.get('baseline_pnn50') is not None]
    
    avg_bf = sum(baseline_bfs) / len(baseline_bfs) if baseline_bfs else 1
    avg_rmssd = sum(baseline_rmssds) / len(baseline_rmssds) if baseline_rmssds else 1
    avg_sdnn = sum(baseline_sdnns) / len(baseline_sdnns) if baseline_sdnns else 1
    avg_pnn50 = sum(baseline_pnn50s) / len(baseline_pnn50s) if baseline_pnn50s else 1
    
    norm_sums = {'base_bf': [], 'base_rmssd': [], 'base_sdnn': [], 'base_pnn50': [],
                'drug_bf': [], 'drug_rmssd': [], 'drug_sdnn': [], 'drug_pnn50': []}
    
    for rec in recordings:
        has_baseline = rec.get('baseline_bf') is not None
        if has_baseline:
            n_base_bf = norm_val(rec.get('baseline_bf'), avg_bf)
            n_base_rmssd = norm_val(rec.get('baseline_ln_rmssd70'), avg_rmssd)
            n_base_sdnn = norm_val(rec.get('baseline_ln_sdnn70'), avg_sdnn)
            n_base_pnn50 = norm_val(rec.get('baseline_pnn50'), avg_pnn50)
            n_drug_bf = norm_val(rec.get('drug_bf'), avg_bf)
            n_drug_rmssd = norm_val(rec.get('drug_ln_rmssd70'), avg_rmssd)
            n_drug_sdnn = norm_val(rec.get('drug_ln_sdnn70'), avg_sdnn)
            n_drug_pnn50 = norm_val(rec.get('drug_pnn50'), avg_pnn50)
        else:
            n_base_bf = n_base_rmssd = n_base_sdnn = n_base_pnn50 = None
            n_drug_bf = n_drug_rmssd = n_drug_sdnn = n_drug_pnn50 = None
        
        data_row = [
            extract_short_name(rec.get('name', '')),
            fmt(n_base_bf, 1), fmt(n_base_rmssd, 1), fmt(n_base_sdnn, 1), fmt(n_base_pnn50, 1),
            fmt(n_drug_bf, 1), fmt(n_drug_rmssd, 1), fmt(n_drug_sdnn, 1), fmt(n_drug_pnn50, 1),
        ]
        for col, value in enumerate(data_row, 1):
            cell = ws3.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if 2 <= col <= 5:
                cell.fill = baseline_fill
            elif 6 <= col <= 9:
                cell.fill = drug_fill
        
        if has_baseline:
            for key, val in [('base_bf', n_base_bf), ('base_rmssd', n_base_rmssd), 
                            ('base_sdnn', n_base_sdnn), ('base_pnn50', n_base_pnn50),
                            ('drug_bf', n_drug_bf), ('drug_rmssd', n_drug_rmssd),
                            ('drug_sdnn', n_drug_sdnn), ('drug_pnn50', n_drug_pnn50)]:
                if val is not None:
                    norm_sums[key].append(val)
        row += 1
    
    # Normalized average row
    norm_avg_row = [
        'Avg',
        fmt(sum(norm_sums['base_bf']) / len(norm_sums['base_bf']) if norm_sums['base_bf'] else None, 1),
        fmt(sum(norm_sums['base_rmssd']) / len(norm_sums['base_rmssd']) if norm_sums['base_rmssd'] else None, 1),
        fmt(sum(norm_sums['base_sdnn']) / len(norm_sums['base_sdnn']) if norm_sums['base_sdnn'] else None, 1),
        fmt(sum(norm_sums['base_pnn50']) / len(norm_sums['base_pnn50']) if norm_sums['base_pnn50'] else None, 1),
        fmt(sum(norm_sums['drug_bf']) / len(norm_sums['drug_bf']) if norm_sums['drug_bf'] else None, 1),
        fmt(sum(norm_sums['drug_rmssd']) / len(norm_sums['drug_rmssd']) if norm_sums['drug_rmssd'] else None, 1),
        fmt(sum(norm_sums['drug_sdnn']) / len(norm_sums['drug_sdnn']) if norm_sums['drug_sdnn'] else None, 1),
        fmt(sum(norm_sums['drug_pnn50']) / len(norm_sums['drug_pnn50']) if norm_sums['drug_pnn50'] else None, 1),
    ]
    for col, value in enumerate(norm_avg_row, 1):
        cell = ws3.cell(row=row, column=col, value=value)
        cell.font = avg_font
        cell.fill = avg_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Set column widths
    for col in range(1, 10):
        ws3.column_dimensions[get_column_letter(col)].width = 12
    
    # ==================== SHEET 4: HEART RATE ADAPTATION ====================
    ws4 = wb.create_sheet("Heart Rate Adaptation")
    
    # Table 4: Light-Induced HRA Data
    ws4['A1'] = 'Table 4 | Light-Induced HRA Data'
    ws4['A1'].font = Font(bold=True, size=12)
    
    hra_headers = ['Stim', 'Baseline BF', 'Avg BF', 'Peak BF', 'Peak %', '1st TTP (s)', 'TTP (s)', 'BF Rec', 'Rec %', 'Amp. BF', 'RoC (1/min)']
    for col, header in enumerate(hra_headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = amber_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    def fmt_ttp(val):
        if val is None:
            return '0.0'
        return f"{val:.1f}"
    
    row = 4
    for rec in recordings:
        data_row = [
            extract_short_name(rec.get('name', '')),
            fmt(rec.get('light_baseline_bf'), 1),
            fmt(rec.get('light_avg_bf'), 1),
            fmt(rec.get('light_peak_bf'), 1),
            fmt(rec.get('light_peak_norm'), 1),
            fmt_ttp(rec.get('light_ttp_first')),
            fmt(rec.get('light_ttp_avg'), 1),
            fmt(rec.get('light_recovery_bf'), 1),
            fmt(rec.get('light_recovery_pct'), 1),
            fmt(rec.get('light_amplitude'), 1),
            fmt(rec.get('light_roc'), 4),
        ]
        for col, value in enumerate(data_row, 1):
            cell = ws4.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if col >= 2:
                cell.fill = light_fill
        row += 1
    
    # Average row
    hra_avg_row = [
        'Avg',
        fmt(hra_averages.get('light_baseline_bf'), 1),
        fmt(hra_averages.get('light_avg_bf'), 1),
        fmt(hra_averages.get('light_peak_bf'), 1),
        fmt(hra_averages.get('light_peak_norm'), 1),
        fmt_ttp(hra_averages.get('light_ttp_first')),
        fmt(hra_averages.get('light_ttp_avg'), 1),
        fmt(hra_averages.get('light_recovery_bf'), 1),
        fmt(hra_averages.get('light_recovery_pct'), 1),
        fmt(hra_averages.get('light_amplitude'), 1),
        fmt(hra_averages.get('light_roc'), 4),
    ]
    for col, value in enumerate(hra_avg_row, 1):
        cell = ws4.cell(row=row, column=col, value=value)
        cell.font = avg_font
        cell.fill = avg_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 2
    
    # Table 5: Light-Induced HRA Normalized Data
    ws4[f'A{row}'] = 'Table 5 | Light-Induced HRA Normalized Data'
    ws4[f'A{row}'].font = Font(bold=True, size=12)
    row += 2
    
    norm_hra_headers = ['Rec', 'Baseline BF%', 'Avg BF%', 'Peak BF%', 'Recovery BF%']
    for col, header in enumerate(norm_hra_headers, 1):
        cell = ws4.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = amber_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    light_baseline_bfs = [r.get('light_baseline_bf') for r in recordings if r.get('light_baseline_bf') is not None]
    avg_light_bf = sum(light_baseline_bfs) / len(light_baseline_bfs) if light_baseline_bfs else 1
    
    norm_hra_sums = {'base': [], 'avg': [], 'peak': [], 'rec': []}
    
    for rec in recordings:
        n_base = norm_val(rec.get('light_baseline_bf'), avg_light_bf)
        n_avg = norm_val(rec.get('light_avg_bf'), avg_light_bf)
        n_peak = norm_val(rec.get('light_peak_bf'), avg_light_bf)
        n_rec = norm_val(rec.get('light_recovery_bf'), avg_light_bf)
        
        data_row = [
            extract_short_name(rec.get('name', '')),
            fmt(n_base, 1), fmt(n_avg, 1), fmt(n_peak, 1), fmt(n_rec, 1),
        ]
        for col, value in enumerate(data_row, 1):
            cell = ws4.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if col >= 2:
                cell.fill = light_fill
        
        for key, val in [('base', n_base), ('avg', n_avg), ('peak', n_peak), ('rec', n_rec)]:
            if val is not None:
                norm_hra_sums[key].append(val)
        row += 1
    
    # Normalized HRA average row
    norm_hra_avg_row = [
        'Avg',
        fmt(sum(norm_hra_sums['base']) / len(norm_hra_sums['base']) if norm_hra_sums['base'] else None, 1),
        fmt(sum(norm_hra_sums['avg']) / len(norm_hra_sums['avg']) if norm_hra_sums['avg'] else None, 1),
        fmt(sum(norm_hra_sums['peak']) / len(norm_hra_sums['peak']) if norm_hra_sums['peak'] else None, 1),
        fmt(sum(norm_hra_sums['rec']) / len(norm_hra_sums['rec']) if norm_hra_sums['rec'] else None, 1),
    ]
    for col, value in enumerate(norm_hra_avg_row, 1):
        cell = ws4.cell(row=row, column=col, value=value)
        cell.font = avg_font
        cell.fill = avg_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Set column widths
    for col in range(1, 12):
        ws4.column_dimensions[get_column_letter(col)].width = 11
    
    # ==================== SHEET 5: DETRENDED HRV ====================
    ws5 = wb.create_sheet("Detrended HRV")
    
    # Table 6: Light-Induced Detrended HRV Data
    ws5['A1'] = 'Table 6 | Light-Induced Detrended HRV Data'
    ws5['A1'].font = Font(bold=True, size=12)
    
    hrv_headers = ['Recording', 'ln(RMSSD₇₀)', 'ln(SDNN₇₀)', 'pNN50₇₀']
    for col, header in enumerate(hrv_headers, 1):
        cell = ws5.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = amber_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row = 4
    for rec in recordings:
        data_row = [
            extract_short_name(rec.get('name', '')),
            fmt(rec.get('light_hrv_ln_rmssd70'), 3),
            fmt(rec.get('light_hrv_ln_sdnn70'), 3),
            fmt(rec.get('light_hrv_pnn50'), 1),
        ]
        for col, value in enumerate(data_row, 1):
            cell = ws5.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if col >= 2:
                cell.fill = light_fill
        row += 1
    
    # Median row
    hrv_median_row = [
        'Median',
        fmt(hrv_averages.get('light_hrv_ln_rmssd70'), 3),
        fmt(hrv_averages.get('light_hrv_ln_sdnn70'), 3),
        fmt(hrv_averages.get('light_hrv_pnn50'), 1),
    ]
    for col, value in enumerate(hrv_median_row, 1):
        cell = ws5.cell(row=row, column=col, value=value)
        cell.font = avg_font
        cell.fill = avg_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Set column widths
    for col in range(1, 5):
        ws5.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
