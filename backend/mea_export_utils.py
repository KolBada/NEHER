"""
MEA Export Utilities - Nature Magazine Style
Clean, professional scientific publication exports
Matches SSE export design for consistency
"""
import io
import zipfile
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional

# Color palette matching NEHER UI - same as SSE
COLORS = {
    'dark': '#18181b',
    'emerald': '#10b981',      # Spikes
    'orange': '#f97316',       # Bursts
    'amber': '#f59e0b',        # Light stims
    'purple': '#a855f7',       # Drug perfusion
    'sky': '#0ea5e9',          # Baseline
    'gray': '#6b7280',
    'line': '#374151',
}

TINTS = {
    'baseline': '#E0F2FE',     # Light blue
    'drug': '#F3E8FF',         # Light purple
    'light': '#FEF3C7',        # Light amber
    'spike': '#d1fae5',        # Light emerald
    'burst': '#ffedd5',        # Light orange
}

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def safe_float(value, default=0.0):
    """Safely convert a value to float."""
    try:
        if value is None:
            return default
        return float(value)
    except (ValueError, TypeError):
        return default


def format_duration(seconds):
    """Format seconds as 'Xm Ys'."""
    if not seconds:
        return "—"
    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    return f"{minutes}m {secs}s"


def calculate_interstimuli_intervals(light_pulses):
    """Calculate intervals between consecutive light pulses.
    Returns a string like '60s-30s-20s-10s' or None if not enough pulses."""
    if not light_pulses or len(light_pulses) < 2:
        return None
    
    intervals = []
    for i in range(1, len(light_pulses)):
        # Interval from end of previous pulse to start of current pulse
        prev_end = light_pulses[i - 1].get('end_sec', 0)
        curr_start = light_pulses[i].get('start_sec', 0)
        interval = curr_start - prev_end
        intervals.append(f"{int(interval)}s")
    
    return "-".join(intervals)


# =============================================================================
# CSV EXPORT - SSE STYLE (Single file, not ZIP)
# =============================================================================

def generate_mea_csv_export(analysis_state: Dict, well_analysis: Dict) -> bytes:
    """Generate a single CSV file matching SSE export format."""
    lines = []
    
    recording_name = analysis_state.get('recordingName', 'MEA Recording')
    selected_well = analysis_state.get('selectedWell', 'N/A')
    
    # Header
    lines.append(f"{recording_name}")
    lines.append("MEA Electrophysiology Analysis Report by NEHER")
    lines.append(f"Recording Date: {analysis_state.get('recordingDate', '')}")
    lines.append("")
    
    # Recording Info
    lines.append("=== RECORDING INFO ===")
    source_files = analysis_state.get('source_files', {})
    if source_files:
        # Each file on its own line
        lines.append("Original Files:")
        for fname in source_files.values():
            lines.append(f"  ,{fname}")
    lines.append(f"Well ID,{selected_well}")
    lines.append(f"Recording Date,{analysis_state.get('recordingDate', '')}")
    
    spikes = analysis_state.get('spikes', [])
    bursts = analysis_state.get('electrode_bursts', []) or analysis_state.get('bursts', [])
    lines.append(f"Total Spikes,{len(spikes)}")
    lines.append(f"Total Bursts,{len(bursts)}")
    lines.append(f"Active Electrodes,{len(analysis_state.get('active_electrodes', []))}")
    lines.append(f"Duration,{format_duration(analysis_state.get('duration_s', 0))}")
    lines.append("")
    
    # Tissue Info
    organoid_info = analysis_state.get('organoidInfo', [])
    if organoid_info and any(o.get('cell_type') for o in organoid_info):
        lines.append("=== TISSUE INFO ===")
        for idx, org in enumerate(organoid_info):
            if org.get('cell_type'):
                lines.append(f"--- Sample {idx + 1} ---")
                cell_type = org.get('other_cell_type') if org.get('cell_type') == 'Other' else org.get('cell_type')
                lines.append(f"Cell Type,{cell_type or ''}")
                if org.get('line_name'):
                    lines.append(f"Line,{org.get('line_name')}")
                if org.get('passage_number'):
                    lines.append(f"Passage,{org.get('passage_number')}")
                if org.get('transfection') and org['transfection'].get('name'):
                    lines.append(f"Transfection,{org['transfection'].get('name')}")
        if analysis_state.get('fusionDate'):
            lines.append(f"Days Since Fusion,{analysis_state.get('fusionDate')}")
        lines.append("")
    
    # Drug Perfusion
    if analysis_state.get('drugEnabled') and analysis_state.get('selectedDrugs'):
        lines.append("=== DRUG PERFUSION ===")
        drugs = analysis_state.get('selectedDrugs', [])
        drug_settings = analysis_state.get('drugSettings', {})
        for drug in drugs:
            lines.append(f"Drug,{drug}")
            settings = drug_settings.get(drug, {})
            if settings.get('concentration'):
                lines.append(f"Concentration,{settings.get('concentration')}µM")
        perf_time = analysis_state.get('drugPerfTime', 0)
        lines.append(f"Perf. Start,{perf_time} min")
        lines.append("")
    
    # Light Stimulation
    light_pulses = analysis_state.get('lightPulses', [])
    if analysis_state.get('lightEnabled') and light_pulses:
        lines.append("=== LIGHT STIMULATION ===")
        lines.append("Status,Enabled")
        lines.append(f"Stims Detected,{len(light_pulses)}")
        if light_pulses:
            first_pulse = light_pulses[0]
            lines.append(f"Stims Start,{first_pulse.get('start_sec', 0) / 60:.2f} min")
        light_params = analysis_state.get('lightParams', {})
        if light_params.get('pulseDuration'):
            lines.append(f"Stim Duration,{light_params.get('pulseDuration')} sec")
        # Add interstimuli interval
        isi = calculate_interstimuli_intervals(light_pulses)
        if isi:
            lines.append(f"Interstimuli Interval,{isi}")
        lines.append("")
    
    # Baseline Readout
    if analysis_state.get('baselineEnabled') and well_analysis:
        lines.append("=== BASELINE READOUT ===")
        lines.append(f"Baseline Minute,{analysis_state.get('baselineMinute', 1)}")
        lines.append(f"Spike Rate,{safe_float(well_analysis.get('baselineSpikeHz'), 0):.4f} Hz")
        lines.append(f"Burst Rate,{safe_float(well_analysis.get('baselineBurstBpm'), 0):.4f} bpm")
        lines.append("")
    
    # Drug Readout
    if analysis_state.get('drugEnabled') and well_analysis:
        lines.append("=== DRUG READOUT ===")
        drugs = analysis_state.get('selectedDrugs', [])
        if drugs:
            lines.append(f"Drug,{drugs[0]}")
        perf_time = analysis_state.get('drugPerfTime', 0)
        lines.append(f"Perf. Time,{perf_time} min")
        lines.append(f"Spike Rate,{safe_float(well_analysis.get('drugSpikeHz'), 0):.4f} Hz")
        lines.append(f"Burst Rate,{safe_float(well_analysis.get('drugBurstBpm'), 0):.4f} bpm")
        lines.append("")
    
    # Light Readout
    light_metrics = analysis_state.get('lightMetrics')
    if analysis_state.get('lightEnabled') and light_metrics:
        avg = light_metrics.get('avg', {})
        lines.append("=== LIGHT READOUT ===")
        lines.append(f"Baseline Spike,{safe_float(avg.get('baselineSpikeHz'), 0):.4f} Hz")
        lines.append(f"Avg Spike,{safe_float(avg.get('avgSpikeHz'), 0):.4f} Hz")
        lines.append(f"Avg Spike (Norm.),{safe_float(avg.get('spikeChangePct'), 0):.1f}%")
        lines.append(f"Peak Spike,{safe_float(avg.get('maxSpikeHz'), 0):.4f} Hz")
        lines.append(f"Peak Spike (Norm.),{safe_float(avg.get('maxSpikeChangePct'), 0):.1f}%")
        lines.append(f"Baseline Burst,{safe_float(avg.get('baselineBurstBpm'), 0):.4f} bpm")
        lines.append(f"Avg Burst,{safe_float(avg.get('avgBurstBpm'), 0):.4f} bpm")
        lines.append(f"Avg Burst (Norm.),{safe_float(avg.get('burstChangePct'), 0):.1f}%")
        lines.append(f"Peak Burst,{safe_float(avg.get('maxBurstBpm'), 0):.4f} bpm")
        lines.append(f"Peak Burst (Norm.),{safe_float(avg.get('maxBurstChangePct'), 0):.1f}%")
        lines.append("")
    
    # Table 1: Per-Minute Spike Data
    per_minute = well_analysis.get('perMinuteCombined', []) if well_analysis else []
    if per_minute:
        lines.append("=== TABLE 1 | PER-MINUTE SPIKE FREQUENCY DATA ===")
        lines.append("Window (min),Spike Rate (Hz),Spike Count")
        for pm in per_minute:
            minute = pm.get('minute', 0)
            spike_rate = safe_float(pm.get('spike_rate_hz'), 0)
            spike_count = pm.get('spike_count', 0)
            lines.append(f"{minute}-{minute+1},{spike_rate:.4f},{spike_count}")
        lines.append("")
    
    # Table 2: Per-Minute Burst Data
    if per_minute:
        lines.append("=== TABLE 2 | PER-MINUTE BURST FREQUENCY DATA ===")
        lines.append("Window (min),Burst Rate (bpm),Burst Count")
        for pm in per_minute:
            minute = pm.get('minute', 0)
            burst_rate = safe_float(pm.get('burst_rate_bpm'), 0)
            burst_count = pm.get('burst_count', 0)
            lines.append(f"{minute}-{minute+1},{burst_rate:.4f},{burst_count}")
        lines.append("")
    
    # Table 3: Light-Induced Spike Response
    if analysis_state.get('lightEnabled') and light_metrics and light_metrics.get('perStim'):
        per_stim = light_metrics.get('perStim', [])
        lines.append("=== TABLE 3 | LIGHT-INDUCED SPIKE RESPONSE ===")
        lines.append("Stim,Baseline (Hz),Avg (Hz),Avg %,Peak (Hz),Peak %,TTP (s)")
        for i, stim in enumerate(per_stim):
            lines.append(f"{i+1},{safe_float(stim.get('baselineSpikeHz'),0):.4f},{safe_float(stim.get('avgSpikeHz'),0):.4f},{safe_float(stim.get('spikeChangePct'),0):.1f},{safe_float(stim.get('maxSpikeHz'),0):.4f},{safe_float(stim.get('maxSpikeChangePct'),0):.1f},{safe_float(stim.get('spikeTimeToPeak'),0):.1f}")
        avg = light_metrics.get('avg', {})
        lines.append(f"Avg,{safe_float(avg.get('baselineSpikeHz'),0):.4f},{safe_float(avg.get('avgSpikeHz'),0):.4f},{safe_float(avg.get('spikeChangePct'),0):.1f},{safe_float(avg.get('maxSpikeHz'),0):.4f},{safe_float(avg.get('maxSpikeChangePct'),0):.1f},{safe_float(avg.get('spikeTimeToPeak'),0):.1f}")
        lines.append("")
    
    # Table 4: Light-Induced Burst Response
    if analysis_state.get('lightEnabled') and light_metrics and light_metrics.get('perStim'):
        per_stim = light_metrics.get('perStim', [])
        lines.append("=== TABLE 4 | LIGHT-INDUCED BURST RESPONSE ===")
        lines.append("Stim,Baseline (bpm),Avg (bpm),Avg %,Peak (bpm),Peak %,TTP (s)")
        for i, stim in enumerate(per_stim):
            lines.append(f"{i+1},{safe_float(stim.get('baselineBurstBpm'),0):.4f},{safe_float(stim.get('avgBurstBpm'),0):.4f},{safe_float(stim.get('burstChangePct'),0):.1f},{safe_float(stim.get('maxBurstBpm'),0):.4f},{safe_float(stim.get('maxBurstChangePct'),0):.1f},{safe_float(stim.get('burstTimeToPeak'),0):.1f}")
        avg = light_metrics.get('avg', {})
        lines.append(f"Avg,{safe_float(avg.get('baselineBurstBpm'),0):.4f},{safe_float(avg.get('avgBurstBpm'),0):.4f},{safe_float(avg.get('burstChangePct'),0):.1f},{safe_float(avg.get('maxBurstBpm'),0):.4f},{safe_float(avg.get('maxBurstChangePct'),0):.1f},{safe_float(avg.get('burstTimeToPeak'),0):.1f}")
        lines.append("")
    
    csv_content = '\n'.join(lines)
    return csv_content.encode('utf-8')


# =============================================================================
# EXCEL EXPORT - SSE STYLE
# =============================================================================

def generate_mea_xlsx_export(analysis_state: Dict, well_analysis: Dict) -> bytes:
    """Generate an Excel workbook matching SSE export format."""
    
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    wb = Workbook()
    
    # Define styles matching SSE
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill_emerald = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_fill_orange = PatternFill(start_color="f97316", end_color="f97316", fill_type="solid")
    header_fill_amber = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
    header_fill_gray = PatternFill(start_color="6b7280", end_color="6b7280", fill_type="solid")
    baseline_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    drug_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    light_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    avg_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    alt_row_spike = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    alt_row_burst = PatternFill(start_color="ffedd5", end_color="ffedd5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    recording_name = analysis_state.get('recordingName', 'MEA Recording')
    selected_well = analysis_state.get('selectedWell', 'N/A')
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    row = 1
    # Title
    ws_summary.cell(row=row, column=1, value=f"{recording_name}").font = Font(bold=True, size=16)
    ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws_summary.cell(row=row, column=1, value="MEA Electrophysiology Analysis Report by NEHER").font = Font(size=10, color='6b7280')
    row += 1
    ws_summary.cell(row=row, column=1, value=f"Recording Date: {analysis_state.get('recordingDate', '')}").font = Font(size=10)
    row += 2
    
    # Recording Info
    ws_summary.cell(row=row, column=1, value="RECORDING INFO").font = Font(bold=True, color='FFFFFF')
    ws_summary.cell(row=row, column=1).fill = header_fill_gray
    ws_summary.cell(row=row, column=2).fill = header_fill_gray
    row += 1
    
    source_files = analysis_state.get('source_files', {})
    if source_files:
        ws_summary.cell(row=row, column=1, value="Original Files")
        # Each file on its own line in column 2
        file_list = list(source_files.values())
        ws_summary.cell(row=row, column=2, value=file_list[0] if file_list else "")
        row += 1
        for fname in file_list[1:]:
            ws_summary.cell(row=row, column=2, value=fname)
            row += 1
    
    ws_summary.cell(row=row, column=1, value="Well ID")
    ws_summary.cell(row=row, column=2, value=selected_well)
    row += 1
    
    spikes = analysis_state.get('spikes', [])
    bursts = analysis_state.get('electrode_bursts', []) or analysis_state.get('bursts', [])
    
    ws_summary.cell(row=row, column=1, value="Total Spikes")
    ws_summary.cell(row=row, column=2, value=len(spikes))
    row += 1
    ws_summary.cell(row=row, column=1, value="Total Bursts")
    ws_summary.cell(row=row, column=2, value=len(bursts))
    row += 1
    ws_summary.cell(row=row, column=1, value="Active Electrodes")
    ws_summary.cell(row=row, column=2, value=len(analysis_state.get('active_electrodes', [])))
    row += 1
    ws_summary.cell(row=row, column=1, value="Duration")
    ws_summary.cell(row=row, column=2, value=format_duration(analysis_state.get('duration_s', 0)))
    row += 2
    
    # Baseline Readout
    if analysis_state.get('baselineEnabled') and well_analysis:
        ws_summary.cell(row=row, column=1, value="BASELINE READOUT").font = Font(bold=True, color='FFFFFF')
        ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
        ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
        row += 1
        ws_summary.cell(row=row, column=1, value="Baseline Minute")
        ws_summary.cell(row=row, column=2, value=analysis_state.get('baselineMinute', 1))
        ws_summary.cell(row=row, column=1).fill = baseline_fill
        ws_summary.cell(row=row, column=2).fill = baseline_fill
        row += 1
        ws_summary.cell(row=row, column=1, value="Spike Rate")
        ws_summary.cell(row=row, column=2, value=f"{safe_float(well_analysis.get('baselineSpikeHz'), 0):.4f} Hz")
        ws_summary.cell(row=row, column=1).fill = baseline_fill
        ws_summary.cell(row=row, column=2).fill = baseline_fill
        row += 1
        ws_summary.cell(row=row, column=1, value="Burst Rate")
        ws_summary.cell(row=row, column=2, value=f"{safe_float(well_analysis.get('baselineBurstBpm'), 0):.4f} bpm")
        ws_summary.cell(row=row, column=1).fill = baseline_fill
        ws_summary.cell(row=row, column=2).fill = baseline_fill
        row += 2
    
    # Drug Readout
    if analysis_state.get('drugEnabled') and well_analysis and analysis_state.get('selectedDrugs'):
        ws_summary.cell(row=row, column=1, value="DRUG READOUT").font = Font(bold=True, color='FFFFFF')
        ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="a855f7", end_color="a855f7", fill_type="solid")
        ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="a855f7", end_color="a855f7", fill_type="solid")
        row += 1
        drugs = analysis_state.get('selectedDrugs', [])
        ws_summary.cell(row=row, column=1, value="Drug")
        ws_summary.cell(row=row, column=2, value=drugs[0] if drugs else "")
        ws_summary.cell(row=row, column=1).fill = drug_fill
        ws_summary.cell(row=row, column=2).fill = drug_fill
        row += 1
        perf_time = analysis_state.get('drugPerfTime', 0)
        ws_summary.cell(row=row, column=1, value="Perf. Time")
        ws_summary.cell(row=row, column=2, value=f"{perf_time} min")
        ws_summary.cell(row=row, column=1).fill = drug_fill
        ws_summary.cell(row=row, column=2).fill = drug_fill
        row += 1
        ws_summary.cell(row=row, column=1, value="Spike Rate")
        ws_summary.cell(row=row, column=2, value=f"{safe_float(well_analysis.get('drugSpikeHz'), 0):.4f} Hz")
        ws_summary.cell(row=row, column=1).fill = drug_fill
        ws_summary.cell(row=row, column=2).fill = drug_fill
        row += 1
        ws_summary.cell(row=row, column=1, value="Burst Rate")
        ws_summary.cell(row=row, column=2, value=f"{safe_float(well_analysis.get('drugBurstBpm'), 0):.4f} bpm")
        ws_summary.cell(row=row, column=1).fill = drug_fill
        ws_summary.cell(row=row, column=2).fill = drug_fill
        row += 2
    
    # Light Stimulation Info (before Light Readout)
    light_pulses = analysis_state.get('lightPulses', [])
    if analysis_state.get('lightEnabled') and light_pulses:
        ws_summary.cell(row=row, column=1, value="LIGHT STIMULATION").font = Font(bold=True, color='FFFFFF')
        ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        row += 1
        ws_summary.cell(row=row, column=1, value="Status")
        ws_summary.cell(row=row, column=2, value="Enabled")
        ws_summary.cell(row=row, column=1).fill = light_fill
        ws_summary.cell(row=row, column=2).fill = light_fill
        row += 1
        ws_summary.cell(row=row, column=1, value="Stims Detected")
        ws_summary.cell(row=row, column=2, value=len(light_pulses))
        ws_summary.cell(row=row, column=1).fill = light_fill
        ws_summary.cell(row=row, column=2).fill = light_fill
        row += 1
        if light_pulses:
            first_pulse = light_pulses[0]
            ws_summary.cell(row=row, column=1, value="Stims Start")
            ws_summary.cell(row=row, column=2, value=f"{first_pulse.get('start_sec', 0) / 60:.2f} min")
            ws_summary.cell(row=row, column=1).fill = light_fill
            ws_summary.cell(row=row, column=2).fill = light_fill
            row += 1
        light_params = analysis_state.get('lightParams', {})
        if light_params.get('pulseDuration'):
            ws_summary.cell(row=row, column=1, value="Stim Duration")
            ws_summary.cell(row=row, column=2, value=f"{light_params.get('pulseDuration')} sec")
            ws_summary.cell(row=row, column=1).fill = light_fill
            ws_summary.cell(row=row, column=2).fill = light_fill
            row += 1
        # Add interstimuli interval
        isi = calculate_interstimuli_intervals(light_pulses)
        if isi:
            ws_summary.cell(row=row, column=1, value="Interstimuli Interval")
            ws_summary.cell(row=row, column=2, value=isi)
            ws_summary.cell(row=row, column=1).fill = light_fill
            ws_summary.cell(row=row, column=2).fill = light_fill
            row += 1
        row += 1
    
    # Light Readout
    light_metrics = analysis_state.get('lightMetrics')
    if analysis_state.get('lightEnabled') and light_metrics:
        avg = light_metrics.get('avg', {})
        ws_summary.cell(row=row, column=1, value="LIGHT STIMULUS READOUT").font = Font(bold=True, color='FFFFFF')
        ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        ws_summary.cell(row=row, column=2).fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        row += 1
        # Baseline metrics in cyan text
        cyan_font = Font(color='22d3ee')
        ws_summary.cell(row=row, column=1, value="Baseline Spike").font = cyan_font
        ws_summary.cell(row=row, column=2, value=f"{safe_float(avg.get('baselineSpikeHz'), 0):.4f} Hz").font = cyan_font
        ws_summary.cell(row=row, column=1).fill = light_fill
        ws_summary.cell(row=row, column=2).fill = light_fill
        row += 1
        # Regular metrics
        metrics = [
            ("Avg Spike", f"{safe_float(avg.get('avgSpikeHz'), 0):.4f} Hz"),
            ("Avg Spike (Norm.)", f"{safe_float(avg.get('spikeChangePct'), 0):.1f}%"),
            ("Peak Spike", f"{safe_float(avg.get('maxSpikeHz'), 0):.4f} Hz"),
            ("Peak Spike (Norm.)", f"{safe_float(avg.get('maxSpikeChangePct'), 0):.1f}%"),
        ]
        for label, value in metrics:
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
            ws_summary.cell(row=row, column=1).fill = light_fill
            ws_summary.cell(row=row, column=2).fill = light_fill
            row += 1
        # Baseline Burst in cyan
        ws_summary.cell(row=row, column=1, value="Baseline Burst").font = cyan_font
        ws_summary.cell(row=row, column=2, value=f"{safe_float(avg.get('baselineBurstBpm'), 0):.4f} bpm").font = cyan_font
        ws_summary.cell(row=row, column=1).fill = light_fill
        ws_summary.cell(row=row, column=2).fill = light_fill
        row += 1
        # More regular metrics
        metrics = [
            ("Avg Burst", f"{safe_float(avg.get('avgBurstBpm'), 0):.4f} bpm"),
            ("Avg Burst (Norm.)", f"{safe_float(avg.get('burstChangePct'), 0):.1f}%"),
            ("Peak Burst", f"{safe_float(avg.get('maxBurstBpm'), 0):.4f} bpm"),
            ("Peak Burst (Norm.)", f"{safe_float(avg.get('maxBurstChangePct'), 0):.1f}%"),
        ]
        for label, value in metrics:
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
            ws_summary.cell(row=row, column=1).fill = light_fill
            ws_summary.cell(row=row, column=2).fill = light_fill
            row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 50
    
    # Sheet 2: Per-Minute Spike Data
    per_minute = well_analysis.get('perMinuteCombined', []) if well_analysis else []
    if per_minute:
        ws_spike = wb.create_sheet("Per-Minute Spike")
        headers = ["Window (min)", "Spike Rate (Hz)", "Spike Count"]
        for col, header in enumerate(headers, 1):
            cell = ws_spike.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill_emerald
            cell.border = thin_border
        
        baseline_minute = analysis_state.get('baselineMinute', 1)
        drug_minute = None
        if analysis_state.get('drugEnabled'):
            drug_minute = analysis_state.get('drugPerfTime', 0) + analysis_state.get('drugReadoutMinute', 0)
        
        for row_num, pm in enumerate(per_minute, 2):
            minute = pm.get('minute', 0)
            ws_spike.cell(row=row_num, column=1, value=f"{minute}-{minute+1}").border = thin_border
            ws_spike.cell(row=row_num, column=2, value=round(safe_float(pm.get('spike_rate_hz'), 0), 4)).border = thin_border
            ws_spike.cell(row=row_num, column=3, value=pm.get('spike_count', 0)).border = thin_border
            
            # Highlight rows
            if minute == baseline_minute:
                for col in range(1, 4):
                    ws_spike.cell(row=row_num, column=col).fill = baseline_fill
            elif drug_minute and minute == drug_minute:
                for col in range(1, 4):
                    ws_spike.cell(row=row_num, column=col).fill = drug_fill
            elif row_num % 2 == 0:
                for col in range(1, 4):
                    ws_spike.cell(row=row_num, column=col).fill = alt_row_spike
        
        for col in range(1, 4):
            ws_spike.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 3: Per-Minute Burst Data
    if per_minute:
        ws_burst = wb.create_sheet("Per-Minute Burst")
        headers = ["Window (min)", "Burst Rate (bpm)", "Burst Count"]
        for col, header in enumerate(headers, 1):
            cell = ws_burst.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill_orange
            cell.border = thin_border
        
        baseline_minute = analysis_state.get('baselineMinute', 1)
        drug_minute = None
        if analysis_state.get('drugEnabled'):
            drug_minute = analysis_state.get('drugPerfTime', 0) + analysis_state.get('drugReadoutMinute', 0)
        
        for row_num, pm in enumerate(per_minute, 2):
            minute = pm.get('minute', 0)
            ws_burst.cell(row=row_num, column=1, value=f"{minute}-{minute+1}").border = thin_border
            ws_burst.cell(row=row_num, column=2, value=round(safe_float(pm.get('burst_rate_bpm'), 0), 4)).border = thin_border
            ws_burst.cell(row=row_num, column=3, value=pm.get('burst_count', 0)).border = thin_border
            
            if minute == baseline_minute:
                for col in range(1, 4):
                    ws_burst.cell(row=row_num, column=col).fill = baseline_fill
            elif drug_minute and minute == drug_minute:
                for col in range(1, 4):
                    ws_burst.cell(row=row_num, column=col).fill = drug_fill
            elif row_num % 2 == 0:
                for col in range(1, 4):
                    ws_burst.cell(row=row_num, column=col).fill = alt_row_burst
        
        for col in range(1, 4):
            ws_burst.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 4: Light Spike Response
    if analysis_state.get('lightEnabled') and light_metrics and light_metrics.get('perStim'):
        ws_light_spike = wb.create_sheet("Light Spike")
        headers = ["Stim", "Baseline (Hz)", "Avg (Hz)", "Avg %", "Peak (Hz)", "Peak %", "TTP (s)"]
        for col, header in enumerate(headers, 1):
            cell = ws_light_spike.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill_amber if col > 1 else header_fill_gray
            cell.border = thin_border
            if col == 2:
                cell.fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
        
        per_stim = light_metrics.get('perStim', [])
        for row_num, stim in enumerate(per_stim, 2):
            ws_light_spike.cell(row=row_num, column=1, value=f"Stim {row_num-1}").border = thin_border
            ws_light_spike.cell(row=row_num, column=2, value=round(safe_float(stim.get('baselineSpikeHz'), 0), 4)).border = thin_border
            ws_light_spike.cell(row=row_num, column=2).fill = baseline_fill
            ws_light_spike.cell(row=row_num, column=3, value=round(safe_float(stim.get('avgSpikeHz'), 0), 4)).border = thin_border
            ws_light_spike.cell(row=row_num, column=4, value=round(safe_float(stim.get('spikeChangePct'), 0), 1)).border = thin_border
            ws_light_spike.cell(row=row_num, column=5, value=round(safe_float(stim.get('maxSpikeHz'), 0), 4)).border = thin_border
            ws_light_spike.cell(row=row_num, column=6, value=round(safe_float(stim.get('maxSpikeChangePct'), 0), 1)).border = thin_border
            ws_light_spike.cell(row=row_num, column=7, value=round(safe_float(stim.get('spikeTimeToPeak'), 0), 1)).border = thin_border
        
        # Average row
        avg = light_metrics.get('avg', {})
        avg_row = len(per_stim) + 2
        ws_light_spike.cell(row=avg_row, column=1, value="Avg").border = thin_border
        ws_light_spike.cell(row=avg_row, column=2, value=round(safe_float(avg.get('baselineSpikeHz'), 0), 4)).border = thin_border
        ws_light_spike.cell(row=avg_row, column=3, value=round(safe_float(avg.get('avgSpikeHz'), 0), 4)).border = thin_border
        ws_light_spike.cell(row=avg_row, column=4, value=round(safe_float(avg.get('spikeChangePct'), 0), 1)).border = thin_border
        ws_light_spike.cell(row=avg_row, column=5, value=round(safe_float(avg.get('maxSpikeHz'), 0), 4)).border = thin_border
        ws_light_spike.cell(row=avg_row, column=6, value=round(safe_float(avg.get('maxSpikeChangePct'), 0), 1)).border = thin_border
        ws_light_spike.cell(row=avg_row, column=7, value=round(safe_float(avg.get('spikeTimeToPeak'), 0), 1)).border = thin_border
        for col in range(1, 8):
            ws_light_spike.cell(row=avg_row, column=col).fill = avg_fill
        
        for col in range(1, 8):
            ws_light_spike.column_dimensions[get_column_letter(col)].width = 14
    
    # Sheet 5: Light Burst Response
    if analysis_state.get('lightEnabled') and light_metrics and light_metrics.get('perStim'):
        ws_light_burst = wb.create_sheet("Light Burst")
        headers = ["Stim", "Baseline (bpm)", "Avg (bpm)", "Avg %", "Peak (bpm)", "Peak %", "TTP (s)"]
        for col, header in enumerate(headers, 1):
            cell = ws_light_burst.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill_amber if col > 1 else header_fill_gray
            cell.border = thin_border
            if col == 2:
                cell.fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
        
        per_stim = light_metrics.get('perStim', [])
        for row_num, stim in enumerate(per_stim, 2):
            ws_light_burst.cell(row=row_num, column=1, value=f"Stim {row_num-1}").border = thin_border
            ws_light_burst.cell(row=row_num, column=2, value=round(safe_float(stim.get('baselineBurstBpm'), 0), 4)).border = thin_border
            ws_light_burst.cell(row=row_num, column=2).fill = baseline_fill
            ws_light_burst.cell(row=row_num, column=3, value=round(safe_float(stim.get('avgBurstBpm'), 0), 4)).border = thin_border
            ws_light_burst.cell(row=row_num, column=4, value=round(safe_float(stim.get('burstChangePct'), 0), 1)).border = thin_border
            ws_light_burst.cell(row=row_num, column=5, value=round(safe_float(stim.get('maxBurstBpm'), 0), 4)).border = thin_border
            ws_light_burst.cell(row=row_num, column=6, value=round(safe_float(stim.get('maxBurstChangePct'), 0), 1)).border = thin_border
            ws_light_burst.cell(row=row_num, column=7, value=round(safe_float(stim.get('burstTimeToPeak'), 0), 1)).border = thin_border
        
        # Average row
        avg = light_metrics.get('avg', {})
        avg_row = len(per_stim) + 2
        ws_light_burst.cell(row=avg_row, column=1, value="Avg").border = thin_border
        ws_light_burst.cell(row=avg_row, column=2, value=round(safe_float(avg.get('baselineBurstBpm'), 0), 4)).border = thin_border
        ws_light_burst.cell(row=avg_row, column=3, value=round(safe_float(avg.get('avgBurstBpm'), 0), 4)).border = thin_border
        ws_light_burst.cell(row=avg_row, column=4, value=round(safe_float(avg.get('burstChangePct'), 0), 1)).border = thin_border
        ws_light_burst.cell(row=avg_row, column=5, value=round(safe_float(avg.get('maxBurstBpm'), 0), 4)).border = thin_border
        ws_light_burst.cell(row=avg_row, column=6, value=round(safe_float(avg.get('maxBurstChangePct'), 0), 1)).border = thin_border
        ws_light_burst.cell(row=avg_row, column=7, value=round(safe_float(avg.get('burstTimeToPeak'), 0), 1)).border = thin_border
        for col in range(1, 8):
            ws_light_burst.cell(row=avg_row, column=col).fill = avg_fill
        
        for col in range(1, 8):
            ws_light_burst.column_dimensions[get_column_letter(col)].width = 14
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# PDF EXPORT - SSE STYLE
# =============================================================================

def generate_mea_pdf_export(analysis_state: Dict, well_analysis: Dict) -> bytes:
    """Generate a professional PDF report matching SSE design."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    import matplotlib.patches as mpatches
    import matplotlib.font_manager as fm
    
    # Font setup (same as SSE)
    body_font = 'DejaVu Sans'
    
    plt.rcParams.update({
        'font.family': 'sans-serif',
        'font.sans-serif': [body_font, 'DejaVu Sans', 'Arial'],
        'font.size': 9,
        'axes.labelsize': 9,
        'axes.titlesize': 10,
        'axes.linewidth': 0.8,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'legend.fontsize': 8,
        'figure.titlesize': 12,
        'axes.spines.top': False,
        'axes.spines.right': False,
    })
    
    recording_name = analysis_state.get('recordingName', 'MEA Recording')
    selected_well = analysis_state.get('selectedWell', 'N/A')
    
    # Colors matching SSE
    COLORS = {
        'dark': '#18181b',
        'gray': '#6b7280',
        'emerald': '#10b981',
        'orange': '#f97316',
        'amber': '#f59e0b',
        'purple': '#a855f7',
        'sky': '#0ea5e9',
        'line': '#374151',
    }
    
    TINTS = {
        'baseline': '#E0F2FE',
        'drug': '#F3E8FF',
        'light': '#FEF3C7',
    }
    
    buf = io.BytesIO()
    
    def add_page_header(fig, section_name):
        """Add header in the SSE style: NEHER section_name"""
        fig.text(0.08, 0.96, 'NEHER', fontsize=12, fontweight='bold', color=COLORS['dark'])
        fig.text(0.16, 0.96, section_name, fontsize=12, fontweight='normal', color=COLORS['gray'])
    
    def add_page_footer(fig, page_num):
        """Add footer: p. XX | Recording Name | MEA Electrophysiology Analysis Report by NEHER"""
        fig.text(0.08, 0.025, f'p. {page_num}', fontsize=10, fontweight='bold', color=COLORS['dark'])
        fig.text(0.125, 0.025, f'|  {recording_name} ({selected_well})  |  MEA Electrophysiology Analysis Report by NEHER', 
                fontsize=10, color=COLORS['gray'])
    
    def draw_header(fig, x, y, text, color, width=0.38):
        """Draw section header with colored background"""
        fig.add_artist(mpatches.Rectangle(
            (x, y - 0.012), width, 0.024,
            facecolor=color, edgecolor='none', transform=fig.transFigure
        ))
        fig.text(x + 0.01, y, text, fontsize=9, fontweight='bold', color='white', va='center')
        return y - 0.028
    
    def draw_row(fig, x, y, label, value, bg_color=None, width=0.38, label_width=0.18):
        """Draw a data row with label and value"""
        if bg_color:
            fig.add_artist(mpatches.Rectangle(
                (x, y - 0.008), width, 0.020,
                facecolor=bg_color, edgecolor='none', transform=fig.transFigure
            ))
        fig.text(x + 0.01, y, label, fontsize=8, color=COLORS['gray'], va='center')
        fig.text(x + label_width, y, str(value), fontsize=8, fontweight='bold', color=COLORS['dark'], va='center')
        return y - 0.020
    
    with PdfPages(buf) as pdf:
        page_num = 0
        
        # ==================== PAGE 1: SUMMARY ====================
        page_num += 1
        fig1 = plt.figure(figsize=(8.5, 11))
        fig1.patch.set_facecolor('white')
        
        add_page_header(fig1, 'summary')
        
        # Title
        fig1.text(0.08, 0.90, recording_name, ha='left', va='top', fontsize=28, fontweight='bold', color=COLORS['dark'])
        fig1.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig1.transFigure))
        
        # Layout
        left_x = 0.08
        right_x = 0.52
        col_width = 0.40
        first_section_y = 0.83
        section_gap = 0.025
        
        # LEFT COLUMN - Recording Info
        y = draw_header(fig1, left_x, first_section_y, 'RECORDING INFO', COLORS['dark'], width=col_width)
        
        source_files = analysis_state.get('source_files', {})
        if source_files:
            file_list = list(source_files.values())
            y = draw_row(fig1, left_x, y, 'Original Files:', '', width=col_width)
            for fname in file_list:
                # Truncate long filenames and use smaller font
                display_name = fname if len(fname) <= 45 else fname[:42] + "..."
                fig1.text(left_x + 0.02, y + 0.006, display_name, fontsize=6, color=COLORS['gray'], va='center')
                y -= 0.016
        
        y = draw_row(fig1, left_x, y, 'Well ID:', selected_well, width=col_width)
        y = draw_row(fig1, left_x, y, 'Recording Date:', analysis_state.get('recordingDate', ''), width=col_width)
        
        spikes = analysis_state.get('spikes', [])
        bursts = analysis_state.get('electrode_bursts', []) or analysis_state.get('bursts', [])
        y = draw_row(fig1, left_x, y, 'Total Spikes:', f"{len(spikes):,}", width=col_width)
        y = draw_row(fig1, left_x, y, 'Total Bursts:', f"{len(bursts):,}", width=col_width)
        y = draw_row(fig1, left_x, y, 'Active Electrodes:', len(analysis_state.get('active_electrodes', [])), width=col_width)
        y = draw_row(fig1, left_x, y, 'Duration:', format_duration(analysis_state.get('duration_s', 0)), width=col_width)
        
        # Tissue Info
        organoid_info = analysis_state.get('organoidInfo', [])
        if organoid_info and any(o.get('cell_type') for o in organoid_info):
            y -= section_gap
            y = draw_header(fig1, left_x, y, 'TISSUE INFO', COLORS['gray'], width=col_width)
            for idx, org in enumerate(organoid_info):
                if org.get('cell_type'):
                    cell_type = org.get('other_cell_type') if org.get('cell_type') == 'Other' else org.get('cell_type')
                    label = f'Cell Type {idx + 1}:' if len(organoid_info) > 1 else 'Cell Type:'
                    y = draw_row(fig1, left_x, y, label, cell_type or '—', width=col_width)
                    if org.get('line_name'):
                        y = draw_row(fig1, left_x, y, 'Line:', org.get('line_name'), width=col_width)
        
        # Drug Perfusion
        if analysis_state.get('drugEnabled') and analysis_state.get('selectedDrugs'):
            y -= section_gap
            y = draw_header(fig1, left_x, y, 'DRUG PERFUSION', COLORS['purple'], width=col_width)
            drugs = analysis_state.get('selectedDrugs', [])
            drug_settings = analysis_state.get('drugSettings', {})
            for drug in drugs:
                y = draw_row(fig1, left_x, y, 'Drug:', drug, TINTS['drug'], width=col_width)
                settings = drug_settings.get(drug, {})
                if settings.get('concentration'):
                    y = draw_row(fig1, left_x, y, 'Concentration:', f"{settings.get('concentration')}µM", TINTS['drug'], width=col_width)
            perf_time = analysis_state.get('drugPerfTime', 0)
            y = draw_row(fig1, left_x, y, 'Perf. Start:', f"{perf_time} min", TINTS['drug'], width=col_width)
        
        # Light Stimulation
        light_pulses = analysis_state.get('lightPulses', [])
        if analysis_state.get('lightEnabled') and light_pulses:
            y -= section_gap
            y = draw_header(fig1, left_x, y, 'LIGHT STIMULATION', COLORS['amber'], width=col_width)
            y = draw_row(fig1, left_x, y, 'Status:', 'Enabled', TINTS['light'], width=col_width)
            y = draw_row(fig1, left_x, y, 'Stims Detected:', len(light_pulses), TINTS['light'], width=col_width)
            if light_pulses:
                first_pulse = light_pulses[0]
                y = draw_row(fig1, left_x, y, 'Stims Start:', f"{first_pulse.get('start_sec', 0) / 60:.2f} min", TINTS['light'], width=col_width)
            light_params = analysis_state.get('lightParams', {})
            if light_params.get('pulseDuration'):
                y = draw_row(fig1, left_x, y, 'Stim Duration:', f"{light_params.get('pulseDuration')} sec", TINTS['light'], width=col_width)
            # Add interstimuli interval
            isi = calculate_interstimuli_intervals(light_pulses)
            if isi:
                y = draw_row(fig1, left_x, y, 'Interstimuli:', isi, TINTS['light'], width=col_width)
        
        # RIGHT COLUMN - Readouts
        y_right = first_section_y
        
        # Baseline Readout
        if analysis_state.get('baselineEnabled') and well_analysis:
            y_right = draw_header(fig1, right_x, y_right, 'BASELINE READOUT', COLORS['sky'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Baseline Minute:', analysis_state.get('baselineMinute', 1), TINTS['baseline'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Spike Rate:', f"{safe_float(well_analysis.get('baselineSpikeHz'), 0):.4f} Hz", TINTS['baseline'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Burst Rate:', f"{safe_float(well_analysis.get('baselineBurstBpm'), 0):.4f} bpm", TINTS['baseline'], width=col_width)
            y_right -= section_gap
        
        # Drug Readout
        if analysis_state.get('drugEnabled') and well_analysis and analysis_state.get('selectedDrugs'):
            y_right = draw_header(fig1, right_x, y_right, 'DRUG READOUT', COLORS['purple'], width=col_width)
            drugs = analysis_state.get('selectedDrugs', [])
            y_right = draw_row(fig1, right_x, y_right, 'Drug:', drugs[0] if drugs else '', TINTS['drug'], width=col_width)
            perf_time = analysis_state.get('drugPerfTime', 0)
            y_right = draw_row(fig1, right_x, y_right, 'Perf. Time:', f"{perf_time} min", TINTS['drug'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Spike Rate:', f"{safe_float(well_analysis.get('drugSpikeHz'), 0):.4f} Hz", TINTS['drug'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Burst Rate:', f"{safe_float(well_analysis.get('drugBurstBpm'), 0):.4f} bpm", TINTS['drug'], width=col_width)
            y_right -= section_gap
        
        # Light Stimulus Readout
        light_metrics = analysis_state.get('lightMetrics')
        if analysis_state.get('lightEnabled') and light_metrics:
            avg = light_metrics.get('avg', {})
            y_right = draw_header(fig1, right_x, y_right, 'LIGHT STIMULUS READOUT', COLORS['amber'], width=col_width)
            # Baseline Spike - cyan text
            if TINTS['light']:
                fig1.add_artist(mpatches.Rectangle(
                    (right_x, y_right - 0.008), col_width, 0.020,
                    facecolor=TINTS['light'], edgecolor='none', transform=fig1.transFigure
                ))
            fig1.text(right_x + 0.01, y_right, 'Baseline Spike:', fontsize=8, color='#22d3ee', va='center')
            fig1.text(right_x + 0.18, y_right, f"{safe_float(avg.get('baselineSpikeHz'), 0):.4f} Hz", fontsize=8, fontweight='bold', color='#22d3ee', va='center')
            y_right -= 0.020
            y_right = draw_row(fig1, right_x, y_right, 'Avg Spike:', f"{safe_float(avg.get('avgSpikeHz'), 0):.4f} Hz", TINTS['light'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Avg Spike (Norm.):', f"{safe_float(avg.get('spikeChangePct'), 0):.1f}%", TINTS['light'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Peak Spike:', f"{safe_float(avg.get('maxSpikeHz'), 0):.4f} Hz", TINTS['light'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Peak Spike (Norm.):', f"{safe_float(avg.get('maxSpikeChangePct'), 0):.1f}%", TINTS['light'], width=col_width)
            # Baseline Burst - cyan text
            if TINTS['light']:
                fig1.add_artist(mpatches.Rectangle(
                    (right_x, y_right - 0.008), col_width, 0.020,
                    facecolor=TINTS['light'], edgecolor='none', transform=fig1.transFigure
                ))
            fig1.text(right_x + 0.01, y_right, 'Baseline Burst:', fontsize=8, color='#22d3ee', va='center')
            fig1.text(right_x + 0.18, y_right, f"{safe_float(avg.get('baselineBurstBpm'), 0):.4f} bpm", fontsize=8, fontweight='bold', color='#22d3ee', va='center')
            y_right -= 0.020
            y_right = draw_row(fig1, right_x, y_right, 'Avg Burst:', f"{safe_float(avg.get('avgBurstBpm'), 0):.4f} bpm", TINTS['light'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Avg Burst (Norm.):', f"{safe_float(avg.get('burstChangePct'), 0):.1f}%", TINTS['light'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Peak Burst:', f"{safe_float(avg.get('maxBurstBpm'), 0):.4f} bpm", TINTS['light'], width=col_width)
            y_right = draw_row(fig1, right_x, y_right, 'Peak Burst (Norm.):', f"{safe_float(avg.get('maxBurstChangePct'), 0):.1f}%", TINTS['light'], width=col_width)
        
        add_page_footer(fig1, page_num)
        pdf.savefig(fig1)
        plt.close(fig1)
        
        # ==================== PAGE 2: SPIKE EVOLUTION ====================
        spike_bins = well_analysis.get('spikeRateBins', []) if well_analysis else []
        if spike_bins:
            page_num += 1
            fig2 = plt.figure(figsize=(8.5, 11))
            fig2.patch.set_facecolor('white')
            
            add_page_header(fig2, 'traces')
            fig2.text(0.08, 0.90, 'Spike Evolution', ha='left', va='top', fontsize=28, fontweight='bold', color=COLORS['dark'])
            fig2.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig2.transFigure))
            
            times = [b.get('time', 0) / 60 for b in spike_bins]
            rates = [b.get('spike_rate_hz', 0) for b in spike_bins]
            time_max = max(times) if times else 10
            baseline_spike = safe_float(well_analysis.get('baselineSpikeHz'), 0) if well_analysis else 0
            
            # Top chart: Raw spike rate
            ax1 = fig2.add_axes([0.15, 0.53, 0.75, 0.30])
            ax1.plot(times, rates, color=COLORS['emerald'], linewidth=1.5)
            ax1.fill_between(times, rates, alpha=0.3, color=COLORS['emerald'])
            ax1.set_ylabel('Spike Rate (Hz)', fontsize=9)
            ax1.set_xlabel('Time (min)', fontsize=9)
            ax1.set_title('Spike Trace Evolution', fontsize=10, fontweight='bold', pad=6)
            ax1.set_xlim(0, time_max * 1.05)
            if baseline_spike > 0:
                ax1.axhline(y=baseline_spike, color=COLORS['sky'], linestyle='--', linewidth=1, label=f'Baseline: {baseline_spike:.2f} Hz')
                ax1.legend(loc='upper right', fontsize=8)
            ax1.grid(True, alpha=0.3)
            
            # Add light stim regions
            light_pulses = analysis_state.get('lightPulses', [])
            if analysis_state.get('lightEnabled') and light_pulses:
                for pulse in light_pulses:
                    start_min = pulse.get('start_sec', 0) / 60
                    end_min = pulse.get('end_sec', 0) / 60
                    ax1.axvspan(start_min, end_min, alpha=0.2, color=COLORS['amber'])
            
            # Bottom chart: Normalized
            if baseline_spike > 0:
                ax2 = fig2.add_axes([0.15, 0.16, 0.75, 0.30])
                normalized = [(r / baseline_spike * 100) for r in rates]
                ax2.plot(times, normalized, color=COLORS['emerald'], linewidth=1.5)
                ax2.fill_between(times, normalized, 100, alpha=0.3, color=COLORS['emerald'])
                ax2.axhline(y=100, color=COLORS['sky'], linestyle='--', linewidth=1, label='Baseline (100%)')
                ax2.set_ylabel('Normalized Spike Rate (%)', fontsize=9)
                ax2.set_xlabel('Time (min)', fontsize=9)
                ax2.set_title('Spike Rate Normalized to Baseline', fontsize=10, fontweight='bold', pad=8)
                ax2.set_xlim(0, time_max * 1.05)
                ax2.set_ylim(0, 200)
                ax2.legend(loc='upper right', fontsize=8)
                ax2.grid(True, alpha=0.3)
                
                if analysis_state.get('lightEnabled') and light_pulses:
                    for pulse in light_pulses:
                        start_min = pulse.get('start_sec', 0) / 60
                        end_min = pulse.get('end_sec', 0) / 60
                        ax2.axvspan(start_min, end_min, alpha=0.2, color=COLORS['amber'])
            
            add_page_footer(fig2, page_num)
            pdf.savefig(fig2)
            plt.close(fig2)
        
        # ==================== PAGE 3: BURST EVOLUTION ====================
        burst_bins = well_analysis.get('burstRateBins', []) if well_analysis else []
        if burst_bins:
            page_num += 1
            fig3 = plt.figure(figsize=(8.5, 11))
            fig3.patch.set_facecolor('white')
            
            add_page_header(fig3, 'traces')
            fig3.text(0.08, 0.90, 'Burst Evolution', ha='left', va='top', fontsize=28, fontweight='bold', color=COLORS['dark'])
            fig3.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig3.transFigure))
            
            times = [b.get('time', 0) / 60 for b in burst_bins]
            rates = [b.get('burst_rate_bpm', 0) for b in burst_bins]
            time_max = max(times) if times else 10
            baseline_burst = safe_float(well_analysis.get('baselineBurstBpm'), 0) if well_analysis else 0
            
            # Top chart: Raw burst rate
            ax1 = fig3.add_axes([0.15, 0.53, 0.75, 0.30])
            ax1.plot(times, rates, color=COLORS['orange'], linewidth=1.5)
            ax1.fill_between(times, rates, alpha=0.3, color=COLORS['orange'])
            ax1.set_ylabel('Burst Rate (bpm)', fontsize=9)
            ax1.set_xlabel('Time (min)', fontsize=9)
            ax1.set_title('Burst Trace Evolution', fontsize=10, fontweight='bold', pad=6)
            ax1.set_xlim(0, time_max * 1.05)
            if baseline_burst > 0:
                ax1.axhline(y=baseline_burst, color=COLORS['sky'], linestyle='--', linewidth=1, label=f'Baseline: {baseline_burst:.2f} bpm')
                ax1.legend(loc='upper right', fontsize=8)
            ax1.grid(True, alpha=0.3)
            
            # Add light stim regions
            light_pulses = analysis_state.get('lightPulses', [])
            if analysis_state.get('lightEnabled') and light_pulses:
                for pulse in light_pulses:
                    start_min = pulse.get('start_sec', 0) / 60
                    end_min = pulse.get('end_sec', 0) / 60
                    ax1.axvspan(start_min, end_min, alpha=0.2, color=COLORS['amber'])
            
            # Bottom chart: Normalized
            if baseline_burst > 0:
                ax2 = fig3.add_axes([0.15, 0.16, 0.75, 0.30])
                normalized = [(r / baseline_burst * 100) for r in rates]
                ax2.plot(times, normalized, color=COLORS['orange'], linewidth=1.5)
                ax2.fill_between(times, normalized, 100, alpha=0.3, color=COLORS['orange'])
                ax2.axhline(y=100, color=COLORS['sky'], linestyle='--', linewidth=1, label='Baseline (100%)')
                ax2.set_ylabel('Normalized Burst Rate (%)', fontsize=9)
                ax2.set_xlabel('Time (min)', fontsize=9)
                ax2.set_title('Burst Rate Normalized to Baseline', fontsize=10, fontweight='bold', pad=8)
                ax2.set_xlim(0, time_max * 1.05)
                ax2.set_ylim(0, 200)
                ax2.legend(loc='upper right', fontsize=8)
                ax2.grid(True, alpha=0.3)
                
                if analysis_state.get('lightEnabled') and light_pulses:
                    for pulse in light_pulses:
                        start_min = pulse.get('start_sec', 0) / 60
                        end_min = pulse.get('end_sec', 0) / 60
                        ax2.axvspan(start_min, end_min, alpha=0.2, color=COLORS['amber'])
            
            add_page_footer(fig3, page_num)
            pdf.savefig(fig3)
            plt.close(fig3)
        
        # ==================== PAGE 4: SPIKE FREQUENCY TABLE ====================
        per_minute = well_analysis.get('perMinuteCombined', []) if well_analysis else []
        if per_minute:
            page_num += 1
            fig4 = plt.figure(figsize=(8.5, 11))
            fig4.patch.set_facecolor('white')
            
            add_page_header(fig4, 'spontaneous activity')
            fig4.text(0.08, 0.90, 'Spike Frequency', ha='left', va='top', fontsize=28, fontweight='bold', color=COLORS['dark'])
            fig4.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig4.transFigure))
            
            fig4.text(0.08, 0.84, 'Table 1 | Per-Minute Spike Frequency Data', fontsize=11, fontweight='bold', color=COLORS['dark'])
            fig4.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig4.transFigure))
            
            ax = fig4.add_axes([0.08, 0.10, 0.84, 0.72])
            ax.axis('off')
            
            headers = ['Window (min)', 'Spike Rate (Hz)', 'Spike Count']
            data = [headers]
            
            baseline_minute = analysis_state.get('baselineMinute', 1)
            drug_minute = None
            if analysis_state.get('drugEnabled'):
                drug_minute = analysis_state.get('drugPerfTime', 0) + analysis_state.get('drugReadoutMinute', 0)
            
            row_colors = []
            for pm in per_minute:
                minute = pm.get('minute', 0)
                spike_rate = safe_float(pm.get('spike_rate_hz'), 0)
                spike_count = pm.get('spike_count', 0)
                data.append([f"{minute}-{minute+1}", f"{spike_rate:.4f}", spike_count])
                
                if minute == baseline_minute:
                    row_colors.append(TINTS['baseline'])
                elif drug_minute and minute == drug_minute:
                    row_colors.append(TINTS['drug'])
                else:
                    row_colors.append(None)
            
            table = ax.table(cellText=data, loc='upper center', cellLoc='center', colWidths=[0.25, 0.25, 0.2])
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.0, 1.8)
            
            for (row, col), cell in table.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white')
                    cell.set_facecolor(COLORS['emerald'])
                elif row > 0 and row <= len(row_colors) and row_colors[row-1]:
                    cell.set_facecolor(row_colors[row-1])
                    cell.set_text_props(fontweight='bold')
                else:
                    cell.set_facecolor('#d1fae5' if row % 2 == 0 else 'white')
            
            add_page_footer(fig4, page_num)
            pdf.savefig(fig4)
            plt.close(fig4)
        
        # ==================== PAGE 5: BURST FREQUENCY TABLE ====================
        if per_minute:
            page_num += 1
            fig5 = plt.figure(figsize=(8.5, 11))
            fig5.patch.set_facecolor('white')
            
            add_page_header(fig5, 'spontaneous activity')
            fig5.text(0.08, 0.90, 'Burst Frequency', ha='left', va='top', fontsize=28, fontweight='bold', color=COLORS['dark'])
            fig5.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig5.transFigure))
            
            fig5.text(0.08, 0.84, 'Table 2 | Per-Minute Burst Frequency Data', fontsize=11, fontweight='bold', color=COLORS['dark'])
            fig5.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig5.transFigure))
            
            ax = fig5.add_axes([0.08, 0.10, 0.84, 0.72])
            ax.axis('off')
            
            headers = ['Window (min)', 'Burst Rate (bpm)', 'Burst Count']
            data = [headers]
            
            row_colors = []
            for pm in per_minute:
                minute = pm.get('minute', 0)
                burst_rate = safe_float(pm.get('burst_rate_bpm'), 0)
                burst_count = pm.get('burst_count', 0)
                data.append([f"{minute}-{minute+1}", f"{burst_rate:.4f}", burst_count])
                
                if minute == baseline_minute:
                    row_colors.append(TINTS['baseline'])
                elif drug_minute and minute == drug_minute:
                    row_colors.append(TINTS['drug'])
                else:
                    row_colors.append(None)
            
            table = ax.table(cellText=data, loc='upper center', cellLoc='center', colWidths=[0.25, 0.25, 0.2])
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.0, 1.8)
            
            for (row, col), cell in table.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white')
                    cell.set_facecolor(COLORS['orange'])
                elif row > 0 and row <= len(row_colors) and row_colors[row-1]:
                    cell.set_facecolor(row_colors[row-1])
                    cell.set_text_props(fontweight='bold')
                else:
                    cell.set_facecolor('#ffedd5' if row % 2 == 0 else 'white')
            
            add_page_footer(fig5, page_num)
            pdf.savefig(fig5)
            plt.close(fig5)
        
        # ==================== PAGE 6: LIGHT SPIKE TABLE ====================
        light_metrics = analysis_state.get('lightMetrics')
        if analysis_state.get('lightEnabled') and light_metrics and light_metrics.get('perStim'):
            page_num += 1
            fig6 = plt.figure(figsize=(8.5, 11))
            fig6.patch.set_facecolor('white')
            
            add_page_header(fig6, 'light stimulus')
            fig6.text(0.08, 0.90, 'Light-Induced Spike Response', ha='left', va='top', fontsize=28, fontweight='bold', color=COLORS['dark'])
            fig6.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig6.transFigure))
            
            fig6.text(0.08, 0.84, 'Table 3 | Per-Stimulus Spike Data', fontsize=11, fontweight='bold', color=COLORS['dark'])
            fig6.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig6.transFigure))
            
            ax = fig6.add_axes([0.08, 0.10, 0.84, 0.72])
            ax.axis('off')
            
            headers = ['Stim', 'Baseline (Hz)', 'Avg (Hz)', 'Avg %', 'Peak (Hz)', 'Peak %', 'TTP (s)']
            data = [headers]
            
            per_stim = light_metrics.get('perStim', [])
            for i, stim in enumerate(per_stim):
                data.append([
                    f"Stim {i+1}",
                    f"{safe_float(stim.get('baselineSpikeHz'), 0):.2f}",
                    f"{safe_float(stim.get('avgSpikeHz'), 0):.2f}",
                    f"{safe_float(stim.get('spikeChangePct'), 0):.1f}",
                    f"{safe_float(stim.get('maxSpikeHz'), 0):.2f}",
                    f"{safe_float(stim.get('maxSpikeChangePct'), 0):.1f}",
                    f"{safe_float(stim.get('spikeTimeToPeak'), 0):.1f}"
                ])
            
            # Average row
            avg = light_metrics.get('avg', {})
            data.append([
                'Avg',
                f"{safe_float(avg.get('baselineSpikeHz'), 0):.2f}",
                f"{safe_float(avg.get('avgSpikeHz'), 0):.2f}",
                f"{safe_float(avg.get('spikeChangePct'), 0):.1f}",
                f"{safe_float(avg.get('maxSpikeHz'), 0):.2f}",
                f"{safe_float(avg.get('maxSpikeChangePct'), 0):.1f}",
                f"{safe_float(avg.get('spikeTimeToPeak'), 0):.1f}"
            ])
            
            table = ax.table(cellText=data, loc='upper center', cellLoc='center', 
                           colWidths=[0.12, 0.14, 0.12, 0.1, 0.12, 0.1, 0.1])
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0, 1.8)
            
            for (row, col), cell in table.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white')
                    if col == 0:
                        cell.set_facecolor('#6b7280')
                    elif col == 1:
                        cell.set_facecolor(COLORS['sky'])
                    else:
                        cell.set_facecolor(COLORS['amber'])
                elif row == len(data) - 1:  # Average row
                    cell.set_facecolor('#FFEBEE')
                    cell.set_text_props(fontweight='bold')
                elif col == 1:
                    cell.set_facecolor(TINTS['baseline'])
            
            add_page_footer(fig6, page_num)
            pdf.savefig(fig6)
            plt.close(fig6)
        
        # ==================== PAGE 7: LIGHT BURST TABLE ====================
        if analysis_state.get('lightEnabled') and light_metrics and light_metrics.get('perStim'):
            page_num += 1
            fig7 = plt.figure(figsize=(8.5, 11))
            fig7.patch.set_facecolor('white')
            
            add_page_header(fig7, 'light stimulus')
            fig7.text(0.08, 0.90, 'Light-Induced Burst Response', ha='left', va='top', fontsize=28, fontweight='bold', color=COLORS['dark'])
            fig7.add_artist(plt.Line2D([0.08, 0.92], [0.865, 0.865], color=COLORS['dark'], linewidth=1.0, transform=fig7.transFigure))
            
            fig7.text(0.08, 0.84, 'Table 4 | Per-Stimulus Burst Data', fontsize=11, fontweight='bold', color=COLORS['dark'])
            fig7.add_artist(plt.Line2D([0.08, 0.92], [0.825, 0.825], color=COLORS['line'], linewidth=0.5, transform=fig7.transFigure))
            
            ax = fig7.add_axes([0.08, 0.10, 0.84, 0.72])
            ax.axis('off')
            
            headers = ['Stim', 'Baseline (bpm)', 'Avg (bpm)', 'Avg %', 'Peak (bpm)', 'Peak %', 'TTP (s)']
            data = [headers]
            
            per_stim = light_metrics.get('perStim', [])
            for i, stim in enumerate(per_stim):
                data.append([
                    f"Stim {i+1}",
                    f"{safe_float(stim.get('baselineBurstBpm'), 0):.2f}",
                    f"{safe_float(stim.get('avgBurstBpm'), 0):.2f}",
                    f"{safe_float(stim.get('burstChangePct'), 0):.1f}",
                    f"{safe_float(stim.get('maxBurstBpm'), 0):.2f}",
                    f"{safe_float(stim.get('maxBurstChangePct'), 0):.1f}",
                    f"{safe_float(stim.get('burstTimeToPeak'), 0):.1f}"
                ])
            
            # Average row
            avg = light_metrics.get('avg', {})
            data.append([
                'Avg',
                f"{safe_float(avg.get('baselineBurstBpm'), 0):.2f}",
                f"{safe_float(avg.get('avgBurstBpm'), 0):.2f}",
                f"{safe_float(avg.get('burstChangePct'), 0):.1f}",
                f"{safe_float(avg.get('maxBurstBpm'), 0):.2f}",
                f"{safe_float(avg.get('maxBurstChangePct'), 0):.1f}",
                f"{safe_float(avg.get('burstTimeToPeak'), 0):.1f}"
            ])
            
            table = ax.table(cellText=data, loc='upper center', cellLoc='center',
                           colWidths=[0.12, 0.14, 0.12, 0.1, 0.14, 0.1, 0.1])
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0, 1.8)
            
            for (row, col), cell in table.get_celld().items():
                cell.set_edgecolor('#e5e7eb')
                if row == 0:
                    cell.set_text_props(fontweight='bold', color='white')
                    if col == 0:
                        cell.set_facecolor('#6b7280')
                    elif col == 1:
                        cell.set_facecolor(COLORS['sky'])
                    else:
                        cell.set_facecolor(COLORS['amber'])
                elif row == len(data) - 1:  # Average row
                    cell.set_facecolor('#FFEBEE')
                    cell.set_text_props(fontweight='bold')
                elif col == 1:
                    cell.set_facecolor(TINTS['baseline'])
            
            add_page_footer(fig7, page_num)
            pdf.savefig(fig7)
            plt.close(fig7)
    
    buf.seek(0)
    return buf.getvalue()
